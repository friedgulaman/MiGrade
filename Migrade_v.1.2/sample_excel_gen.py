import openpyxl

# Use the absolute path to the Excel file
file_path = r'C:\Users\angelo\Documents\GitHub\ces_migrade\MiGrade\Migrade_v.1.2\TEMPLATE - SF1.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Print all sheet names
sheet_names = workbook.sheetnames
print("Sheet Names:", sheet_names)

# Select the desired sheet (use the correct sheet name from the output)
desired_sheet_name = 'Sheet1'  # Replace 'Sheet1' with the correct sheet name
sheet = workbook[desired_sheet_name]

# The value you want to write to cell CDEF7
value_to_write = "maam gina"

# Specify the column and row coordinates (CDEF7 in this case)
column_coordinates = 2  # Column index for 'CDEF' (0-based index)
row_coordinates = 12  # Row index for '7' (0-based index)

# Write the value to the specified cell
sheet.cell(row=row_coordinates, column=column_coordinates, value=value_to_write)

# Save the changes to the workbook
workbook.save(file_path)

print(f"Updated cell ({row_coordinates}, {column_coordinates}) with the value: {value_to_write}")
