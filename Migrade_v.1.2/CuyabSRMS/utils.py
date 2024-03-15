from .models import ActivityLog
from django.http import HttpResponse
import openpyxl
import pandas as pd
import shutil
from datetime import datetime
from openpyxl.styles import Font
from .models import GradeScores
import json


def transmuted_grade(initial_grade):
                    if initial_grade is None:
                        return None
                    
                    if 98.40 <= initial_grade <= 99.99:
                        return 99
                    elif 96.80 <= initial_grade <= 98.39:
                        return 98
                    elif 95.20 <= initial_grade <= 96.79:
                        return 97
                    elif 93.60 <= initial_grade <= 95.19:
                        return 96
                    elif 92.00 <= initial_grade <= 93.59:
                        return 95
                    elif 90.40 <= initial_grade <= 91.99:
                        return 94
                    elif 88.80 <= initial_grade <= 90.39:
                        return 93
                    elif 87.20 <= initial_grade <= 88.79:
                        return 92
                    elif 85.60 <= initial_grade <= 87.19:
                        return 91
                    elif 84.00 <= initial_grade <= 85.59:
                        return 90
                    elif 82.40 <= initial_grade <= 83.99:
                        return 89
                    elif 80.80 <= initial_grade <= 82.39:
                        return 88
                    elif 79.20 <= initial_grade <= 80.79:
                        return 87
                    elif 77.60 <= initial_grade <= 79.19:
                        return 86
                    elif 76.00 <= initial_grade <= 77.59:
                        return 85
                    elif 74.40 <= initial_grade <= 75.99:
                        return 84
                    elif 72.80 <= initial_grade <= 74.39:
                        return 83
                    elif 71.20 <= initial_grade <= 72.79:
                        return 82
                    elif 69.60 <= initial_grade <= 71.19:
                        return 81
                    elif 68.00 <= initial_grade <= 69.59:
                        return 80
                    elif 66.40 <= initial_grade <= 67.99:
                        return 79
                    elif 64.80 <= initial_grade <= 66.39:
                        return 78
                    elif 63.20 <= initial_grade <= 64.79:
                        return 77
                    elif 61.60 <= initial_grade <= 63.19:
                        return 76
                    elif 60.00 <= initial_grade <= 61.59:
                        return 75
                    elif 56.00 <= initial_grade <= 59.99:
                        return 74
                    elif 52.00 <= initial_grade <= 55.99:
                        return 73
                    elif 48.00 <= initial_grade <= 51.99:
                        return 72
                    elif 44.00 <= initial_grade <= 47.99:
                        return 71
                    elif 40.00 <= initial_grade <= 43.99:
                        return 70
                    elif 36.00 <= initial_grade <= 39.99:
                        return 69
                    elif 32.00 <= initial_grade <= 35.99:
                        return 68
                    elif 28.00 <= initial_grade <= 31.99:
                        return 67
                    elif 24.00 <= initial_grade <= 27.99:
                        return 66
                    elif 20.00 <= initial_grade <= 23.99:
                        return 65
                    elif 16.00 <= initial_grade <= 19.99:
                        return 64
                    elif 12.00 <= initial_grade <= 15.99:
                        return 63
                    elif 8.00 <= initial_grade <= 11.99:
                        return 62
                    elif 4.00 <= initial_grade <= 7.99:
                        return 61
                    elif 0 <= initial_grade <= 3.99:
                        return 0
                    else:
                        return initial_grade
                    
def log_activity(user, action, details=''):
    ActivityLog.objects.create(user=user, action=action, details=details)

def write_school_info(sheet, school_info, grade, section, teacher_name):
    column_coordinates_region = 7
    column_coordinates_school_name = 7
    column_coordinates_district = 24
    column_coordinates_school_year = 33
    column_coordinates_schoolID = 24
    column_coordinates_division = 15
    column_coordinates_grade_section = 11
    column_coordinates_teacher = 19
    row_coordinates_4 = 4
    row_coordinates_5 = 5
    row_coordinates_7 = 7

    for info in school_info:
        value_region = info.region
        value_school_name = info.school_name
        value_division = info.division
        value_school_id = info.school_id
        value_district = info.district

        value_school_year = info.school_year
        value_grade_section =  f"{grade} - {section}"
        value_teacher = teacher_name
        
        # Write region
        sheet.cell(row=row_coordinates_4, column=column_coordinates_region, value=value_region)

        # Write school name
        sheet.cell(row=row_coordinates_5, column=column_coordinates_school_name, value=value_school_name)
        sheet.cell(row=row_coordinates_4, column=column_coordinates_division, value=value_division)
        sheet.cell(row=row_coordinates_5, column=column_coordinates_schoolID, value=value_school_id)
        sheet.cell(row=row_coordinates_4, column=column_coordinates_district, value=value_district)
        sheet.cell(row=row_coordinates_5, column=column_coordinates_school_year, value=value_school_year)
        sheet.cell(row=row_coordinates_7, column=column_coordinates_grade_section, value=value_grade_section)
        sheet.cell(row=row_coordinates_7, column=column_coordinates_teacher, value=value_teacher)


    
# STUDENT NAME
def write_student_names(sheet, grade_scores_queryset):
    column_coordinates_student_name = 2
    row_coordinates_male = 12
    row_coordinates_female = 64

    while sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name).value:
        row_coordinates_male += 1

    while sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name).value:
        row_coordinates_female += 1

    for score_student_name in grade_scores_queryset:
        if score_student_name.student.sex == 'M':
            value_to_write = score_student_name.student.name
            sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name, value=value_to_write)
            row_coordinates_male += 1
        elif score_student_name.student.sex == 'F':
            value_to_write = score_student_name.student.name
            sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name, value=value_to_write)
            row_coordinates_female += 1


def write_scores_hps_written(sheet, grade_scores_queryset):

    # SCORES IN HIGHEST POSSIBLE SCORE IN WRITTEN WORKS
    column_coordinates_score_hps_written = 6
    row_coordinates_scores_hps_written = 10

    while sheet.cell(row=row_coordinates_scores_hps_written, column=column_coordinates_score_hps_written).value:
        column_coordinates_score_hps_written += 1

    for score_hps_written in grade_scores_queryset:
        values_to_write = score_hps_written.scores_hps_written
        for value in values_to_write:
            sheet.cell(row=row_coordinates_scores_hps_written, column=column_coordinates_score_hps_written, value=value)
            column_coordinates_score_hps_written += 1

        column_coordinates_score_hps_written = 6

    # TOTAL HIGHEST POSSIBLE SCORE IN WRITTEN WORKS
        
    column_coordinates_total_scores_hps_written = 16
    row_coordinate_total_scores_hps_written = 10

    for scores_total_ww_hps in grade_scores_queryset:
        values_to_write = scores_total_ww_hps.total_ww_hps


        value_to_write = str(values_to_write) if values_to_write is not None and values_to_write !=0 else ""

        sheet.cell(row=row_coordinate_total_scores_hps_written, column=column_coordinates_total_scores_hps_written, value=value_to_write)
        column_coordinates_total_scores_hps_written += 1

        column_coordinates_total_scores_hps_written = 16

    # PERCENTAGE SCORE IN HIGHEST POSSIBLE SCORE IN WRITTEN WORKS
    column_coordinate_total_ww_ps = 17
    row_coordinate_total_ww_ps = 10

    for scores_total_ww_ps in grade_scores_queryset:
        value_to_write = 100

        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_total_ww_ps, column=column_coordinate_total_ww_ps, value=value_to_write)
        column_coordinate_total_ww_ps += 1

        column_coordinate_total_ww_ps = 17

    # WEIGHT INPUT IN HIGHEST POSSIBLE SCORE IN WRITTEN WORKS
    column_coordinate_weight_input_written = 18
    row_coordinate_weight_input_written = 10

    for scores_weight_input_written in grade_scores_queryset:
        value_to_write = scores_weight_input_written.weight_input_written
        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_weight_input_written, column=column_coordinate_weight_input_written, value=value_to_write)
        column_coordinate_weight_input_written += 1

        column_coordinate_weight_input_written = 18

def write_scores_hps_performance(sheet, grade_scores_queryset):
    column_coordinate = 19
    row_coordinate = 10

    while sheet.cell(row=row_coordinate, column=column_coordinate). value:
        column_coordinate += 1

    for score in grade_scores_queryset:
        values_to_write = score.scores_hps_performance
        for value in values_to_write:
            sheet.cell(row=row_coordinate, column=column_coordinate, value=value)
            column_coordinate += 1

        column_coordinate = 19

    # TOTAL HIGHEST POSSIBLE SCORE IN PERFORMANCE TASKS
        
    column_coordinates_total_scores_hps_performance = 29
    row_coordinate_total_scores_hps_performance = 10

    for scores_total_pt_hps in grade_scores_queryset:
        values_to_write = scores_total_pt_hps.total_pt_hps


        value_to_write = str(values_to_write) if values_to_write is not None and values_to_write !=0 else ""

        sheet.cell(row=row_coordinate_total_scores_hps_performance, column=column_coordinates_total_scores_hps_performance, value=value_to_write)
        column_coordinates_total_scores_hps_performance += 1

        column_coordinates_total_scores_hps_performance = 29

    # PERCENTAGE SCORE IN HIGHEST POSSIBLE SCORE IN WRITTEN WORKS
    column_coordinate_total_pt_ps = 30
    row_coordinate_total_pt_ps = 10

    for scores_total_pt_ps in grade_scores_queryset:
        value_to_write = 100

        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_total_pt_ps, column=column_coordinate_total_pt_ps, value=value_to_write)
        column_coordinate_total_pt_ps += 1

        column_coordinate_total_pt_ps = 30

    # WEIGHT INPUT IN HIGHEST POSSIBLE SCORE
    column_coordinate_weight_input_performance = 31
    row_coordinate_weight_input_performance = 10

    for scores_weight_input_performance in grade_scores_queryset:
        value_to_write = scores_weight_input_performance.weight_input_performance
        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_weight_input_performance, column=column_coordinate_weight_input_performance, value=value_to_write)
        column_coordinate_weight_input_performance += 1

        column_coordinate_weight_input_performance = 31

def write_scores_hps_quarterly(sheet, grade_scores_queryset):

    # TOTAL HIGHEST POSSIBLE SCORE IN QUARTERLY ASSESSMENT
        
    column_coordinates_total_scores_hps_performance = 32
    row_coordinate_total_scores_hps_performance = 10

    for scores_total_pt_hps in grade_scores_queryset:
        values_to_write = scores_total_pt_hps.total_qa_hps


        value_to_write = str(values_to_write) if values_to_write is not None and values_to_write !=0 else ""

        sheet.cell(row=row_coordinate_total_scores_hps_performance, column=column_coordinates_total_scores_hps_performance, value=value_to_write)
        column_coordinates_total_scores_hps_performance += 1

        column_coordinates_total_scores_hps_performance = 32

    # PERCENTAGE SCORE IN HIGHEST POSSIBLE SCORE IN QUARTERLY ASSESSMENT
    column_coordinate_total_pt_ps = 33
    row_coordinate_total_pt_ps = 10

    for scores_total_pt_ps in grade_scores_queryset:
        value_to_write = 100

        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_total_pt_ps, column=column_coordinate_total_pt_ps, value=value_to_write)
        column_coordinate_total_pt_ps += 1

        column_coordinate_total_pt_ps = 33

    # WEIGHT INPUT IN HIGHEST POSSIBLE SCORE QUARTERLY ASSESSMENT
    column_coordinate_weight_input_performance = 34
    row_coordinate_weight_input_performance = 10

    for scores_weight_input_performance in grade_scores_queryset:
        value_to_write = scores_weight_input_performance.weight_input_quarterly
        value_to_write = str(value_to_write)
        sheet.cell(row=row_coordinate_weight_input_performance, column=column_coordinate_weight_input_performance, value=value_to_write)
        column_coordinate_weight_input_performance += 1

        column_coordinate_weight_input_performance = 34


def write_written_works_scores(sheet, grade_scores_queryset):
    column_coordinates = 6
    max_column_index = 15
    row_coordinates_male = 12
    row_coordinates_female = 64

    for score in grade_scores_queryset:
     

        # Get the written_works_scores_list directly from the score object
        written_works_scores_list = score.written_works_scores

        # Set the row coordinates based on the sex of the student
        if score.student.sex == 'M':

            for value in written_works_scores_list:
                value_to_write = str(value)

                sheet.cell(row=row_coordinates_male, column=column_coordinates, value=value_to_write)
                column_coordinates += 1

                # If we reach the last column, move to the next row
                if column_coordinates > max_column_index:
                    column_coordinates = 6  # Reset column index to the starting column
                    row_coordinates_male += 1  # Move to the next row

        elif score.student.sex == 'F':

            for value in written_works_scores_list:
                value_to_write = str(value)
                sheet.cell(row=row_coordinates_female, column=column_coordinates, value=value_to_write)
                column_coordinates += 1

            # If we reach the last column, move to the next row
            if column_coordinates > max_column_index:
                column_coordinates = 6  # Reset column index to the starting column
                row_coordinates_female += 1  # Move to the next row

        else:
            print(f"Unknown sex for student: {score.student.name}")
            continue




    # TOTAL WRITTEN WORKS SCORE
        column_coordinates_total_written_works_score = 16
        row_coordinates_total_written_works_score_male = 12
        row_coordinates_total_written_works_score_female = 64

        for score in grade_scores_queryset:
            total_written_works_score = score.total_score_written

            # Convert the float value to a string or use an empty string if it's None
            value_to_write = str(total_written_works_score) if total_written_works_score is not None and total_written_works_score != 0 else ""

            if score.student.sex == 'M':
                sheet.cell(row=row_coordinates_total_written_works_score_male, column=column_coordinates_total_written_works_score, value=value_to_write)

                column_coordinates_total_written_works_score += 1

                if column_coordinates_total_written_works_score > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates_total_written_works_score = 16  # Reset column index to the starting column
                    row_coordinates_total_written_works_score_male += 1

            elif score.student.sex == 'F':
                sheet.cell(row=row_coordinates_total_written_works_score_female, column=column_coordinates_total_written_works_score, value=value_to_write)

                column_coordinates_total_written_works_score += 1

                if column_coordinates_total_written_works_score > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates_total_written_works_score = 16  # Reset column index to the starting column
                    row_coordinates_total_written_works_score_female += 1
            else:
                print(f"Unknown sex for student: {score.student.name}")
                continue

        # TOTAL PERCENTAGE SCORE WRITTEN WORKS
        column_coordinates_percentage_score_written = 17
        row_coordinates_percentage_score_written_male = 12
        row_coordinates_percentage_score_written_female = 64


        for scores in grade_scores_queryset:
            percentage_score_written = scores.percentage_score_written

            if percentage_score_written is not None:
                rounded_percentage_score = round(percentage_score_written, 2)
                value_to_write = str(rounded_percentage_score)
            else:
                value_to_write = ""  # or any other placeholder for None

            if scores.student.sex == 'M':

                sheet.cell(row=row_coordinates_percentage_score_written_male, column=column_coordinates_percentage_score_written, value=value_to_write)
                column_coordinates_percentage_score_written += 1

                if column_coordinates_percentage_score_written > max_column_index:
                    column_coordinates_percentage_score_written = 17
                    row_coordinates_percentage_score_written_male += 1


            elif scores.student.sex == 'F':
                sheet.cell(row= row_coordinates_percentage_score_written_female, column=column_coordinates_percentage_score_written, value=value_to_write)
                column_coordinates_percentage_score_written += 1

                if column_coordinates_percentage_score_written > max_column_index:
                    column_coordinates_percentage_score_written = 17
                    row_coordinates_percentage_score_written_female += 1

            else:
                print(f"Unknown sex for student: {scores.student.name}")
                continue

        # TOTAL WEIGHTED SCORE WRITTEN WORKS
        column_coordinates_weighted_score_written = 18
        row_coordinates_weighted_score_written_male = 12
        row_coordinates_weighted_score_written_female = 64

        for scores in grade_scores_queryset:
            weighted_score_written = scores.weighted_score_written

            if weighted_score_written is not None:
                rounded_weighted_score_written = round(weighted_score_written, 2)
                value_to_write = str(rounded_weighted_score_written)
            else:
                value_to_write = ""  # or any other placeholder for None

            if scores.student.sex == 'M':  
                sheet.cell(row=row_coordinates_weighted_score_written_male, column=column_coordinates_weighted_score_written, value=value_to_write)
                column_coordinates_weighted_score_written += 1

                if column_coordinates_weighted_score_written > max_column_index:
                    column_coordinates_weighted_score_written = 18
                    row_coordinates_weighted_score_written_male += 1


            elif scores.student.sex == 'F':
             
                sheet.cell(row=row_coordinates_weighted_score_written_female, column=column_coordinates_weighted_score_written, value=value_to_write)
                column_coordinates_weighted_score_written += 1

                if column_coordinates_weighted_score_written > max_column_index:
                    column_coordinates_weighted_score_written = 18
                    row_coordinates_weighted_score_written_female += 1
                
            else:
                print(f"Unknown sex for student: {scores.student.name}")
                continue

def write_performance_tasks_scores(sheet, grade_scores_queryset):

    # PERFORMANCE TASKS SCORES
    column_coordinates = 19
    row_coordinates_male = 12
    row_coordinates_female = 64
    max_column_index = 28

    for score in grade_scores_queryset:
        performance_tasks_scores_list = score.performance_task_scores

        for value in performance_tasks_scores_list:
            value_to_write = str(value)

            if score.student.sex == 'M':  
                sheet.cell(row=row_coordinates_male, column=column_coordinates, value=value_to_write)
                column_coordinates += 1

                # If we reach the last column, move to the next row
                if column_coordinates > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates = 19  # Reset column index to the starting column
                    row_coordinates_male += 1  # Move to the next row

            elif score.student.sex == 'F':
                sheet.cell(row=row_coordinates_female, column=column_coordinates, value=value_to_write)
                column_coordinates += 1

                # If we reach the last column, move to the next row
                if column_coordinates > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates = 19  # Reset column index to the starting column
                    row_coordinates_female += 1  # Move to the next row
                
            else:
                print(f"Unknown sex for student: {score.student.name}")
                continue


        # TOTAL PERFORMANCE TASKS SCORE
        column_coordinates_total_performance_tasks_score = 29
        row_coordinates_total_performance_tasks_score_male = 12
        row_coordinates_total_performance_tasks_score_female = 64

        for scores in grade_scores_queryset:
            total_performance_tasks_score = scores.total_score_performance

            # Convert the float value to a string or set to empty string if None
            value_to_write = str(total_performance_tasks_score) if total_performance_tasks_score is not None and total_performance_tasks_score != 0 else ""

            if scores.student.sex == 'M':    
                sheet.cell(row=row_coordinates_total_performance_tasks_score_male, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
                column_coordinates_total_performance_tasks_score += 1

                if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates_total_performance_tasks_score = 29  # Reset column index to the starting column
                    row_coordinates_total_performance_tasks_score_male += 1  

            elif scores.student.sex == 'F':
                
                sheet.cell(row=row_coordinates_total_performance_tasks_score_female, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
                column_coordinates_total_performance_tasks_score += 1

                if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                    column_coordinates_total_performance_tasks_score = 29  # Reset column index to the starting column
                    row_coordinates_total_performance_tasks_score_female += 1  

            else:
                print(f"Unknown sex for student: {scores.student.name}")
                continue


            # TOTAL PERCENTAGE SCORE WRITTEN WORKS
            column_coordinates_percentage_score_performance = 30
            row_coordinates_percentage_score_performance_male = 12
            row_coordinates_percentage_score_performance_female = 64


            for scores in grade_scores_queryset:
                percentage_score_performance = scores.percentage_score_performance

                # Check if the value is not None before rounding
                if percentage_score_performance is not None:
                    rounded_percentage_score = round(percentage_score_performance, 2)
                    value_to_write = str(rounded_percentage_score)
                else:
                    value_to_write = ""  # or any other placeholder for None

                if scores.student.sex == 'M':    
                    sheet.cell(row=row_coordinates_percentage_score_performance_male, column=column_coordinates_percentage_score_performance, value=value_to_write)
                    column_coordinates_percentage_score_performance += 1

                    if column_coordinates_percentage_score_performance > max_column_index:
                        column_coordinates_percentage_score_performance = 30
                        row_coordinates_percentage_score_performance_male += 1

                elif scores.student.sex == 'F':
                    sheet.cell(row=row_coordinates_percentage_score_performance_female, column=column_coordinates_percentage_score_performance, value=value_to_write)
                    column_coordinates_percentage_score_performance += 1

                    if column_coordinates_percentage_score_performance > max_column_index:
                        column_coordinates_percentage_score_performance = 30
                        row_coordinates_percentage_score_performance_female += 1
                
                else:
                    print(f"Unknown sex for student: {scores.student.name}")
                    continue


        # TOTAL WEIGHTED SCORE WRITTEN WORKS
        column_coordinates_weighted_score_performance = 31
        row_coordinates_weighted_score_performance_male = 12
        row_coordinates_weighted_score_performance_female = 64
        

        for scores in grade_scores_queryset:
            weighted_score_performance = scores.weighted_score_performance

            # Check if the value is not None before rounding
            if weighted_score_performance is not None:
                rounded_weighted_score_performance = round(weighted_score_performance, 2)
                value_to_write = str(rounded_weighted_score_performance)
            else:
                value_to_write = ""  # or any other placeholder for None

            if scores.student.sex == 'M':    

                sheet.cell(row=row_coordinates_weighted_score_performance_male, column=column_coordinates_weighted_score_performance, value=value_to_write)
                column_coordinates_weighted_score_performance += 1

                if column_coordinates_weighted_score_performance > max_column_index:
                    column_coordinates_weighted_score_performance = 31
                    row_coordinates_weighted_score_performance_male += 1

            elif scores.student.sex == 'F':

                sheet.cell(row=row_coordinates_weighted_score_performance_female, column=column_coordinates_weighted_score_performance, value=value_to_write)
                column_coordinates_weighted_score_performance += 1

                if column_coordinates_weighted_score_performance > max_column_index:
                    column_coordinates_weighted_score_performance = 31
                    row_coordinates_weighted_score_performance_female += 1

            else:
                print(f"Unknown sex for student: {scores.student.name}")
                continue


def write_quarterly_assessment_scores(sheet, grade_scores_queryset):
    #    # PERFORMANCE TASKS SCORES
    # column_coordinates = 19
    # row_coordinates = 12
    max_column_index = 31

    # for score in grade_scores_queryset:
    #     performance_tasks_scores_list = score.performance_task_scores

    #     for value in performance_tasks_scores_list:
    #         value_to_write = str(value)
    #         sheet.cell(row=row_coordinates, column=column_coordinates, value=value_to_write)
    #         column_coordinates += 1

    #         # If we reach the last column, move to the next row
    #         if column_coordinates > max_column_index:  # Update max_column_index with the actual maximum column index
    #             column_coordinates = 19  # Reset column index to the starting column
    #             row_coordinates += 1  # Move to the next row

    # TOTAL QUARTERLY ASSESSMENT SCORE
    column_coordinates_total_performance_tasks_score = 32
    row_coordinates_total_performance_tasks_score_male = 12
    row_coordinates_total_performance_tasks_score_female = 64

    for scores in grade_scores_queryset:
        total_performance_tasks_score = scores.total_score_quarterly

        # Convert the float value to a string or set to empty string if None
        value_to_write = str(total_performance_tasks_score) if total_performance_tasks_score is not None and total_performance_tasks_score != 0 else ""

        
        if scores.student.sex == 'M':    
            sheet.cell(row=row_coordinates_total_performance_tasks_score_male, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 32  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_male += 1 

        elif scores.student.sex == 'F':
            sheet.cell(row=row_coordinates_total_performance_tasks_score_female, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 32  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_female += 1 
            
        else:
            print(f"Unknown sex for student: {scores.student.name}")
            continue
        

        # TOTAL PERCENTAGE SCORE WRITTEN WORKS
        column_coordinates_percentage_score_performance = 33
        row_coordinates_percentage_score_performance_male = 12
        row_coordinates_percentage_score_performance_female = 64


        for scores in grade_scores_queryset:
            percentage_score_quarterly = scores.percentage_score_quarterly

            # Check if the value is not None before rounding
            if percentage_score_quarterly is not None:
                rounded_percentage_score = round(percentage_score_quarterly, 2)
                value_to_write = str(rounded_percentage_score)
            else:
                value_to_write = ""  # or any other placeholder for None

            if scores.student.sex == 'M':    

                sheet.cell(row=row_coordinates_percentage_score_performance_male, column=column_coordinates_percentage_score_performance, value=value_to_write)
                column_coordinates_percentage_score_performance += 1

                if column_coordinates_percentage_score_performance > max_column_index:
                    column_coordinates_percentage_score_performance = 33
                    row_coordinates_percentage_score_performance_male += 1

            elif scores.student.sex == 'F':

                sheet.cell(row=row_coordinates_percentage_score_performance_female, column=column_coordinates_percentage_score_performance, value=value_to_write)
                column_coordinates_percentage_score_performance += 1

                if column_coordinates_percentage_score_performance > max_column_index:
                    column_coordinates_percentage_score_performance = 33
                    row_coordinates_percentage_score_performance_female += 1
            
            else:
                print(f"Unknown sex for student: {scores.student.name}")
                continue


    # TOTAL WEIGHTED SCORE WRITTEN WORKS
    column_coordinates_weighted_score_performance = 34
    row_coordinates_weighted_score_performance_male = 12
    row_coordinates_weighted_score_performance_female = 64

    for scores in grade_scores_queryset:
        weighted_score_performance = scores.weighted_score_quarterly

        # Check if the value is not None before rounding
        if weighted_score_performance is not None:
            rounded_weighted_score_performance = round(weighted_score_performance, 2)
            value_to_write = str(rounded_weighted_score_performance)
        else:
            value_to_write = ""  # or any other placeholder for None

        if scores.student.sex == 'M':    
            sheet.cell(row=row_coordinates_weighted_score_performance_male, column=column_coordinates_weighted_score_performance, value=value_to_write)
            column_coordinates_weighted_score_performance += 1

            if column_coordinates_weighted_score_performance > max_column_index:
                column_coordinates_weighted_score_performance = 34
                row_coordinates_weighted_score_performance_male += 1

        elif scores.student.sex == 'F':
            sheet.cell(row=row_coordinates_weighted_score_performance_female, column=column_coordinates_weighted_score_performance, value=value_to_write)
            column_coordinates_weighted_score_performance += 1

            if column_coordinates_weighted_score_performance > max_column_index:
                column_coordinates_weighted_score_performance = 34
                row_coordinates_weighted_score_performance_female += 1
                

        else:
            print(f"Unknown sex for student: {scores.student.name}")
            continue

    

def write_initial_grade(sheet, grade_scores_queryset):
    max_column_index = 35
    column_coordinates_total_performance_tasks_score = 35
    row_coordinates_total_performance_tasks_score_male = 12
    row_coordinates_total_performance_tasks_score_female = 64

    for scores in grade_scores_queryset:
        total_performance_tasks_score = scores.initial_grades

        if total_performance_tasks_score is not None:
            rounded_performance_tasks_score = round(total_performance_tasks_score, 2)
            value_to_write = str(rounded_performance_tasks_score)
        else:
            value_to_write = ""

        if scores.student.sex == 'M':  
            sheet.cell(row=row_coordinates_total_performance_tasks_score_male, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 35  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_male += 1    

        elif scores.student.sex == 'F':

            sheet.cell(row=row_coordinates_total_performance_tasks_score_female, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 35  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_female += 1  
            
        else:
            print(f"Unknown sex for student: {scores.student.name}")
            continue



def write_transmuted_grade(sheet, grade_scores_queryset):
    max_column_index = 36
    column_coordinates_total_performance_tasks_score = 36 
    row_coordinates_total_performance_tasks_score_male = 12
    row_coordinates_total_performance_tasks_score_female = 64

    for scores in grade_scores_queryset:
        total_performance_tasks_score = scores.transmuted_grades

                # Convert the float value to a string
        if total_performance_tasks_score is not None:
            rounded_performance_tasks_score = round(total_performance_tasks_score, 2)
            value_to_write = str(rounded_performance_tasks_score)
        else:
            value_to_write = ""

        if scores.student.sex == 'M':    
            
            sheet.cell(row=row_coordinates_total_performance_tasks_score_male, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 36  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_male += 1  

        elif scores.student.sex == 'F':

            sheet.cell(row=row_coordinates_total_performance_tasks_score_female, column=column_coordinates_total_performance_tasks_score, value=value_to_write)
            column_coordinates_total_performance_tasks_score += 1

            if column_coordinates_total_performance_tasks_score > max_column_index:  # Update max_column_index with the actual maximum column index
                column_coordinates_total_performance_tasks_score = 36  # Reset column index to the starting column
                row_coordinates_total_performance_tasks_score_female += 1  
            
        else:
            print(f"Unknown sex for student: {scores.student.name}")
            continue

def write_sf9_data(front_sheet, student):
    # Assuming Y16 is the cell where you want to write the student's name
    column_coordinates_student_name = 17  # Column Y
    row_coordinates_student_name = 22
    column_coordinates_lrn = 19
    row_coordinates_lrn = 24
    row_coordinates_sex = 26
    column_coordinates_sex = 20
    row_coordinates_grade = 28
    column_coordinates_grade = 17
    row_coordinates_section = 28
    column_coordinates_section = 20
    row_coordinates_age = 26
    column_coordinates_age = 17
    # birthdate = student.birthday

    # Get the current date
    current_date = datetime.now()

    bold_font = Font(bold=True, name='Aparajita')

    # Calculate the age
    # age = current_date.year - birthdate.year - ((current_date.month, current_date.day) < (birthdate.month, birthdate.day))

    # Write the student's name to the specified cell
    front_sheet.cell(row=row_coordinates_student_name, column=column_coordinates_student_name, value=student.name if student else "")
    front_sheet.cell(row=row_coordinates_lrn, column=column_coordinates_lrn, value=student.lrn if student else "")
    front_sheet.cell(row=row_coordinates_sex, column=column_coordinates_sex, value=student.sex if student else "").font = bold_font
    front_sheet.cell(row=row_coordinates_grade, column=column_coordinates_grade, value=student.grade if student else "").font = bold_font
    front_sheet.cell(row=row_coordinates_section, column=column_coordinates_section, value=student.section if student else "").font = bold_font
    # sheet.cell(row=row_coordinates_age, column=column_coordinates_age, value=age if student else "N/A")
    # print(student.name)


def write_sf9_school_info(front_sheet, school_info_queryset):
    row = 40
    column = 16

    # Assuming there's only one principal for a single school information record
    for school_info in school_info_queryset:
        principal = school_info.principal_name
        front_sheet.cell(row=row, column=column, value=principal)


def write_sf9_grades(back_sheet, advisory_class, general_average):
    # Retrieve grades data from the AdvisoryClass instance
    grades_data = advisory_class.grades_data
    
    # Define the row coordinates for subjects
    row_coordinates = {
        "FILIPINO": 7,
        "ENGLISH": 8,
        "MATHEMATICS": 9,
        "SCIENCE": 11,
        "ARALING PANLIPUNAN": 12, 
        "EDUKASYON SA PAGPAPAKATAO": 13,
        "EDUKASYONG PANTAHANAN AT PANGKABUHAYAN": 15,
        "MAPEH": 17,
        "Music": 18,
        "Arts": 19,
        "PE": 20,
        "Health": 21,
        # Add more subjects as needed
    }

    # Write the grades for each subject
    for subject, data in grades_data.items():
        row = row_coordinates.get(subject)
        if row is None:
            continue  # Skip if subject not found in the coordinates

        # Write quarter grades and final grade to the back sheet
        back_sheet.cell(row=row, column=14, value=data.get("first_quarter", ""))
        back_sheet.cell(row=row, column=15, value=data.get("second_quarter", ""))
        back_sheet.cell(row=row, column=16, value=data.get("third_quarter", ""))
        back_sheet.cell(row=row, column=17, value=data.get("fourth_quarter", ""))
        back_sheet.cell(row=row, column=18, value=data.get("final_grade", ""))

        # Write general average
        back_sheet.cell(row=22, column=18, value=general_average.general_average)

def write_sf9_total_attendance(front_sheet, attendance_record):
    attendance_data = attendance_record.attendance_record
    row_total_school_days = 7
    row_total_days_present = 9
    row_total_days_absent = 12 

    column_coordinates = 13

    for month, record in attendance_data.items():
        total_school_days = record.get('Total School Days', '')
        total_days_present = record.get('Total Days Present', '')
        total_days_absent = record.get('Total Days Absent', '')



        front_sheet.cell(row=row_total_school_days, column=column_coordinates, value=total_school_days)
        front_sheet.cell(row=row_total_days_present, column=column_coordinates, value=total_days_present)
        front_sheet.cell(row=row_total_days_absent, column=column_coordinates, value=total_days_absent)

def write_sf9_attendance(front_sheet, attendance_record):
    attendance_data = attendance_record.attendance_record
    row_school_days = 7 
    row_days_present = 9
    row_days_absent =  12
    row_total_school_days = 7
    row_total_days_present = 9
    row_total_days_absent = 12 # Starting row for school days
    column_coordinates = {
        "JUNE": 2,
        "JULY": 3,
        "AUGUST": 4,
        "SEPTEMBER": 5,
        "OCTOBER": 6,
        "NOVEMBER": 7,
        "DECEMBER": 8,
        "JANUARY": 9,
        "FEBRUARY": 10,
        "MARCH": 11,
        "APRIL": 12,
        "TOTAL": 13,
        # Add more months as needed
    }

    # Iterate over the months and their attendance records
    for month, record in attendance_data.items():
        print(f"month {month}")
        # Access the number of school days for the current month
        school_days = record.get('No. of School Days', '')
        days_present = record.get('No. of Days Present', '')
        days_absent = record.get('No. of Days Absent', '')

        # print(f"school days:{school_days}")
        # print(f"days present:{days_present}")
        # print(f"days absent:{days_absent}")
        
        # Get the column index based on the month
        column_index = column_coordinates.get(month, -1)
        # print(f"month: {column_index}")
            # Write the school days to the respective column for the current month
            # Adjust row and column indices here
        front_sheet.cell(row=row_school_days, column=column_index, value=school_days)
        front_sheet.cell(row=row_days_present, column=column_index, value=days_present)
        front_sheet.cell(row=row_days_absent, column=column_index, value=days_absent)


def get_behavior_rows():
    # Define behavior rows
    behavior_rows = {
        "Expresses one’s spiritual beliefs while respecting the spiritual beliefs of others": 7,
        "Show adherence to ethical principles by upholding truth": 10,
        "Is sensitive to individual, social and cultural differences": 13,
        "Demonstrates contributions toward solidarity": 16,
        "Cares for the environment and utilizes resources wisely, judiciously, and economically.": 18,
        "Demonstrates pride in being a Filipino; exercises the rights and responsibilities of a Filipino citizen": 20,
        "Demonstrates appropriate behavior in carrying out activities in the school, community, and country": 22
    }
    return behavior_rows

def write_sf9_learners_observation(back_sheet, learners_observation, quarter, column_index):
    # Retrieve the observations for the specified quarter
    observation_data = getattr(learners_observation, quarter, None)
    if observation_data:
        behavior_rows = get_behavior_rows()
        print(f"observation: {observation_data}")

        for domain, behaviors in observation_data.items():
            for behavior in behaviors:
                behavior_statement = behavior['behavior_statement']
                marking = behavior['marking']

                print(f"behavior_statement: {behavior_statement}")
                print(f"marking: {marking}")

                # Check if the behavior statement exists in the behavior_rows dictionary
                if behavior_statement in behavior_rows:
                    row_index = behavior_rows[behavior_statement]
                    back_sheet.cell(row=row_index, column=column_index, value=marking)
    #             # Iterate through the behavior_rows dictionary to print each marking
    # for row, row_index in behavior_rows.items():
    #             print(f"marking: {marking}")
    #             print(f"row {row_index}")
    #             back_sheet.cell(row=row_index, column=column_index, value=marking)
                
                

  

#SUMMARRY OF QUARTERLY GRADES
        
def write_school_info_quarterly(input_sheet, school_info, quarterly_grades_query):
    column_coordinates_region = 7
    column_coordinates_school_name = 7
    column_coordinates_district = 24
    column_coordinates_school_year = 33
    column_coordinates_schoolID = 24
    column_coordinates_division = 15
    row_coordinates_4 = 4
    row_coordinates_5 = 5
    column_coordinates_grade_section = 11
    column_coordinates_teacher = 19
    row_coordinates = 7

    # Iterate over school information and quarterly grades
    for info, grade_instance in zip(school_info, quarterly_grades_query):
        # School Information
        value_region = info.region
        value_school_name = info.school_name
        value_division = info.division
        value_school_id = info.school_id
        value_district = info.district
        value_school_year = info.school_year

        # Quarterly Grades Information
        grade = grade_instance.student.grade
        section = grade_instance.student.section
        teacher_name = f"{grade_instance.student.teacher.user.first_name} {grade_instance.student.teacher.user.last_name}"


        print(grade)
        print(section)
        print(teacher_name)

        grade_section = f"{grade} - {section}"

        # Write school information
        input_sheet.cell(row=row_coordinates_4, column=column_coordinates_region, value=value_region)
        input_sheet.cell(row=row_coordinates_5, column=column_coordinates_school_name, value=value_school_name)
        input_sheet.cell(row=row_coordinates_4, column=column_coordinates_division, value=value_division)
        input_sheet.cell(row=row_coordinates_5, column=column_coordinates_schoolID, value=value_school_id)
        input_sheet.cell(row=row_coordinates_4, column=column_coordinates_district, value=value_district)
        input_sheet.cell(row=row_coordinates_5, column=column_coordinates_school_year, value=value_school_year)

        # Write grade, section, and teacher
        input_sheet.cell(row=row_coordinates, column=column_coordinates_grade_section, value=grade_section)
        input_sheet.cell(row=row_coordinates, column=column_coordinates_teacher, value=teacher_name)

        

def write_student_name_quarterly(input_sheet, quarterly_grades_query):

    column_coordinates_student_name = 2
    row_coordinates_male = 12
    row_coordinates_female = 62

    while input_sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name).value:
        row_coordinates_male += 1

    while input_sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name).value:
        row_coordinates_female += 1

    for quarterly_grade in quarterly_grades_query:  # Retrieve the related student
         # Check if the related student exists
       

        if quarterly_grade.student.sex == 'M':
            value_to_write = quarterly_grade.student.name
            input_sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name, value=value_to_write)
            row_coordinates_male += 1
        elif quarterly_grade.student.sex == 'F':
            value_to_write = quarterly_grade.student.name
            input_sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name, value=value_to_write)
            row_coordinates_female += 1


def write_quarterly_grade_AP(sheet, quarterly_grades_query):
    column_coordinates_AP = 8
    row_coordinates_male = 11
    row_coordinates_female = 62

    for quarterly_grade in quarterly_grades_query:
        # Parse the JSON field to extract the value of "AP"
        grades_data = quarterly_grade.grades


        if quarterly_grade.student.sex == 'M':
                value_to_write = grades_data.get("AP")
                sheet.cell(row=row_coordinates_male, column=column_coordinates_AP, value=value_to_write)
                row_coordinates_male += 1
        elif quarterly_grade.student.sex == 'F':
                value_to_write = grades_data.get("AP")
                sheet.cell(row=row_coordinates_female, column=column_coordinates_AP, value=value_to_write)
                row_coordinates_female += 1

   

def write_quarterly_grade_ENGLISH(sheet, quarterly_grades_query):
    column_coordinates_ENGLISH = 6
    row_coordinates_male = 11
    row_coordinates_female = 62

    for quarterly_grade in quarterly_grades_query:
        # Parse the JSON field to extract the value of "AP"
        grades_data = quarterly_grade.grades



        if quarterly_grade.student.sex == 'M':
                value_to_write = grades_data.get("English")
                sheet.cell(row=row_coordinates_male, column=column_coordinates_ENGLISH, value=value_to_write)
                row_coordinates_male += 1
        elif quarterly_grade.student.sex == 'F':
                value_to_write = grades_data.get("ENGLISH")
                sheet.cell(row=row_coordinates_female, column=column_coordinates_ENGLISH, value=value_to_write)
                row_coordinates_female += 1            




#FINAL GRADE AND GENERAL AVERAGE

def write_school_info_general_average(sheet, school_info, general_grades_query):
    column_coordinates_region = 7
    column_coordinates_school_name = 7
    # column_coordinates_district = 24
    column_coordinates_school_year = 37
    column_coordinates_schoolID = 21
    column_coordinates_division = 21
    row_coordinates_3 = 3
    # row_coordinates_4 = 4
    row_coordinates_5 = 5
    column_coordinates_grade_section = 37
    # column_coordinates_teacher = 32
    # row_coordinates = 7

    # Iterate over school information and quarterly grades
    for info, grade_instance in zip(school_info, general_grades_query):
        # School Information
        value_region = info.region
        value_school_name = info.school_name
        value_division = info.division
        value_school_id = info.school_id
        # value_district = info.district
        value_school_year = info.school_year

        # Quarterly Grades Information
        grade = grade_instance.student.grade
        section = grade_instance.student.section
        # teacher_name = f"{grade_instance.student.teacher.user.first_name} {grade_instance.student.teacher.user.last_name}"


        print(grade)
        print(section)
        # print(teacher_name)

        grade_section = f"{grade} - {section}"

        # Write school information
        sheet.cell(row=row_coordinates_3, column=column_coordinates_region, value=value_region)
        sheet.cell(row=row_coordinates_5, column=column_coordinates_school_name, value=value_school_name)
        sheet.cell(row=row_coordinates_3, column=column_coordinates_division, value=value_division)
        sheet.cell(row=row_coordinates_5, column=column_coordinates_schoolID, value=value_school_id)
        # sheet.cell(row=row_coordinates_4, column=column_coordinates_district, value=value_district)
        sheet.cell(row=row_coordinates_3, column=column_coordinates_school_year, value=value_school_year)

        # Write grade, section, and teacher
        sheet.cell(row=row_coordinates_5, column=column_coordinates_grade_section, value=grade_section)
        # sheet.cell(row=row_coordinates, column=column_coordinates_teacher, value=teacher_name)


def write_student_name_general_average(sheet, general_grades_query):
    column_coordinates_student_name = 2
    column_coordinates_general_average = 46
    row_coordinates_male = 11
    row_coordinates_female = 62

    while sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name).value:
        row_coordinates_male += 1

    while sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name).value:
        row_coordinates_female += 1

    for general_average_grade in general_grades_query:  # Retrieve the related student
        # Check if the related student exists

        if general_average_grade.student.sex == 'M':
            student_name = general_average_grade.student.name
            general_average = general_average_grade.general_average
            sheet.cell(row=row_coordinates_male, column=column_coordinates_student_name, value=student_name)
            sheet.cell(row=row_coordinates_male, column=column_coordinates_general_average, value=general_average)
            row_coordinates_male += 1
            print(f'male {student_name}')
        elif general_average_grade.student.sex == 'F':
            student_name = general_average_grade.student.name
            general_average = general_average_grade.general_average
            sheet.cell(row=row_coordinates_female, column=column_coordinates_student_name, value=student_name)
            sheet.cell(row=row_coordinates_female, column=column_coordinates_general_average, value=general_average)
            row_coordinates_female += 1
            print(f'female {student_name}')

def write_final_grade_subject(sheet, advisory_class_query, general_grades_query, subject_name, column_start, column_end):
    
    row_coordinates_male = 11
    row_coordinates_female = 62

    for advisory_class in advisory_class_query:
        grades_data = advisory_class.grades_data
        subject_data = grades_data.get(subject_name)

        if not subject_data:
            continue

        first_quarter_grade = convert_to_int(subject_data.get('first_quarter', ''))
        second_quarter_grade = convert_to_int(subject_data.get('second_quarter', ''))
        third_quarter_grade = convert_to_int(subject_data.get('third_quarter', ''))
        fourth_quarter_grade = convert_to_int(subject_data.get('fourth_quarter', ''))
        final_grade = convert_to_int(subject_data.get('final_grade', ''))

        if advisory_class.student.sex == 'M':
            sheet.cell(row=row_coordinates_male, column=column_start, value=first_quarter_grade)
            sheet.cell(row=row_coordinates_male, column=column_start + 1, value=second_quarter_grade)
            sheet.cell(row=row_coordinates_male, column=column_start + 2, value=third_quarter_grade)
            sheet.cell(row=row_coordinates_male, column=column_start + 3, value=fourth_quarter_grade)
            sheet.cell(row=row_coordinates_male, column=column_end, value=final_grade)
            row_coordinates_male += 1
        elif advisory_class.student.sex == 'F':
            sheet.cell(row=row_coordinates_female, column=column_start, value=first_quarter_grade)
            sheet.cell(row=row_coordinates_female, column=column_start + 1, value=second_quarter_grade)
            sheet.cell(row=row_coordinates_female, column=column_start + 2, value=third_quarter_grade)
            sheet.cell(row=row_coordinates_female, column=column_start + 3, value=fourth_quarter_grade)
            sheet.cell(row=row_coordinates_female, column=column_end, value=final_grade)
            row_coordinates_female += 1

# Example usage:

def convert_to_int(grade):
    try:
        if grade is not None:
            # Try converting to float first
            grade_float = float(grade)
            # Then convert to int
            return int(grade_float)
        else:
            # Return 0 if grade is None
            return ""
    except ValueError:
        # Return 0 if conversion fails
        return ""
