from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponseNotFound
from .models import Student, FinalGrade, GeneralAverage, ClassRecord, Subject, SchoolInformation, QuarterlyGrades, AdvisoryClass, AttendanceRecord, LearnersObservation,BehaviorStatement,  SchoolInformation
from .views import log_activity
import os
import openpyxl
import pandas as pd
import shutil
from datetime import datetime
import urllib.parse
from django.contrib import messages 
from django.conf import settings

from .utils import (
    write_student_names,
    write_scores_hps_written,
    write_scores_hps_performance,
    write_scores_hps_quarterly,
    write_written_works_scores,
    write_performance_tasks_scores,
    write_quarterly_assessment_scores,
    write_initial_grade,
    write_transmuted_grade,
    write_sf9_data,
    write_sf9_grades,
    write_school_info,
    write_sf9_attendance,
    write_sf9_total_attendance,
    write_sf9_learners_observation,
    write_sf9_school_info,

    #SUMMARY OF QUARTERLY
    write_school_info_quarterly,
    write_student_name_quarterly,
    write_quarterly_grade_AP,
    write_quarterly_grade_ENGLISH,

    #FINAL GRADE AND GENERAL AVERAGE
    write_school_info_general_average,
    write_student_name_general_average,
    write_final_grade_subject
)
from .models import GradeScores

def generate_summary_of_quarterly_grades(request, grade, section, quarter):
    # Assuming `grade`, `section`, and `quarter` are passed in the URL parameters

    # Query the QuarterlyGrades model
    grade_scores_queryset = GradeScores.objects.filter(
        class_record__quarters=quarter, student__grade=grade, student__section=section,
    )

    grade_name = grade
    section_name = section
    quarter_name = quarter

    school_info = SchoolInformation.objects.all()

    print(grade_scores_queryset)

    # Initialize lists to hold data for each subject
    subject_data = {}
    for grade_score in GradeScores.objects.filter(class_record__quarters=quarter, student__grade=grade, student__section=section):
        class_record_name = grade_score.class_record.name
        teacher_name = f"{grade_score.class_record.teacher.user.first_name} {grade_score.class_record.teacher.user.last_name}".upper()

        # Assuming `subjects` is the related field to GradeScores
        subject_name = grade_score.class_record.subject

        if subject_name not in subject_data:
            subject_data[subject_name] = {'data': [], 'gradescores': []}
        
        # Append the grade score and its corresponding GradeScores object to the dictionary
        subject_data[subject_name]['data'].append((class_record_name, teacher_name))
        subject_data[subject_name]['gradescores'].append(grade_score)

    user = request.user
    action = f'{user} generate a Summary of Quarterly Grades ({quarter}) of {grade} {section}'
    details = f'{user} generate a Summary of Quarterly Grades ({quarter}) of {grade} {section} in the system.'
    log_activity(user, action, details)

    if grade_name == "Grade 1":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 1_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 1_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')

        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 1_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 1_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')

    elif grade_name == "Grade 2":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 2_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 2_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
        
        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 2_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 2_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
        
    elif grade_name == "Grade 3":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 3_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 3_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')

        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 3_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 3_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            


    original_file_path = os.path.join(media_directory, excel_file_name)

    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-sf9')
    if not os.path.exists(created_directory):
        os.makedirs(created_directory)

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'SUMMARY OF QUARTERLY GRADES {quarter} {timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

    # Open the copied SF9 Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

    sheet_input_data = 'INPUT DATA'
    sheet_input = workbook[sheet_input_data]

    for subject_name, subject_info in subject_data.items():
        # Extract data and GradeScores objects for the current subject
        data = subject_info['data']
        grade_scores = subject_info['gradescores']

        # Determine the Excel sheet name for each subject
    
        if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
        elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
        elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
        elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
        elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
        elif subject_name in ["ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
        elif subject_name == "MUSIC":
                desired_sheet_name = 'MUSIC '
        elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
        elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
        else:
                desired_sheet_name = 'Sheet1'
       

        # Select the desired sheet (use the correct sheet name from the output)
        sheet = workbook[desired_sheet_name]

        write_scores_hps_written(sheet, grade_scores)
        write_scores_hps_performance(sheet, grade_scores)
        write_scores_hps_quarterly(sheet, grade_scores)
        write_written_works_scores(sheet, grade_scores)
        write_performance_tasks_scores(sheet, grade_scores)
        write_quarterly_assessment_scores(sheet, grade_scores)
        write_initial_grade(sheet, grade_scores)
        write_transmuted_grade(sheet, grade_scores)
        
    write_student_names(sheet_input, grade_scores)
    write_school_info(sheet_input, school_info, grade, section, teacher_name)

    # Save the changes to the SF9 workbook
    workbook.save(copied_file_path)

    # Create an HTTP response with the updated SF9 Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'SUMMARY OF QUARTERLY GRADES ({quarter}) {grade} - {section} {excel_file_name}_{timestamp}.xlsx'
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

    with open(copied_file_path, 'rb') as excel_file:
        response.write(excel_file.read())

    return response

def generate_summary_of_mapeh(request, grade, section, subject, quarter,):
    # Assuming `grade`, `section`, and `quarter` are passed in the URL parameters

    # Query the QuarterlyGrades model
    grade_scores_queryset = GradeScores.objects.filter(
        class_record__quarters=quarter, student__grade=grade, student__section=section,
    )

    grade_name = grade
    section_name = section
    quarter_name = quarter
    subject_name = subject

    school_info = SchoolInformation.objects.all()

    print(grade_scores_queryset)

    # Initialize lists to hold data for each subject
    subject_data = {}
    for grade_score in GradeScores.objects.filter(class_record__quarters=quarter, student__grade=grade, student__section=section):
        class_record_name = grade_score.class_record.name
        teacher_name = f"{grade_score.class_record.teacher.user.first_name} {grade_score.class_record.teacher.user.last_name}".upper()

        # Assuming `subjects` is the related field to GradeScores
        subject_name = grade_score.class_record.subject

        if subject_name not in subject_data:
            subject_data[subject_name] = {'data': [], 'gradescores': []}
        
        # Append the grade score and its corresponding GradeScores object to the dictionary
        subject_data[subject_name]['data'].append((class_record_name, teacher_name))
        subject_data[subject_name]['gradescores'].append(grade_score)

    user = request.user
    action = f'{user} generate MAPEH Class Record of {grade} {section} ({quarter})'
    details = f'{user} generate MAPEH Class Record of {grade} {section} ({quarter}) in the system.'
    log_activity(user, action, details)

    

    if quarter_name == "1st Quarter":
                    excel_file_name = "GRADE 4-6_ 1ST QUARTER MAPEH.xlsx"
                    media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
    elif quarter_name == "2nd Quarter":
                    excel_file_name = "GRADE 4-6_ 2ND QUARTER MAPEH.xlsx"
                    media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
    elif quarter_name == "3rd Quarter":
                    excel_file_name = "GRADE 4-6_ 3RD QUARTER MAPEH.xlsx"
                    media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
    elif quarter_name == "4th Quarter":
                    excel_file_name = "GRADE 4-6_ 4TH QUARTER MAPEH.xlsx"
                    media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
    
            


    original_file_path = os.path.join(media_directory, excel_file_name)

    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-sf9')
    if not os.path.exists(created_directory):
        os.makedirs(created_directory)

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'Class Record {grade_name}_{section_name}_{subject_name}_{quarter_name}_{timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

    # Open the copied SF9 Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

    sheet_input_data = 'INPUT DATA'
    sheet_input = workbook[sheet_input_data]

    for subject_name, subject_info in subject_data.items():
        # Extract data and GradeScores objects for the current subject
        data = subject_info['data']
        grade_scores = subject_info['gradescores']

        # Determine the Excel sheet name for each subject
    

        if subject_name == "ARTS":
                desired_sheet_name = 'ARTS'
        elif subject_name == "MUSIC":
                desired_sheet_name = 'MUSIC'
        elif subject_name == "PE":
                desired_sheet_name = 'PE'
        elif subject_name == "HEALTH":
                desired_sheet_name = 'HEALTH'
        else:
                desired_sheet_name = 'Sheet1'
       

        # Select the desired sheet (use the correct sheet name from the output)
        sheet = workbook[desired_sheet_name]

        write_scores_hps_written(sheet, grade_scores)
        write_scores_hps_performance(sheet, grade_scores)
        write_scores_hps_quarterly(sheet, grade_scores)
        write_written_works_scores(sheet, grade_scores)
        write_performance_tasks_scores(sheet, grade_scores)
        write_quarterly_assessment_scores(sheet, grade_scores)
        write_initial_grade(sheet, grade_scores)
        write_transmuted_grade(sheet, grade_scores)
        
    write_student_names(sheet_input, grade_scores)
    write_school_info(sheet_input, school_info, grade, section, teacher_name)

    # Save the changes to the SF9 workbook
    workbook.save(copied_file_path)

    # Create an HTTP response with the updated SF9 Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Class Record {grade_name}_{section_name}_{subject}_{quarter_name}_{timestamp}.xlsx'
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

    with open(copied_file_path, 'rb') as excel_file:
        response.write(excel_file.read())

    response.content += b"""
    <script>
        // Refresh the page after download
        setTimeout(function() {
            window.location.reload(true);
        }, 1000); // 1000 milliseconds = 1 second
    </script>
    """
    return response



def generate_final_and_general_grades(request, grade, section):

    general_grades_query = GeneralAverage.objects.filter(
        student__grade=grade, student__section=section,
    )

    advisory_class_query = AdvisoryClass.objects.filter(
        grade=grade,
        section=section
    )

    school_info = SchoolInformation.objects.all()


    user = request.user
    action = f'{user} generate the Summary of Final Grade and General Average of {grade} {section}'
    details = f'{user} generate the Summary of Final Grade and General Average of {grade} {section} in the system.'
    log_activity(user, action, details)


    print(grade)
    print(section)

    # students = [quarterly_grade.student for quarterly_grade in quarterly_grades_query]

    # # Now you can access the students
    # for student in students:
    #     print(student.name)

     # Original file path for SF9 template
        # Original file path
    if grade == 'Grade 1':
        excel_file_name = "GRADE 1_SUMMARY FINAL GRADES.xlsx"
    elif grade == 'Grade 2':
        excel_file_name = "GRADE 2_SUMMARY FINAL GRADES.xlsx"
    elif grade == 'Grade 3':
        excel_file_name = "GRADE 3_SUMMARY FINAL GRADES.xlsx"
    elif grade == 'Grade 4' or grade == 'Grade 5' or grade == 'Grade 6' :
        excel_file_name = "GRADE 4-6_SUMMARY FINAL GRADES.xlsx"
   

    # Get the current working directory
    # current_directory = os.getcwd()
    media_directory = os.path.join(settings.MEDIA_ROOT, 'excel-files')
    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-sf9')

    if not os.path.exists(media_directory):
        os.makedirs(media_directory)

    if not os.path.exists(created_directory):
        os.makedirs(created_directory)

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create the path for the original file
    original_file_path = os.path.join(media_directory, excel_file_name)

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'Final and General Average {grade} {section} {timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

    
        # Open the copied SF9 Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

            # Select the desired sheet (use the correct sheet name from the output)
        # input_sheet_name = 'INPUT DATA'
        # input_sheet = workbook[input_sheet_name]

    sheet_name = 'SUMMARY - FINAL GRADES'
    sheet = workbook[sheet_name]
    if grade == 'Grade 1':
        write_school_info_general_average(sheet, school_info, general_grades_query)
        write_student_name_general_average(sheet, general_grades_query)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MOTHER TONGUE', 6, 10)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'FILIPINO', 11, 15)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ENGLISH', 16, 20)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MATHEMATICS', 21, 25)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ARALING PANLIPUNAN', 26, 30)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'EDUKASYON SA PAGPAPAKATAO', 31, 35)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MAPEH', 36, 40)

    elif grade == 'Grade 2':
        write_school_info_general_average(sheet, school_info, general_grades_query)
        write_student_name_general_average(sheet, general_grades_query)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MOTHER TONGUE', 6, 10)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'FILIPINO', 11, 15)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ENGLISH', 16, 20)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MATHEMATICS', 21, 25)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ARALING PANLIPUNAN', 26, 30)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'EDUKASYON SA PAGPAPAKATAO', 31, 35)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MAPEH', 36, 40)

    elif grade == 'Grade 3':
        write_school_info_general_average(sheet, school_info, general_grades_query)
        write_student_name_general_average(sheet, general_grades_query)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MOTHER TONGUE', 6, 10)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'FILIPINO', 11, 15)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ENGLISH', 16, 20)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MATHEMATICS', 21, 25)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'SCIENCE', 26, 30)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ARALING PANLIPUNAN', 31, 35)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'EDUKASYON SA PAGPAPAKATAO', 36, 40)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MAPEH', 41, 45)

    elif grade == 'Grade 4' or grade == 'Grade 5' or grade == 'Grade 6' :
        write_school_info_general_average(sheet, school_info, general_grades_query)
        write_student_name_general_average(sheet, general_grades_query)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'FILIPINO', 6, 10)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ENGLISH', 11, 15)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MATHEMATICS', 16, 20)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'SCIENCE', 21, 25)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'ARALING PANLIPUNAN', 26, 30)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'EPP', 31, 35)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'MAPEH', 36, 40)
        write_final_grade_subject(sheet, advisory_class_query, general_grades_query, 'EDUKASYON SA PAGPAPAKATAO', 41, 45)

        # write_quarterly_grade_AP(sheet, quarterly_grades_query)
        # write_quarterly_grade_ENGLISH(sheet, quarterly_grades_query)

            # Save the changes to the SF9 workbook
    workbook.save(copied_file_path)

            # Create an HTTP response with the updated SF9 Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Final and General Average {grade} {section} {timestamp}.xlsx'
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

    with open(copied_file_path, 'rb') as excel_file:
                response.write(excel_file.read())

    return response

def generate_excel_for_grades(request, grade, section, subject, quarter):
    # Get GradeScores for the specified grade, section, and subject
    grade_scores_queryset = GradeScores.objects.filter(class_record__grade=grade,
                                                        class_record__section=section,
                                                        class_record__subject=subject,
                                                        class_record__quarters=quarter)
    school_info = SchoolInformation.objects.all()

    grade_name = grade
    section_name = section
    subject_name = subject
    quarter_name = quarter

    print(quarter_name)
    print(grade_name)
    print(subject_name)
    print(section_name)

    # print(quarter)
   
    for grade_score in grade_scores_queryset:
        class_record_name = grade_score.class_record.name
        teacher_name = f"{grade_score.class_record.teacher.user.first_name} {grade_score.class_record.teacher.user.last_name}".upper()
        

    user = request.user
    action = f'{user} generate a Class Record name "{class_record_name}"'
    details = f'{user} generate a Class Record named "{class_record_name}" in the system.'
    log_activity(user, action, details)


        # Original file path
    if grade_name in ["Grade 4", "Grade 5", "Grade 6"]:
        if subject_name == "MATHEMATICS":
            excel_file_name = "GRADE 4-6_MATHEMATICS.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name == "ARALING PANLIPUNAN":
            excel_file_name = "GRADE 4-6_ARALING PANLIPUNAN.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
            excel_file_name = "GRADE 4-6_EDUKASYON SA PAGPAPAKATAO.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'            
            

        elif subject_name == "EPP":
            excel_file_name = "GRADE 4-6_EDUKASYONG PANTAHANAN AT PANGKABUHAYAN.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name == "ENGLISH":
            excel_file_name = "GRADE 4-6_ENGLISH.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name == "FILIPINO":
            excel_file_name = "GRADE 4-6_FILIPINO.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name == "SCIENCE":
            excel_file_name = "GRADE 4-6_SCIENCE.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
            if quarter_name == "1st Quarter":
                desired_sheet_name = 'Q1'
            elif quarter_name == "2nd Quarter":
                desired_sheet_name = 'Q2'
            elif quarter_name == "3rd Quarter":
                desired_sheet_name = 'Q3'
            elif quarter_name == "4th Quarter":
                desired_sheet_name = 'Q4'

        elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
            if quarter_name == "1st Quarter":
                excel_file_name = "GRADE 4-6_ 1ST QUARTER MAPEH.xlsx"
                media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
                if subject_name == 'MUSIC':
                    desired_sheet_name = 'MUSIC'
                elif subject_name == 'ARTS':
                    desired_sheet_name = 'ARTS'
                elif subject_name == 'PE':
                    desired_sheet_name = 'PE'
                elif subject_name == 'HEALTH':
                    desired_sheet_name = 'HEALTH'
            elif quarter_name == "2nd Quarter":
                excel_file_name = "GRADE 4-6_ 2ND QUARTER MAPEH.xlsx"
                media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
                if subject_name == 'MUSIC':
                    desired_sheet_name = 'MUSIC'
                elif subject_name == 'ARTS':
                    desired_sheet_name = 'ARTS'
                elif subject_name == 'PE':
                    desired_sheet_name = 'PE'
                elif subject_name == 'HEALTH':
                    desired_sheet_name = 'HEALTH'
            elif quarter_name == "3rd Quarter":
                excel_file_name = "GRADE 4-6_ 3RD QUARTER MAPEH.xlsx"
                media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
                if subject_name == 'MUSIC':
                    desired_sheet_name = 'MUSIC'
                elif subject_name == 'ARTS':
                    desired_sheet_name = 'ARTS'
                elif subject_name == 'PE':
                    desired_sheet_name = 'PE'
                elif subject_name == 'HEALTH':
                    desired_sheet_name = 'HEALTH'
            elif quarter_name == "4th Quarter":
                excel_file_name = "GRADE 4-6_ 4TH QUARTER MAPEH.xlsx"
                media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1', 'MAPEH')
                if subject_name == 'MUSIC':
                    desired_sheet_name = 'MUSIC'
                elif subject_name == 'ARTS':
                    desired_sheet_name = 'ARTS'
                elif subject_name == 'PE':
                    desired_sheet_name = 'PE'
                elif subject_name == 'HEALTH':
                    desired_sheet_name = 'HEALTH'


    elif grade_name == "Grade 1":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 1_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')
            print(media_directory)
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 1_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'            
        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 1_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 1_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        else:
            excel_file_name = "TEMPLATE - CLASSRECORD (2).xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-1_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

    elif grade_name == "Grade 2":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 2_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 2_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 2_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 2_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        else:
            excel_file_name = "TEMPLATE - CLASSRECORD (2).xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-2_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'
            
    elif grade_name == "Grade 3":
        if quarter_name == "1st Quarter":
            excel_file_name = f"GRADE 3_1ST QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "2nd Quarter":
            excel_file_name = f"GRADE 3_2ND QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "3rd Quarter":
            excel_file_name = f"GRADE 3_3RD QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        elif quarter_name == "4th Quarter":
            excel_file_name = f"GRADE 3_4TH QUARTER.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

        else:
            excel_file_name = "TEMPLATE - CLASSRECORD (2).xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-3_E-Class-Record-Templates')
            if subject_name == "MOTHER TONGUE":
                desired_sheet_name = 'MTB'
            elif subject_name == "FILIPINO":
                desired_sheet_name = 'FILIPINO'
            elif subject_name == "ENGLISH":
                desired_sheet_name = 'ENGLISH'
            elif subject_name == "MATHEMATICS":
                desired_sheet_name = 'MATH'
            elif subject_name == "ARALING PANLIPUNAN":
                desired_sheet_name = 'AP'
            elif subject_name in ["MUSIC", "ARTS", "PE", "HEALTH"]:
                desired_sheet_name = subject_name
            elif subject_name == "EDUKASYON SA PAGPAPAKATAO":
                desired_sheet_name = 'ESP'
            elif subject_name == "SCIENCE":
                desired_sheet_name = 'SCIENCE'
            else:
                desired_sheet_name = 'Sheet1'

    else:
        excel_file_name = "TEMPLATE - CLASSRECORD (2).xlsx"
        media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord')




    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-class-record')

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create the path for the original file
    original_file_path = os.path.join(media_directory, excel_file_name)

    print("media",media_directory)
    print("excel",excel_file_name)

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'SF1 {grade_name}_{section_name}_{excel_file_name}{timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

        # Open the copied Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

        # Select the desired sheet (use the correct sheet name from the output)
    sheet = workbook[desired_sheet_name]

    sheet_input_data = 'INPUT DATA'
    sheet_input = workbook[sheet_input_data]

        # # Write student names
    write_student_names(sheet_input, grade_scores_queryset)

        # Write scores_hps_written
    write_scores_hps_written(sheet, grade_scores_queryset)

        # Write scores_hps_performance
    write_scores_hps_performance(sheet, grade_scores_queryset)

        # Write scores_hps_quarterly
    write_scores_hps_quarterly(sheet, grade_scores_queryset)

        # Write Written_works_score
    write_written_works_scores(sheet, grade_scores_queryset)

        # Write performance_tasks_score
    write_performance_tasks_scores(sheet, grade_scores_queryset)

        # Write quarterly assessment score
    write_quarterly_assessment_scores(sheet, grade_scores_queryset)

        # Write Initial Grade
    write_initial_grade(sheet, grade_scores_queryset)

        # Write transmuted Grade
    write_transmuted_grade(sheet, grade_scores_queryset)


    write_school_info(sheet_input, school_info, grade, section, teacher_name)


        # Save the changes to the workbook
    workbook.save(copied_file_path)

        # Create an HTTP response with the updated Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Class Record {grade_name}_{section_name}_{subject_name}_{quarter_name}_{timestamp}.xlsx'

    with open(copied_file_path, 'rb') as excel_file:
            response.write(excel_file.read())
    
    return response

    # except Exception as e:
    #     # Handle exceptions, such as a corrupted file
    #     return HttpResponse(f"An error occurred: {e}")
    
def generate_excel_for_all_subjects(request, grade, section, quarter):
    # Get GradeScores for the specified grade, section, and quarter
    grade_scores_queryset = GradeScores.objects.filter(class_record__grade=grade,
                                                        class_record__section=section,
                                                        class_record__quarters=quarter)
    
    print(grade_scores_queryset)
    school_info = SchoolInformation.objects.all()
    
    grade_name = grade
    section_name = section
    quarter_name = quarter
    # print(grade_scores_queryset)

    # Logging activity
    user = request.user
    class_record_name = ""  # Initialize class record name
    for grade_score in grade_scores_queryset:
        class_record_name = grade_score.class_record.name

    action = f'{user} generate a Class Record name "{class_record_name}"'
    details = f'{user} generate a Class Record named "{class_record_name}" in the system.'
    log_activity(user, action, details)

    # File paths
    excel_file_name = "Template - ClassRecord.xlsx"
    media_directory = os.path.join(settings.MEDIA_ROOT, 'excel-files')

    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-class-record')
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    original_file_path = os.path.join(media_directory, excel_file_name)
    print("media",media_directory)
    print("excel",excel_file_name)
    copied_file_path = os.path.join(created_directory, f'ClassRecord_{grade_name}_{section_name}_AllSubjects_{quarter_name}_{timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)


        # Open the copied Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

        # Iterate through all sheets in the workbook
    
            # Identify the subject based on the sheet name
    sheet_input_data = 'INPUT DATA'
    sheet = workbook[sheet_input_data]

            # Write data to the selected sheet
    write_student_names(sheet, grade_scores_queryset)
    write_scores_hps_written(sheet, grade_scores_queryset)
    write_scores_hps_performance(sheet, grade_scores_queryset)
    write_scores_hps_quarterly(sheet, grade_scores_queryset)
    write_written_works_scores(sheet, grade_scores_queryset)
    write_performance_tasks_scores(sheet, grade_scores_queryset)
    write_quarterly_assessment_scores(sheet, grade_scores_queryset)
    write_initial_grade(sheet, grade_scores_queryset)
    write_transmuted_grade(sheet, grade_scores_queryset)
    write_school_info(sheet, school_info)

        # Save changes to the workbook
    workbook.save(copied_file_path)

        # Prepare HTTP response with updated Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ClassRecord_{grade_name}_{section_name}_AllSubjects_{quarter_name}_{timestamp}.xlsx'

    with open(copied_file_path, 'rb') as excel_file:
            response.write(excel_file.read())

    return response


def generate_summary_for_grades_4_to_6(request, grade, section, subject, quarter):
    # Query the GradeScores model
    grade_scores_queryset = GradeScores.objects.filter(
        class_record__grade=grade,
        class_record__section=section,
        class_record__subject=subject,
    )

    quarter_name = quarter
    selected_quarter = quarter

    school_info = SchoolInformation.objects.all()

    # Initialize lists to hold data for each subject
    quarter_data = {}
    for grade_score in grade_scores_queryset:
        class_record_name = grade_score.class_record.name
        teacher_name = f"{grade_score.class_record.teacher.user.first_name} {grade_score.class_record.teacher.user.last_name}".upper()

        if quarter_name not in quarter_data:
            quarter_data[quarter_name] = {'data': [], 'gradescores': []}
        
        # Append the grade score and its corresponding GradeScores object to the dictionary
        quarter_data[quarter_name]['data'].append((class_record_name, teacher_name))
        quarter_data[quarter_name]['gradescores'].append(grade_score)

    user = request.user
    action = f'{user} generate a Summary of Quarterly Grades for {grade} {section} in {subject}'
    details = f'{user} generate a Summary of Quarterly Grades for {grade} {section} in {subject} in the system.'
    log_activity(user, action, details)

    if grade in ["Grade 4", "Grade 5", "Grade 6"]:
        if subject == "MATHEMATICS":
            excel_file_name = "GRADE 4-6_MATHEMATICS.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "ARALING PANLIPUNAN":
            excel_file_name = "GRADE 4-6_ARALING PANLIPUNAN.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "EDUKASYON SA PAGPAPAKATAO":
            excel_file_name = "GRADE 4-6_EDUKASYON SA PAGPAPAKATAO.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "EPP":
            excel_file_name = "GRADE 4-6_EDUKASYONG PANTAHANAN AT PANGKABUHAYAN.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "ENGLISH":
            excel_file_name = "GRADE 4-6_ENGLISH.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "FILIPINO":
            excel_file_name = "GRADE 4-6_FILIPINO.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')

        elif subject == "SCIENCE":
            excel_file_name = "GRADE 4-6_SCIENCE.xlsx"
            media_directory = os.path.join(settings.MEDIA_ROOT, 'classrecord', 'GRADE-4-6_E-Class-Record-Templates-1')
    else:
        # Handle other grades
        pass

    original_file_path = os.path.join(media_directory, excel_file_name)

    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-sf9')
    if not os.path.exists(created_directory):
        os.makedirs(created_directory)

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'SUMMARY OF QUARTERLY GRADES {timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

    # Open the copied SF9 Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

    sheet_input_data = 'INPUT DATA'
    sheet_input = workbook[sheet_input_data]

    for quarter_info in quarter_data.values():
        # Extract data and GradeScores objects for the current subject
        data = quarter_info['data']
        grade_scores = quarter_info['gradescores']

        # Initialize dictionary to hold data for each quarter
        quarter_sheets = {'1st Quarter': [], '2nd Quarter': [], '3rd Quarter': [], '4th Quarter': []}

        # Organize grade scores by quarter
        for grade_score in grade_scores:
            quarter_name = grade_score.class_record.quarters  # Assuming 'quarter' field exists in ClassRecord model
            quarter_sheets[quarter_name].append(grade_score)

        # Map quarter names to desired sheet names
        quarter_sheet_mapping = {
            "1st Quarter": 'Q1',
            "2nd Quarter": 'Q2',
            "3rd Quarter": 'Q3',
            "4th Quarter": 'Q4'
        }

        print(selected_quarter)
        # Process data for each quarter
        for quarter_name, desired_sheet_name in quarter_sheet_mapping.items():
            if quarter_name <= selected_quarter:
                # Select the desired sheet
                sheet = workbook[desired_sheet_name]

                # Write data to the selected sheet
                write_student_names(sheet_input, quarter_sheets[quarter_name])
                write_scores_hps_written(sheet, quarter_sheets[quarter_name])
                write_scores_hps_performance(sheet, quarter_sheets[quarter_name])
                write_scores_hps_quarterly(sheet, quarter_sheets[quarter_name])
                write_written_works_scores(sheet, quarter_sheets[quarter_name])
                write_performance_tasks_scores(sheet, quarter_sheets[quarter_name])
                write_quarterly_assessment_scores(sheet, quarter_sheets[quarter_name])
                write_initial_grade(sheet, quarter_sheets[quarter_name])
                write_transmuted_grade(sheet, quarter_sheets[quarter_name])

# Now the data will be printed on the corresponding quarter's sheet based on the actual quarter each data belongs to.

            
    write_student_names(sheet_input, grade_scores)
    write_school_info(sheet_input, school_info, grade, section, teacher_name)

    # Save the changes to the SF9 workbook
    workbook.save(copied_file_path)

    # Create an HTTP response with the updated SF9 Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'SUMMARY OF QUARTERLY GRADES ({quarter}) {grade} - {section} {excel_file_name}_{timestamp}.xlsx'
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

    with open(copied_file_path, 'rb') as excel_file:
        response.write(excel_file.read())

    return response

    # except Exception as e:
    #     # Handle exceptions, such as a corrupted file
    #     return HttpResponse(f"An error occurred: {e}")
  

def generate_excel_for_sf9(request, student_id):
    # Retrieve the student object
    student = get_object_or_404(Student, id=student_id)

    user = request.user
    teacher = request.user.teacher
    teacher_name = f"{teacher.user.first_name} {teacher.user.middle_ini}. {teacher.user.last_name}" 

    # Query for other related objects
    advisory_class = AdvisoryClass.objects.filter(student=student).first()
    general_average = GeneralAverage.objects.filter(student=student).first()
    attendance_record = AttendanceRecord.objects.filter(student=student).first()
    learners_observation = LearnersObservation.objects.filter(student=student).first()
    school_info =  SchoolInformation.objects.all()

    # Check if any required records are missing
    if advisory_class is None:
        return render(request, 'teacher_template/adviserTeacher/generate_excel_for_sf9.html', {'advisory_class_missing': True})
    student_name = student.name
    grade_name = student.grade
    section_name = student.section

    user = request.user
    action = f'{user} generate the SF9 of student "{student_name}" in {grade_name} {section_name}'
    details = f'{user} generate the SF9 of student "{student_name}" in {grade_name} {section_name}.'
    log_activity(user, action, details)

    # Proceed with generating the Excel file
    student_name = student.name
    grade_name = student.grade
    section_name = student.section

    # Original file path for SF9 template
    excel_file_name = "SF9.xlsx"

    # Define directories
    media_directory = os.path.join(settings.MEDIA_ROOT, 'excel-files')
    created_directory = os.path.join(settings.MEDIA_ROOT, 'created-sf9')

    # Create directories if they don't exist
    if not os.path.exists(media_directory):
        os.makedirs(media_directory)
    if not os.path.exists(created_directory):
        os.makedirs(created_directory)

    # Generate a timestamp for the copy
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')

    # Create the path for the original file
    original_file_path = os.path.join(media_directory, excel_file_name)

    # Create a path for the copied file with a timestamp
    copied_file_path = os.path.join(created_directory, f'SF9 {student_name}_{grade_name}_{section_name}_{timestamp}.xlsx')

    # Copy the Excel file
    shutil.copyfile(original_file_path, copied_file_path)

    # Open the copied SF9 Excel file
    workbook = openpyxl.load_workbook(copied_file_path)

    # Select the desired sheet (use the correct sheet name from the output)
    front_sheet_name = 'FRONT'
    front_sheet = workbook[front_sheet_name]
    back_sheet_name = 'BACK'
    back_sheet = workbook[back_sheet_name]

    # Write SF9-specific data using utility functions
    write_sf9_data(front_sheet, student)
    write_sf9_school_info(front_sheet, school_info, teacher_name)
    write_sf9_grades(back_sheet, advisory_class, general_average)
    write_sf9_attendance(front_sheet, attendance_record)
    write_sf9_total_attendance(front_sheet, attendance_record)
    write_sf9_learners_observation(back_sheet, learners_observation, 'quarter_1', 25)
    write_sf9_learners_observation(back_sheet, learners_observation, 'quarter_2', 26)
    write_sf9_learners_observation(back_sheet, learners_observation, 'quarter_3', 27)
    write_sf9_learners_observation(back_sheet, learners_observation, 'quarter_4', 28)

    # Save the changes to the SF9 workbook
    workbook.save(copied_file_path)

            # Create an HTTP response with the updated SF9 Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'SF9_{student_name}_{grade_name}_{section_name}_{timestamp}.xlsx'
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'

    with open(copied_file_path, 'rb') as excel_file:
                response.write(excel_file.read())

    return response

    # Convert Excel to PDF using libreoffice
  
    # return response

    # except Exception as e:
    #     # Handle exceptions, such as a corrupted file
    #     return HttpResponse(f"An error occurred: {e}")



@login_required
def generate_per_subject_view(request):
    # Retrieve the user and check if the user is a teacher
    user = request.user
    if hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        # Filter class records based on the teacher
        class_records = ClassRecord.objects.filter(teacher=teacher)

        # Keep track of unique grade and section combinations
        unique_combinations = set()
        unique_class_records = []

        # Iterate through class records to filter out duplicates
        for record in class_records:
            combination = (record.grade, record.section)
            # Check if the combination is unique
            if combination not in unique_combinations:
                unique_combinations.add(combination)
                unique_class_records.append(record)

        context = {
            'class_records': unique_class_records,
        }

        return render(request, 'teacher_template/adviserTeacher/generate_per_subject.html', context)
    
    else:
        # Handle the case where the user is not a teacher
        return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")


@login_required
def generate_per_all_subject_view(request):
    # Retrieve the user and check if the user is a teacher
    user = request.user
    if hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        # Filter class records based on the teacher
        class_records = ClassRecord.objects.filter(teacher=teacher)

        # Create a dictionary to store data organized by quarter
        quarters_data = {}

        # Iterate through class records to organize data by quarter
        for record in class_records:
            quarter_data = quarters_data.setdefault(record.quarters, {'subjects': set(), 'grade_sections': set()})
            quarter_data['grade_sections'].add((record.grade, record.section, record.school_year))
            quarter_data['subjects'].add(record.subject)

        # Sort subjects alphabetically
        for quarter_data in quarters_data.values():
            quarter_data['subjects'] = sorted(quarter_data['subjects'])

        print(quarters_data)
        context = {
            'quarters_data': quarters_data,
        }

        return render(request, 'teacher_template/adviserTeacher/generate_per_all_subject.html', context)
    
    else:
        # Handle the case where the user is not a teacher
        return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")
    
def generate_grade_section_list(request):
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        grade = request.GET.get('grade')
        section = request.GET.get('section')
        
        # Filter class records based on the teacher
        class_records = ClassRecord.objects.filter(teacher=teacher, grade=grade, section=section)
        print(class_records)

        # Fetch distinct subjects based on grade and section
        subjects = class_records.values_list('subject', flat=True).distinct()

        for class_record in class_records:
            subject = class_record.subject

    
        context = {
            'grade': grade,
            'section': section,
            'class_records': class_records,
            'subjects': subjects,
        }

        return render(request, 'teacher_template/adviserTeacher/generate_grade_section_list.html', context)
    else:
        # Handle the case where the user is not a teacher
        return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")
    
@login_required
def generate_final_grade_view(request):
    # Retrieve the user and check if the user is a teacher
    user = request.user
    if hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        # Filter class records based on the teacher
        class_records = AdvisoryClass.objects.filter(teacher=teacher)

        # Create a set to store grade_sections
        grade_sections = set()

        # Iterate through class records to organize data by grade and section
        for record in class_records:
            grade_sections.add((record.student.grade, record.student.section, record.student.school_year))

        # Convert the set to a sorted list
        grade_sections = sorted(grade_sections)

        context = {
            'grade_sections': grade_sections,
        }

        return render(request, 'teacher_template/adviserTeacher/generate_final_grade.html', context)
    
    else:
        # Handle the case where the user is not a teacher
        return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")