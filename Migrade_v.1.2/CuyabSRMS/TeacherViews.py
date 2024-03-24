
import json
import os
import io
import re
from dateutil import parser
from requests import request
import base64
from CuyabSRMS.utils import transmuted_grade, log_activity
from .utils import log_activity
from django import forms
import openpyxl
from django.contrib import messages
from .models import AdvisoryClass, Grade, GradeScores, SchoolInformation, Section, Student, Teacher, Subject, Quarters, ClassRecord, FinalGrade, GeneralAverage, QuarterlyGrades
from .models import AdvisoryClass, Grade, GradeScores, Section, Student, Teacher, Subject, Quarters, ClassRecord, FinalGrade, GeneralAverage, QuarterlyGrades, AttendanceRecord, LearnersObservation, CoreValues
from django.contrib.auth import get_user_model  # Add this import statement
from django.urls import reverse
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect
import pandas as pd
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.errors import HttpError
import logging
from django.views.decorators.csrf import csrf_exempt
from django.core.serializers.json import DjangoJSONEncoder
from django.conf import settings
from django.db import IntegrityError
from django.contrib import messages
#OCR
from .forms import DocumentUploadForm, DocumentBatchUploadForm, CoreValuesForm, BehaviorStatementForm, LearnersObservationForm
from .models import ProcessedDocument, ExtractedData, Section, BehaviorStatement
from google.cloud import documentai_v1beta3 as documentai
from django.shortcuts import render
from .views import *
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.template.loader import render_to_string
from django.db.models import Avg
from django.template import RequestContext
#Grade
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.core.files.uploadedfile import TemporaryUploadedFile
import openpyxl
from django.http import HttpResponseForbidden
from django.utils import timezone
#Grade
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404
from .forms import GradeScoresForm
from django.views.decorators.http import require_POST
from django.utils import translation
from django.db import transaction
from statistics import mean
from django.core.exceptions import MultipleObjectsReturned
from openpyxl import load_workbook
from django.utils.timezone import now
from django.core.exceptions import MultipleObjectsReturned
import logging
import requests
from bs4 import BeautifulSoup
import os
from dotenv import load_dotenv
load_dotenv()
from json import loads as json_loads
from django.db.models import Q
import datetime
from datetime import datetime
from google.cloud import documentai_v1beta3 as documentai
from google.api_core.client_options import ClientOptions
import calendar
from calendar import month_name


def faq_view(request):
    # Dito mo ilalagay ang iyong logic para sa FAQ page
    return render(request, 'teacher_template/adviserTeacher/faq.html')  #saka etooo
def doc_view(request):
    # Dito mo ilalagay ang iyong logic para sa FAQ page
    return render(request, 'teacher_template/adviserTeacher/documentation.html')  #saka etooo
@login_required
def home_teacher(request):
    announcements = Announcement.objects.all()
    return render(request, 'teacher_template/home_teacher.html', {'announcements': announcements})
@login_required
def upload_adviser_teacher(request):
    return render(request, 'teacher_template/adviserTeacher/upload.html')
@login_required
def new_classrecord(request):
        return render(request, 'teacher_template/adviserTeacher/new_classrecord.html')
@login_required
def classes(request):
        return render(request, 'teacher_template/adviserTeacher/classes.html')




# adviser
@login_required
def home_adviser_teacher(request):
    url = "https://www.deped.gov.ph/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36"
    }

    response = requests.get(url, headers=headers)
    recent_posts_data = []
    recent_deped_memoranda = []

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')

        # Recent Posts
        recent_posts_container = soup.find('aside', id='panel-bottom-1')
        if recent_posts_container:
            posts = recent_posts_container.find('ul', id='lcp_instance_listcategorypostswidget-2').find_all('li')
            for post in posts:
                title = post.find('a').text.strip()
                link = post.find('a')['href']
                recent_posts_data.append({'title': title, 'link': link})

        # Recent DepEd Memoranda
        recent_deped_memoranda_container = soup.find('aside', id='panel-bottom-2')
        if recent_deped_memoranda_container:
            posts = recent_deped_memoranda_container.find('ul', id='lcp_instance_listcategorypostswidget-3').find_all('li')
            for post in posts:
                title = post.find('a').text.strip()
                link = post.find('a')['href']
                recent_deped_memoranda.append({'title': title, 'link': link})

    return render(request, 'teacher_template/adviserTeacher/home_adviser_teacher.html', {'recent_posts_data': recent_posts_data, 'recent_deped_memoranda': recent_deped_memoranda})
@login_required
def dashboard(request):
     # Get the currently logged-in teacher
    teacher = request.user.teacher

    # Retrieve the related section
    section = teacher.sections.first()  # Assuming a teacher can have multiple sections

    if section:
        # Get the grade from the related section
        grade = section.grade
    else:
        # Handle the case where no section is related to the teacher
        grade = None

    # Filter students based on the teacher's ID
    students = Student.objects.filter(teacher=teacher)

    # Count the number of students
    num_students = students.count()

    # Count the number of male (M) and female (F) students
    num_male_students = students.filter(sex='M').count()
    num_female_students = students.filter(sex='F').count()

    context = {
        'num_students': num_students,
        'grade': grade,  # Include the grade in the context
        'section': section,
        'num_male_students': num_male_students,
        'num_female_students': num_female_students,
    }
    return render(request, 'teacher_template/home_teacher.html', context)

# Your combined view
def process_google_sheet(spreadsheet_id, sheet_name):
    # Read environment variables
    project_id = os.getenv("SHEET_PROJECT_ID")
    private_key_id = os.getenv("SHEET_PRIVATE_KEY_ID")
    private_key = os.getenv("SHEET_PRIVATE_KEY").replace('\\n', '\n')  # Replace escaped newlines
    client_email = os.getenv("SHEET_CLIENT_EMAIL")
    client_id = os.getenv("SHEET_CLIENT_ID")
    token_uri = "https://oauth2.googleapis.com/token"  # Provide the token_uri
    # Construct credentials from environment variables
    credentials = service_account.Credentials.from_service_account_info({
        "type": "service_account",
        "project_id": project_id,
        "private_key_id": private_key_id,
        "private_key": private_key,
        "client_email": client_email,
        "client_id": client_id,
        "token_uri": token_uri  # Include the token_uri field
    })

    # Build the service
    service = build('sheets', 'v4', credentials=credentials)

    try:
        def get_sheet_values(spreadsheet_id, start_range, end_range):
            result = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=f"{start_range}:{end_range}",
            ).execute()
            values = result.get('values', [])
            return values

        def get_rows(sheet_values):
            rows_list = []

            for i in range(5):
                if i < len(sheet_values):
                    row = sheet_values[i]
                    non_empty_values = [value.strip() for value in row if value.strip() != '']
                    rows_list.append(non_empty_values)
                else:
                    rows_list.append([])

            return rows_list

        def find_lrn_and_store_as_dict(sheet_values):
            lrn_row_index = None
            lrn_col_index = None
            lrn_data = {}

            for i, row in enumerate(sheet_values):
                if "LRN" in row:
                    lrn_row_index = i
                    lrn_col_index = row.index("LRN")
                    break

            if lrn_row_index is not None:
                current_lrn = None

                for i in range(lrn_row_index + 1, len(sheet_values)):
                    row = sheet_values[i]
                    if lrn_col_index < len(row):
                        lrn_value = row[lrn_col_index]
                        if len(lrn_value) == 12 and lrn_value.isdigit():
                            current_lrn = lrn_value
                            lrn_data[current_lrn] = []
                        else:
                            current_lrn = None

                    if current_lrn:
                        non_empty_fields = [field for field in row if field.strip() != '']
                        lrn_data[current_lrn].append(non_empty_fields)

            return lrn_data

        # Get the values from the existing sheet
        existing_sheet_values = get_sheet_values(spreadsheet_id, "A1", "ZZ1000")

        if existing_sheet_values:
            rows_list = get_rows(existing_sheet_values)
            lrn_data = find_lrn_and_store_as_dict(existing_sheet_values)

            # Extract key-value pairs from the first few rows of the sheet
            terms_to_find = ["School ID", "School Name", "Division", "District", "School Year", "Grade Level", "Section", "Age"]
            key_value_pairs = {}

            for i, row in enumerate(rows_list, start=1):
                # Inside the process_google_sheet function
                for term in terms_to_find:
                    if term in row:
                        index = row.index(term)
                        if index + 1 < len(row):
                            key_value_pairs[term.replace(" ", "_")] = row[index + 1]
                        else:
                            key_value_pairs[term.replace(" ", "_")] = None


            return {'lrn_data': lrn_data, 'key_value_pairs': key_value_pairs}
        else:
            return None
    except Exception as e:
        print(f"An error occurred in process_google_sheet: {e}")
        return None

def get_sections(request):
    grade_id = request.GET.get('grade_id')
    
    # Query your database to get sections for the selected grade
    sections = Section.objects.filter(grade_id=grade_id)

    # Serialize the sections into a JSON response
    sections_data = [{'id': section.id, 'name': section.name} for section in sections]

    return JsonResponse({'sections': sections_data})

def get_sheet_values(sheet, start_range, end_range):
    values = []
    for row in sheet.iter_rows(min_row=start_range[0], min_col=start_range[1], max_row=end_range[0], max_col=end_range[1]):
        row_values = [cell.value for cell in row]
        values.append(row_values)
    return values

def find_lrn_and_store_as_dict(sheet_values):
    lrn_row_index = None
    lrn_col_index = None
    lrn_data = {}

    for i, row in enumerate(sheet_values):
        if "LRN" in row:
            lrn_row_index = i
            lrn_col_index = row.index("LRN")
            break

    if lrn_row_index is not None:
        current_lrn = None

        for i in range(lrn_row_index + 1, len(sheet_values)):
            row = sheet_values[i]
            if lrn_col_index < len(row):
                lrn_value = row[lrn_col_index]
                if len(str(lrn_value)) == 12 and str(lrn_value).isdigit():
                    current_lrn = lrn_value
                    lrn_data[current_lrn] = []
                else:
                    current_lrn = None

                if current_lrn:
                    non_empty_fields = [field for field in row if field is not None and str(field).strip() != '']
                    lrn_data[current_lrn].append(non_empty_fields)

    return lrn_data

def get_rows(sheet_values):
    rows_list = list(sheet_values)[:5]  # Convert the generator to a list and take the first 5 rows

    for i in range(5 - len(rows_list)):
        rows_list.append([])

    # Remove empty fields from each row and strip white spaces
    non_empty_rows = [[str(cell.value).strip() for cell in row if cell.value is not None] for row in rows_list]

    return non_empty_rows

def process_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        existing_sheet_name = workbook.sheetnames[0]  # Use the first sheet
        existing_sheet = workbook[existing_sheet_name]
        existing_sheet_values = get_sheet_values(existing_sheet, (1, 1), (existing_sheet.max_row, existing_sheet.max_column))
        row_list = get_rows(existing_sheet.iter_rows())

        if existing_sheet_values:
            lrn_data = find_lrn_and_store_as_dict(existing_sheet_values)

            # Extract key-value pairs from the first few rows of the sheet
            terms_to_find = ["School ID", "School Name", "Division", "District", "School Year", "Grade Level", "Grade", "Section", "Age"]
            key_value_pairs = {}

            for i, row in enumerate(row_list, start=1):
                for term in terms_to_find:
                    if term in row:
                        index = row.index(term)
                        if index + 1 < len(row):
                            key_value_pairs[term.replace(" ", "_")] = row[index + 1]
                        else:
                            key_value_pairs[term.replace(" ", "_")] = None

            return {'lrn_data': lrn_data, 'key_value_pairs': key_value_pairs}
        else:
            return None
    except Exception as e:
        print(f"An error occurred in process_excel_file: {e}")
        return None
    
def upload(request):
    if request.method == 'POST':
        google_sheet_link = request.POST.get('google_sheet_link')
        excel_file = request.FILES.get('excel_file')

        if google_sheet_link:
            # Process Google Sheet
            spreadsheet_id_match = re.search(r'/d/([a-zA-Z0-9-_]+)', google_sheet_link)

            if spreadsheet_id_match:
                spreadsheet_id = spreadsheet_id_match.group(1)
                sheet_name_match = re.search(r'#gid=([0-9]+)', google_sheet_link)
                sheet_name = sheet_name_match.group(1) if sheet_name_match else None

                result_data = process_google_sheet(spreadsheet_id, sheet_name)

                if result_data is not None:
                    return render(request, 'teacher_template/adviserTeacher/tempo_newupload.html', result_data)
                else:
                    messages.error(request, "Failed to process the Google Sheet")
            else:
                messages.error(request, "Invalid Google Sheet link")
        elif excel_file:
            # Save the uploaded Excel file temporarily
            temp_file_path = 'temp.xlsx'
            valid_excel_extensions = ['.xls', '.xlsx']
            file_extension = os.path.splitext(excel_file.name)[1].lower()

            if file_extension not in valid_excel_extensions:
                messages.error(request, "Invalid Excel file. Please upload a valid Excel file.")
                return render(request, 'teacher_template/adviserTeacher/upload.html')

            # If the file is in .xls format, convert it to .xlsx
            if file_extension == '.xls':
                df = pd.read_excel(excel_file)
                df.to_excel(temp_file_path, index=False)
            else:
                with open(temp_file_path, 'wb') as temp_file:
                    for chunk in excel_file.chunks():
                        temp_file.write(chunk)

            # Process Excel file
            result_data = process_excel_file(temp_file_path)

            if result_data is not None:
                return render(request, 'teacher_template/adviserTeacher/tempo_newupload.html', result_data)
            else:
                messages.error(request, "Failed to process the Excel File")

        else:
            messages.error(request, "Invalid file")

    return render(request, 'teacher_template/adviserTeacher/tempo_newupload.html')


@csrf_exempt
@login_required
def save_json_data(request):
    if request.method == 'POST':
        
        if not hasattr(request.user, 'teacher'):
            response_data = {'message': 'User is not a teacher.'}
            return JsonResponse(response_data, status=403)

        try:
            received_data = json.loads(request.body)
            teacher = request.user.teacher  # Get the currently logged-in teacher

            # Get the data from the request
            school_id = received_data.get('school_id', '')
            district = received_data.get('district', '')
            division = received_data.get('division', '')
            school_name = received_data.get('school_name', '')
            school_year = received_data.get('school_year', '')
            grade_name = received_data.get('grade', '')
            section_name = received_data.get('section', '')
            age = received_data.get('age', '')
            class_type_data = received_data.get('classType', '')   # New field for class type

            subject = ''
            if class_type_data == 'Advisory':
                subject = 'Advisory Class'
            elif class_type_data == 'Subject':
                subject = 'Subject Class'
            elif class_type_data == 'advisory_Subject_g1' or class_type_data == 'advisory_Subject_g4':
                subject = 'Advisory Class, Subject Class'

            teacher_id = teacher.id
            class_type = {teacher_id: subject}

            user = request.user
            action = f'{user} create a Class {grade_name} {section_name}'
            details = f'{user} created a Class named {grade_name} {section_name} in the system.'
            log_activity(user, action, details)

            logs = user, action, details    
            grade, _ = Grade.objects.get_or_create(name=grade_name)
            section, _ = Section.objects.get_or_create(name=section_name, grade=grade)


            if 'Advisory Class, Subject Class' in subject:
                existing_asc = Student.objects.filter(
                    Q(grade=grade_name) &
                    Q(section=section_name) &
                    Q(school_year=school_year) &
                    Q(class_type__icontains="Advisory Class, Subject Class")
                )
                
                if existing_asc.exists():
                    return JsonResponse({'status': 'error', 'message': f"Advisory class already exists for this {grade_name} {section_name} - School Year: {school_year}."})

                existing_subject = Student.objects.filter(
                    Q(grade=grade_name) &
                    Q(section=section_name) &
                    Q(school_year=school_year) &
                   Q(class_type__contains={teacher_id: "Subject Class"})
                )
                
                if existing_subject.exists():
                    return JsonResponse({'status': 'error', 'message': f"Advisory class already exists for this {grade_name} {section_name} - School Year: {school_year}."})
                             
            elif 'Subject Class' in subject:
                existing_subject_class = Student.objects.filter(
                    Q(grade=grade_name) &
                    Q(section=section_name) &
                    Q(school_year=school_year) &
                    Q(class_type__contains={teacher_id: "Subject Class"})
                )

                # Check if a subject class already exists
                if existing_subject_class.exists():
                    return JsonResponse({'status': 'error', 'message': f"Subject class already exists for this {grade_name} {section_name} - School Year: {school_year}."})

            elif 'Advisory Class' in subject:
                existing_sub = Student.objects.filter(
                    Q(grade=grade_name) &
                    Q(section=section_name) &
                    Q(school_year=school_year) &
                    Q(class_type__icontains="Subject Class")
                )
                
                # Check if a subject class already exists
                if existing_sub.exists():
                    return JsonResponse({'status': 'error', 'message': f"Subject class already exists for this {grade_name} {section_name} - School Year: {school_year}. If you wish to add Advisory Class, please refer to the Advisory Class and Subject Class and select the appropriate Class type."})

                existing_adv = Student.objects.filter(
                    Q(grade=grade_name) &
                    Q(section=section_name) &
                    Q(school_year=school_year) &
                    (Q(class_type__icontains="Advisory Class") | Q(class_type__contains={teacher_id: "Advisory Class, Subject Class"}))
                )
                
                # Check if an advisory class already exists
                if existing_adv.exists():
                    return JsonResponse({'status': 'error', 'message': f"Advisory class already exists for this {grade_name} {section_name} - School Year: {school_year}."})

            
                            # Increment the total_students field for the respective section
            section.total_students += 1
            # Initialize or get the existing class_type dictionary for the section
            existing_class_type = section.class_type or {}

            # Check if the class_type already exists
 

            # Update the existing class_type with the new value
            existing_class_type.update(class_type)

            # Save the class_type on the Section object
            section.class_type = existing_class_type

            # Save the Section object
            section.save()

            for item in received_data['rows']:
                lrn = item.get('LRN')
                name = item.get('Name')
                sex = item.get('Sex')
                birthday = item.get('Birthday')
                age = item.get('Age')

                # Create or update the Grade object

                # Create or update the Section object

                # Initialize or get the existing grade_section dictionary
                existing_grade_section = teacher.grade_section or {}

                grade_section = f"{grade.name} {section.name}"

                # Save the grade_section in the Teacher model
                existing_grade_section[grade_section] = subject
                teacher.grade_section = existing_grade_section
                teacher.save()

                

                # Get or create the Student object based on LRN and teacher
                student, created = Student.objects.get_or_create(
                    lrn=lrn,
                    school_year=school_year,
                    defaults={
                        'name': name,
                        'sex': sex,
                        'age': age,
                        'birthday': birthday,
                        'lrn':lrn,
                        'school_id': school_id,
                        'district': district,
                        'division': division,
                        'school_name': school_name,
                        'grade': grade.name,
                        'section': section.name,
                        'class_type': class_type  # Save the class type on Student
                    }
                )
   
                if not created:
                    
                 
                    # Check if the existing student has the same school year
                    if student.school_year == school_year:
           
                            student.class_type =  existing_class_type
                            student.save()
                            
                    else:
                        # If LRN exists but different school year, create a new entry
                        student = Student.objects.create(
                            lrn=lrn,
                            name=name,
                            sex=sex,
                            age=age,
                            birthday=birthday,
                            school_id=school_id,
                            district=district,
                            division=division,
                            school_name=school_name,
                            school_year=school_year,
                            grade=grade.name,
                            section=section.name,
                            class_type=class_type
                        )


                # Save the associated objects before saving the student
                grade.save()
                student.save()

            response_data = {'status': 'success', 'message': 'The class has been successfully created'}
            return JsonResponse(response_data)
        
        except ValidationError as e:
        # Return a JSON response with the error message
            return JsonResponse({'error': str(e)})
        
    else:
        response_data = {'message': 'Method not allowed'}
        return JsonResponse(response_data, status=405)



def get_grades_and_sections(request):
    teacher = get_object_or_404(Teacher, user=request.user)  # Assuming you have a User associated with Teacher

    # Retrieve a list of grades and sections that have the specific teacher
    grades = Grade.objects.filter(sections__teacher=teacher).distinct().values('id', 'name')
    sections = Section.objects.filter(teacher=teacher).values('id', 'name')

    # Convert the grades and sections to a dictionary
    data = {
        'grades': list(grades),
        'sections': list(sections),
    }

    return JsonResponse(data)


def class_record(request):
    return render(request, 'teacher_template/adviserTeacher/class_record.html')


def get_grade_details(request):

    user = request.user
    teacher = user.teacher
    teacher_id = teacher.id
    selected_grade = request.GET.get('grade')
    selected_section = request.GET.get('section')

    teacher = Teacher.objects.get(user=user)
    grades = Student.objects.filter(Q(class_type__contains={teacher_id :"Advisory Class, Subject Class"}) | Q(class_type__contains={teacher_id :"Subject Class"})).values_list('grade', flat=True).distinct()
    sections = Student.objects.values_list('section', flat=True).distinct()
    subjects = Subject.objects.values_list('name', flat=True).distinct()
    quarters = Quarters.objects.values_list('quarters', flat=True).distinct()
    
    context = {
        'teacher': teacher,
        'grades': grades,
        'sections': sections,
        'subjects': subjects,
        'quarters': quarters,
        'selected_grade': selected_grade, 
        'selected_section': selected_section,
    }


    return render(request, 'teacher_template/adviserTeacher/new_classrecord.html', context)
   # Replace with the actual URL of your new_classrecord.html
   
def get_sections_classrecord(request):
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        teacher = request.user.teacher
        teacher_id = teacher.id
        grade_id = request.GET.get('grade_id')
        sections = Student.objects.filter(Q(grade=grade_id), Q(class_type__contains={teacher_id :"Advisory Class, Subject Class"}) | Q(class_type__contains={teacher_id :"Subject Class"}) ).values_list('section', flat=True).distinct()
        sections_list = list(sections)
        return JsonResponse({'sections': sections_list})
    else:
        return JsonResponse({'error': 'Invalid request'})

# views.py
def get_students_by_grade_and_section(request):
    if request.method == "POST":
        try:
            grade_name = request.POST.get("grade")
            section_name = request.POST.get("section")
            subject_name = request.POST.get("subject")
            quarter_name = request.POST.get("quarter")

            user = request.user

            all_school_info = SchoolInformation.objects.all()

            for school_info in all_school_info:
                school_year = school_info.school_year
            

            teacher = get_object_or_404(Teacher, user=user)

            # Use a more unique identifier for the class record name
            classrecord_name = f'{grade_name} - {section_name} - {subject_name} - {quarter_name}'

            classrecord = ClassRecord(
                name=classrecord_name,
                grade=grade_name,
                section=section_name,
                subject=subject_name,
                quarters=quarter_name,
                teacher=teacher,
                school_year=school_year,
            )

            classrecord.save()


            # Query the database to retrieve students based on the selected grade and section
            students = Student.objects.filter(grade=grade_name, section=section_name)

            subject = Subject.objects.get(name=subject_name)
  
            assessments = subject.assessment
            assessment_types = list(assessments.keys())
            # Create a dictionary to store scores for each student
            student_scores = {}

            # Loop through the students and retrieve their scores
            for student in students:
                # Assuming you have a field to store the student's name in the Student model
                student_name = student.name

                # Query the database to retrieve the scores for the current student
                # You should modify this query based on your data model
                scores = GradeScores.objects.filter(student=student)

                # Store the scores in the dictionary with the student's name as the key
                student_scores[student_name] = scores

            context = {
                'students': students,
                'subject_name': subject_name,
                'quarters': quarter_name,
                'grade_name': grade_name,
                'section_name': section_name,
                'student_scores': student_scores,
                'assessment': assessments,
                 'assessment_types': assessment_types,
                 'assessment_values': list(assessments.values()),
            }

            return render(request, 'teacher_template/adviserTeacher/class_record.html', context)
        
        except IntegrityError as e:
            error_message = 'Duplicate entry. The record already exists.'
            messages.error(request, error_message)
            messages.info(request, 'You are being redirected back to the previous page.')
            return redirect('get_grade_details')

    
    return render(request, 'teacher_template/adviserTeacher/home_adviser_teacher.html')




def calculate_grades(request):
    if request.method == "POST":
        grade_name = request.POST.get("hidden_grade")
        section_name = request.POST.get("hidden_section")
        subject_name = request.POST.get("hidden_subject")
        quarters_name = request.POST.get("hidden_quarter")
        subject = Subject.objects.get(name=subject_name)
        assessments = subject.assessment
        sanitized_assessment_types = list(assessments.keys())
        assessmentTypes = [assessmentType.replace(' ', '-').lower() for assessmentType in sanitized_assessment_types]

        scores_hps = {}  # Store scores for each assessment type

        for assessmentType in assessmentTypes:
            scores_hps[assessmentType] = {}
            # Extract scores for each assessment type
            scores_hps[assessmentType]['SCORES'] = [request.POST.get(f"max_{assessmentType}_{i}") for i in range(1, 11)]

            scores_hps[assessmentType]['TOTAL_HPS'] = request.POST.get(f"total_max_{assessmentType}")
            scores_hps[assessmentType]['WEIGHT'] = request.POST.get(f"weight_{assessmentType}")


        # Extract total scores for each assessment type
        total_hps = {
            assessmentType: request.POST.get(f"total_max_{assessmentType}")
            for assessmentType in assessmentTypes
        }
        weights = {}
        for assessmentType in assessmentTypes:
                weights[assessmentType] = request.POST.get(f"weight_{assessmentType}")

        class_record = ClassRecord.objects.get(grade=grade_name, section=section_name, subject=subject_name, quarters=quarters_name)
        class_record_name = class_record.name

        user = request.user
        action = f'{user} create a Classrecord "{class_record_name}"'
        details = f'{user} created a Classrecord named "{class_record_name}" in the system.'
        log_activity(user, action, details)

        logs = user, action, details    
  
        students = Student.objects.filter(grade=grade_name, section=section_name)

        for student in students:
            total_weighted_score = 0
            total_weight = 0
            scores_assessment = {}
            
            # Loop through each assessment type for the current student
            for assessmentType in assessmentTypes:
            
                total_score = 0
                total_max_score = 0
                scores_assessment[assessmentType] = {
                    'scores': [],
                    'total_score': total_score,
                    'percentage_score': None,
                    'total_weighted_score': None
                }
                
                # Retrieve scores and maximum scores for each component
                for i in range(1, 11):
            
                    score = request.POST.get(f"scores_{assessmentType}_{student.id}_{i}")
                    max_score = request.POST.get(f"max_{assessmentType}_{i}")
                    
                    # Check if score is None before converting to float and appending to scores list
                    if score is not None:
                        if score != '':
                            scores_assessment[assessmentType]['scores'].append(float(score))
                        else:
                            scores_assessment[assessmentType]['scores'].append(score)
                    else:
                        scores_assessment[assessmentType]['scores'].append(0)
                        
                    total_score += float(score) if score is not None and score.isnumeric() else 0
                    total_max_score += float(max_score) if max_score is not None and max_score.isnumeric() else 0
                    scores_assessment[assessmentType]['total_score'] = total_score
                
                # Extract weight for the current assessment type
                    weight_input = request.POST.get(f"weight_{assessmentType}")
                    if weight_input is not None:
                        weight_input = float(weight_input)
                    else:
                        weight_input = 0 
                    
                    # Perform calculations based on the assessment type
                    if total_max_score > 0:
                        percentage_score = round((total_score / total_max_score) * 100, 2)
                    else:
                        percentage_score = None if total_max_score == 0 else 0

                    weighted_score = (percentage_score / 100) * weight_input if percentage_score is not None else None
                    rounded_weighted_score = round(weighted_score, 2) if weighted_score is not None else None
                    

                    # Accumulate total weighted score and total weight for all assessment types
                    if weighted_score is not None:
                        total_weighted_score += weighted_score
                        total_weight += weight_input


                    scores_assessment[assessmentType]['percentage_score'] = percentage_score
                    scores_assessment[assessmentType]['total_weighted_score'] = weighted_score
            
            # Calculate initial grades for the current student after processing all assessment types
            if total_weight > 0:
                initial_grades = total_weighted_score / 10
            else:
                initial_grades = None  # Handle the case where total weight is 0 or not provided
            
            # Round off initial grades
            rounded_initial_grades = round(initial_grades, 2) if initial_grades is not None else None
            
            # Calculate the transmuted grades (if needed) based on the initial grades
            transmuted_grades = transmuted_grade(initial_grades) 
            
            # Round off transmuted grades
            rounded_transmuted_grades = round(transmuted_grades, 2) if transmuted_grades is not None else None


            grade_scores_data = {
                    'scores_hps': scores_hps,
                    'scores_per_assessment' : scores_assessment
                }
            
            # grade_scores_json = json.dumps(grade_scores_data)


            # Create a new GradeScores object and populate its fields
            grade_scores = GradeScores(
                student=student,
                class_record=class_record,
                initial_grades= rounded_initial_grades,
                transmuted_grades= rounded_transmuted_grades,
                grade_scores=grade_scores_data 
            )

            # Save the GradeScores object to the database
            grade_scores.save()


            # grade_scores = GradeScores.objects.filter(grade=grade_name, section=section_name)

            # context = {

            #     'grade_scores': grade_scores
            # }

            # return render(request, "teacher_template/adviserTeacher/display_classrecord.html", context)

        # Redirect to a success page or render a response as needed
        # return redirect('display_classrecord')
        return redirect('display_classrecord', class_record_id=class_record.id)

    return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")

def display_classrecord(request, class_record_id=None):
     # If class_record_id is provided, retrieve the ClassRecord object
    class_record = get_object_or_404(ClassRecord, id=class_record_id)
    subject_name = class_record.subject
    teacher = request.user.teacher
    teacher_id = teacher.id
    if teacher_id != class_record.teacher_id:
        return HttpResponseForbidden("You don't have permission to access this class record.")


    grade_scores = GradeScores.objects.filter(class_record=class_record)
    subject = Subject.objects.get(name=subject_name)
    assessments = subject.assessment

    assessment_types = list(assessments.keys())
    processed_types = []
    for assessment_type in assessment_types:
        processed_type = assessment_type.replace(' ', '-').lower()
        processed_types.append(processed_type)

    print(processed_types)

    assessment_type_processed = None

    context = {
            'class_record': class_record,
            'gradescores': grade_scores,
            'assessment_types': processed_types,
            'assessment_values': list(assessments.values()),
        }

    return render(request, 'teacher_template/adviserTeacher/display_classrecord.html', context)
    
    
def view_classrecord(request):
    
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        # Filter class records based on the teacher
        class_records = ClassRecord.objects.filter(teacher=teacher)

        context = {
            'class_records': class_records,
        }


        return render(request, 'teacher_template/adviserTeacher/view_classrecord.html', context)
    
    else:
        # Handle the case where the user is not a teacher
        return render(request, "teacher_template/adviserTeacher/home_adviser_teacher.html")
    
def display_students(request):
    user = request.user

    if user.user_type == 2:
        teacher = get_object_or_404(Teacher, user=user)
        teacher_id = str(teacher.id)  # Convert teacher id to string for comparison

        # Fetch distinct school years from the Student model
        unique_school_years = Student.objects.values_list('school_year', flat=True).distinct()

        # Reverse the order of unique_school_years
        unique_school_years = list(unique_school_years)
        unique_school_years.reverse()
        
        # Find the latest school year
        latest_school_year = max(unique_school_years, default=None)

        # Fetch the school year from the SchoolInformation model if available
        try:
            school_info = SchoolInformation.objects.latest('id')
            default_school_year = school_info.school_year
        except SchoolInformation.DoesNotExist:
            default_school_year = 'None'  # Set a default value if no school year is found

        # Use the latest school year if available, otherwise fallback to default
        default_school_year = latest_school_year or default_school_year

        # Get the school year from the request parameters or use the default value
        school_year = request.GET.get('school_year', default_school_year)
        
        # If the 'display_students' view is accessed without a specific school year, redirect to the same view with the latest school year
        if not request.GET.get('school_year'):
            return redirect(reverse('display_students') + f'?school_year={default_school_year}')
        
        # Filter students based on the teacher_id and school_year in class_type
        students = Student.objects.filter(class_type__has_key=teacher_id, school_year=school_year)
        unique_combinations = students.values('grade', 'section', 'class_type').distinct()
        class_type_list = []

        # Fetch the class type separately
        for combination in unique_combinations:
            class_type_dict = combination['class_type']
            if class_type_dict:  # Ensure class_type is not None
                if teacher_id in class_type_dict:  # Check if teacher_id exists in class_type_dict keys
                    class_type_value = class_type_dict[teacher_id]  # Get the value using the teacher_id key
                    combination['class_type'] = class_type_value
                    class_type_list.append(combination)

        context = {
            'teacher': teacher,
            'unique_grades_sections': class_type_list,
            'latest_school_year': latest_school_year,
            'default_school_year': default_school_year,
            'selected_school_year': school_year,  # Pass the selected school year to the template
            'unique_school_years': unique_school_years,  # Pass distinct school years to the template
        }
        return render(request, 'teacher_template/adviserTeacher/classes.html', context)

    return render(request, 'teacher_template/adviserTeacher/classes.html')
def toggle_class_type_function(student):
    # Toggle the class type for the given student
    if student.class_type == 'Advisory':
        student.class_type = 'Subject'
    else:
        student.class_type = 'Advisory'
    
    # Save the changes
    student.save()

    # Return the updated class type
    return student.class_type


def toggle_class_type(request):
    if request.method == 'POST':
        try:
            # Retrieve the POST data
            data = json.loads(request.body)

            # Extract the grade, section, and current_class_type from the data
            grade = data.get('grade')
            section = data.get('section')
            current_class_type = data.get('current_class_type')

            # Retrieve all students with the given grade and section
            students = Student.objects.filter(grade=grade, section=section)

            # Toggle the class type for each student
            for student in students:
                if current_class_type == 'Advisory':
                    student.class_type = 'Subject'
                else:
                    student.class_type = 'Advisory'

                # Save the changes
                student.save()

            response_data = {'message': 'Class type updated successfully'}
            return JsonResponse(response_data)

        except Exception as e:
            response_data = {'message': f'Error: {str(e)}'}
            return JsonResponse(response_data, status=400)

    response_data = {'message': 'Invalid request method'}
    return JsonResponse(response_data, status=405)

def sf9(request):
    # Query all students from the database
    teacher = request.user.teacher
    teacher_id = teacher.id
    students = Student.objects.filter(
        Q(class_type__contains={teacher_id: "Advisory Class"}) | Q(class_type__contains={teacher_id: "Advisory Class, Subject Class"}) 
    )
    # Pass the queryset to the template context
    context = {'all_students': students}

    # Render the template with the context
    return render(request, 'teacher_template/adviserTeacher/sf9.html', context)

@login_required
def delete_class(request, grade, section):
    user = request.user

    if user.user_type == 2:
        teacher = get_object_or_404(Teacher, user=user)
        students = Student.objects.filter(grade=grade, section=section)

        if students.exists():
            # Assuming you have some permission checks here before deleting
            students.delete()

            # Delete associated ClassRecord records
            ClassRecord.objects.filter(grade=grade, section=section, teacher=teacher).delete()

            # Delete the section
            section_instance = Section.objects.filter(grade__name=grade, name=section).first()
            if section_instance:
                section_instance.delete()

            # Remove the class type entry from the teacher's grade_section
            if teacher.grade_section:
                class_type_to_remove = f"{grade} {section}"
                if class_type_to_remove in teacher.grade_section:
                    del teacher.grade_section[class_type_to_remove]

                # Save the updated grade_section
                teacher.save()

            action = f'{user} delete a Class name {grade} {section}'
            details = f'{user} deleted a Class named {grade} {section} in the system.'
            log_activity(user, action, details)

            # Redirect to a different page after deletion
            return redirect('display_students')  # Replace with your actual view name
        else:
            # Redirect to a different page if no students found
            return redirect('display_students')  # Replace with your actual view name

    # If the user is not a teacher or if the permissions check fails
    return JsonResponse({'message': 'Unable to delete students. Permission denied.'}, status=403)

@login_required
def delete_class_subject(request, grade, section):
    user = request.user

    if user.user_type == 2:  # Assuming user_type 2 represents a teacher
        teacher = get_object_or_404(Teacher, user=user)
        students = Student.objects.filter(grade=grade, section=section)

        if students.exists():
            for student in students:
                class_type = student.class_type
                if str(teacher.id) in class_type:
                    if class_type.get(str(teacher.id)) == 'Advisory Class, Subject Class':
                        # Update the class_type to contain only 'Advisory Class'
                        class_type[str(teacher.id)] = 'Advisory Class'

                        if teacher.grade_section:
                            for key, value in teacher.grade_section.items():
                                if key == f"{grade} {section}":
                                    teacher.grade_section[key] = 'Advisory Class'
                                    break

                    else:
                        del class_type[str(teacher.id)]  # Delete the key associated with the teacher's ID

                    student.class_type = class_type
                    student.save()

            # Delete associated ClassRecord records
            ClassRecord.objects.filter(grade=grade, section=section, teacher=teacher).delete()

            # Delete the section
            section_instance = Section.objects.filter(grade__name=grade, name=section).first()
            if section_instance:
                # Update the class_type for the section
                class_type = section_instance.class_type
                if str(teacher.id) in class_type:
                    if class_type[str(teacher.id)] == 'Advisory Class, Subject Class':
                        # Update the class_type to contain only 'Advisory Class'
                        class_type[str(teacher.id)] = 'Advisory Class'

                    else:
                        del class_type[str(teacher.id)]  # Delete the key associated with the teacher's ID

                    section_instance.class_type = class_type
                    section_instance.save()

                # Delete the section if it has no associated students
                if section_instance.total_students == 0:
                    section_instance.delete()


            action = f'{user} deleted a Class named {grade} {section}'
            details = f'{user} deleted a Class named {grade} {section} in the system.'
            log_activity(user, action, details)

            teacher.save()

            return redirect('display_students')  # Redirect to your desired page after deletion
        else:
            return redirect('display_students')  # Redirect if no students found
    else:
        return JsonResponse({'message': 'Unable to delete students. Permission denied.'}, status=403)


def student_list_for_subject(request):
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher
        teacher_id = teacher.id

        grade = request.GET.get('grade')
        section = request.GET.get('section')
        class_type = request.GET.get('class_type')
        # quarter = request.GET.get('quarter')
        quarters = {"1st Quarter", "2nd Quarter", "3rd Quarter", "4th Quarter"}

        # Filter class records based on the teacher
        class_records = ClassRecord.objects.filter(teacher=teacher, grade=grade, section=section)

        # Fetch students based on grade and section
        students = Student.objects.filter(grade=grade, section=section)
  

        # Fetch distinct subjects based on grade and section
        subjects = ClassRecord.objects.filter(grade=grade, section=section, teacher=teacher).values('subject').distinct()
   

        # Initialize a dictionary to store highest initial grade and transmuted grades per subject
        subject_grades = {}

        # Loop through each subject
        for subject in subjects:
            subject_name = subject['subject']
            subject_students = students

            # Initialize a dictionary to store grades for each quarter
            subject_grades[subject_name] = {}

            # Fetch highest initial grade and transmuted grades for students in this subject
            for quarter in quarters:
                grades_for_quarter = []  # Initialize a list to store grades for the current quarter
                for student in subject_students:
                    # Retrieve the student's highest initial grade
                    highest_initial_grade = GradeScores.objects.filter(student=student, class_record__subject=subject_name).order_by('-initial_grades').first()

                    # Retrieve the student's transmuted grade for the specific subject and quarter
                    transmuted_grade = GradeScores.objects.filter(student=student, class_record__subject=subject_name, class_record__quarters=quarter).order_by('transmuted_grades').first()
                    if transmuted_grade:
                        quarter = transmuted_grade.class_record.quarters

                    # Determine remarks based on transmuted grade
                    remarks = determine_remarks(transmuted_grade.transmuted_grades) if transmuted_grade else 'No Grade'
                    status = determine_status(transmuted_grade.transmuted_grades) if transmuted_grade else 'No Grade'

                    # If both initial grade and transmuted grade exist, append the data to the grades_for_quarter list
                    if highest_initial_grade is not None and transmuted_grade is not None:
                        # Append the student's data to the grades_for_quarter list
                        grades_for_quarter.append({
                            'student_name': student.name,
                            'highest_initial_grade': highest_initial_grade.initial_grades,
                            'transmuted_grade': transmuted_grade.transmuted_grades,
                            'remarks': remarks,
                            'status': status,
                            'quarter': quarter  # Include the quarter information
                        })

                # Sort the grades for this quarter by highest initial grade
                grades_for_quarter.sort(key=lambda x: x['transmuted_grade'] if x['transmuted_grade'] is not None else 0, reverse=True)

                # Store the grades for this quarter in the subject_grades dictionary
                subject_grades[subject_name][quarter] = grades_for_quarter

        # print(subject_grades)

    context = {
        'grade': grade,
        'section': section,
        'class_type': class_type,
        'students': students,
        'class_records': class_records,
        'subjects': subjects,
        'subject_grades': subject_grades,
    }

    return render(request, 'teacher_template/adviserTeacher/student_list_for_subject.html', context)


def student_list_for_advisory(request):
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher
        teacher_id = teacher.id 

        grade = request.GET.get('grade')    
        section = request.GET.get('section')
        class_type = request.GET.get('class_type')
        quarter = request.GET.get('quarter', '1st Quarter')  # Default to 1st Quarter if not provided

        # Fetch advisory classes based on teacher, grade, and section
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section, teacher=teacher)


        unique_keys = set()  # Initialize an empty set here

        if advisory_classes.exists():
            for advisory_class in advisory_classes:
                grades_data = advisory_class.grades_data
                if grades_data:
                    for key, value in grades_data.items():
                        unique_keys.add((key, value.get('from_teacher_id')))  # Add (key, from_teacher_id) tuple to the set


        else:
            print("No AdvisoryClass objects found for the specified criteria")

      
        # Filter students based on the class type
        students = Student.objects.filter(grade=grade, section=section)
        students_filtered = []
        for student in students:
            class_type_json = student.class_type
            if class_type_json and str(teacher_id) in class_type_json and class_type_json[str(teacher_id)] == class_type:
                students_filtered.append(student)

        # Unique keys context
        unique_keys_context = list(unique_keys)

        # Filter quarterly grades based on the selected quarter
        quarterly_grades = QuarterlyGrades.objects.filter(student__grade=grade, student__section=section, quarter=quarter)

        # Prepare data to pass to the template
        data = []
        for index, grades in enumerate(quarterly_grades, start=1):
            student_name = grades.student.name
            subjects_grades = grades.grades
            average_score = subjects_grades.pop('average_score', None)
            subjects_data = [{'subject': subject, 'score': score} for subject, score in subjects_grades.items()]

            data.append({
                'no': index,
                'student_name': student_name,
                'subjects_data': subjects_data,
                'average_score': average_score,
            })
        
        students = AdvisoryClass.objects.filter(grade=grade, section=section)
        # Dictionary to store subject-wise grades and average score for each student
        subject_grades = {}
        quarter_mapping = {
            '1st Quarter': 'first_quarter',
            '2nd Quarter': 'second_quarter',
            '3rd Quarter': 'third_quarter',
            '4th Quarter': 'fourth_quarter',
        }

        # Fetch subject-wise grades for each student
        for student in students:
            grades_data = student.grades_data
            subject_grades[student.student.name] = {}
            grades = []  # List to store grades for calculating mean

            for subject, grades_info in grades_data.items():
                # Skip subjects 'MUSIC', 'ARTS', 'PE', and 'HEALTH'
                if subject in ['MUSIC', 'ARTS', 'PE', 'HEALTH']:
                    continue
                
                if quarter_mapping[quarter] in grades_info:
                    subject_grade = grades_info[quarter_mapping[quarter]]
         
                    subject_grade_str = str(subject_grade) if subject_grade is not None else ""
                    if subject_grade_str.strip():  # Check if the string is not empty after stripping whitespace
                        subject_grades[student.student.name][subject] = subject_grade_str
                        if subject_grade is not None:
                            grades.append(float(subject_grade))

            # Calculate average score
            if grades:
                subject_grades[student.student.name]['average_score'] = round(mean(grades), 2)
            else:
                subject_grades[student.student.name]['average_score'] = ""

            # Check if QuarterlyGrades entry already exists for this student and quarter
            existing_entry = QuarterlyGrades.objects.filter(student=student.student, quarter=quarter).first()
            if not existing_entry:
                # Save grades to QuarterlyGrades model
                QuarterlyGrades.objects.create(
                    student=student.student,
                    quarter=quarter,
                    grades=subject_grades[student.student.name]
                )
            elif existing_entry.grades != subject_grades[student.student.name]:
                # Update the existing entry if the grades are different
                existing_entry.grades = subject_grades[student.student.name]
                existing_entry.save()

        subjects = list(students.first().grades_data.keys()) if students else []

        students = Student.objects.filter(grade=grade, section=section)
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section)

        final_grades = []
        for student in students:
            student_data = {
                'id': student.id,
                'name': student.name,
                'grade': grade,
                'section': section,
                'subjects': [],
                'student': student
            }

            for advisory_class in advisory_classes.filter(student=student):
                grades_data = advisory_class.grades_data
                for subject, subject_info in grades_data.items():
                    if subject.upper() in ['MUSIC', 'ARTS', 'PE', 'HEALTH']:
                        continue
                    # Access grades data for each subject
                    subject_data = {
                        'subject': subject,
                        'quarter_grades': {
                            'first_quarter': subject_info.get('first_quarter', ''),
                            'second_quarter': subject_info.get('second_quarter', ''),
                            'third_quarter': subject_info.get('third_quarter', ''),
                            'fourth_quarter': subject_info.get('fourth_quarter', ''),
                            # Add more quarters if available
                        },
                        'final_grade': subject_info.get('final_grade', ''),
                        'teacher_name': subject_info.get('from_teacher_id', 'Unknown Teacher')
                    }
                    student_data['subjects'].append(subject_data)

            # Append student data to the final grades
            final_grades.append(student_data)
        
        # Compute the general average and save it for each student
        for student_data in final_grades:
            total_final_grade = 0
            num_subjects = len(student_data['subjects'])
            for subject_info in student_data['subjects']:
                final_grade = subject_info['final_grade']
                try:    
                    if final_grade is not None:
                        total_final_grade += float(final_grade) 
                except ValueError:
                    # Handle the case where final_grade is not a valid number
                    pass

            student_data['general_average'] = total_final_grade / num_subjects if num_subjects > 0 else 0
            save_general_average(student_data, grade, section)

        sorted_final_grades = sorted(final_grades, key=lambda x: x.get('general_average', 0), reverse=True)
        highest_per_quarter = {
            'first_quarter': [],
            'second_quarter': [],
            'third_quarter': [],
            'fourth_quarter': [],
        }

        # Populate data for each quarter
        for quarters in ['first_quarter', 'second_quarter', 'third_quarter', 'fourth_quarter']:
            sorted_students = []
            for student in final_grades:
                if 'subjects' in student and student['subjects']:  # Check if 'subjects' list exists and is not empty
                    quarter_grades = student['subjects'][0]['quarter_grades'].get(quarters)
                    if quarter_grades is not None and quarter_grades != '':
                        sorted_students.append(student)
            sorted_students = sorted(sorted_students, key=lambda x: float(x['subjects'][0]['quarter_grades'].get(quarters, '0') or '0'), reverse=True)
            highest_per_quarter[quarters] = sorted_students



            general_averages = GeneralAverage.objects.filter(grade=grade, section=section)

        # Sort GeneralAverage instances based on the general average from highest to lowest
            sorted_general_averages = general_averages.order_by('-general_average')

        user = request.user
        if hasattr(user, 'teacher'):
            # Retrieve the teacher associated with the user
            teacher = user.teacher

            # Filter class records based on the teacher
            class_records = AdvisoryClass.objects.filter(teacher=teacher)

            # Keep track of unique grade and section combinations
            unique_combinations = set()
            unique_class_records = []

     

            # Iterate through class records to filter out duplicates
            for record in class_records:
                combination = (record.grade, record.section)
                # Check if the combination is unique
                if combination not in unique_combinations:
                    unique_combinations.add(combination)
                    unique_class_records.append(record)

            
        

        context = {
            'grade': grade,
            'section': section,
            'unique_keys': unique_keys_context,
            'students': students_filtered,
            'class_type': class_type,
            'data': data,
            'quarter': quarter,
            'final_grades': sorted_final_grades,
            'highest_per_quarter': highest_per_quarter,
            'general_averages': sorted_general_averages,
            'class_records': unique_class_records,

         
        }

        return render(request, 'teacher_template/adviserTeacher/student_list_for_advisory.html', context)
    
def advisory_quarterly_grades(request):
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher
        teacher_id = teacher.id 

        grade = request.GET.get('grade')    
        section = request.GET.get('section')
        class_type = request.GET.get('class_type')
        quarter = request.GET.get('quarter', '1st Quarter')  # Default to 1st Quarter if not provided

        # Fetch advisory classes based on teacher, grade, and section
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section, teacher=teacher)


        unique_keys = set()  # Initialize an empty set here

        if advisory_classes.exists():
            for advisory_class in advisory_classes:
                grades_data = advisory_class.grades_data
                if grades_data:
                    for key, value in grades_data.items():
                        unique_keys.add((key, value.get('from_teacher_id')))  # Add (key, from_teacher_id) tuple to the set

        else:
            print("No AdvisoryClass objects found for the specified criteria")

      
        # Filter students based on the class type
        students = Student.objects.filter(grade=grade, section=section)
        students_filtered = []
        for student in students:
            class_type_json = student.class_type
            if class_type_json and str(teacher_id) in class_type_json and class_type_json[str(teacher_id)] == class_type:
                students_filtered.append(student)

        # Unique keys context
        unique_keys_context = list(unique_keys)

        # Filter quarterly grades based on the selected quarter
        quarterly_grades = QuarterlyGrades.objects.filter(student__grade=grade, student__section=section, quarter=quarter)

        # Prepare data to pass to the template
        data = []
        for index, grades in enumerate(quarterly_grades, start=1):
            student_name = grades.student.name
            subjects_grades = grades.grades
            average_score = subjects_grades.pop('average_score', None)
            subjects_data = [{'subject': subject, 'score': score} for subject, score in subjects_grades.items()]

            data.append({
                'no': index,
                'student_name': student_name,
                'subjects_data': subjects_data,
                'average_score': average_score,
            })
        
        students = AdvisoryClass.objects.filter(grade=grade, section=section)
        # Dictionary to store subject-wise grades and average score for each student
        subject_grades = {}
        quarter_mapping = {
            '1st Quarter': 'first_quarter',
            '2nd Quarter': 'second_quarter',
            '3rd Quarter': 'third_quarter',
            '4th Quarter': 'fourth_quarter',
        }

        # Fetch subject-wise grades for each student
        for student in students:
            grades_data = student.grades_data
            subject_grades[student.student.name] = {}
            grades = []  # List to store grades for calculating mean

            for subject, grades_info in grades_data.items():
                # Skip subjects 'MUSIC', 'ARTS', 'PE', and 'HEALTH'
                if subject in ['MUSIC', 'ARTS', 'PE', 'HEALTH']:
                    continue
                
                if quarter_mapping[quarter] in grades_info:
                    subject_grade = grades_info[quarter_mapping[quarter]]
      
                    subject_grade_str = str(subject_grade) if subject_grade is not None else ""
                    if subject_grade_str.strip():  # Check if the string is not empty after stripping whitespace
                        subject_grades[student.student.name][subject] = subject_grade_str
                        if subject_grade is not None:
                            grades.append(float(subject_grade))

            # Calculate average score
            if grades:
                subject_grades[student.student.name]['average_score'] = round(mean(grades), 2)
            else:
                subject_grades[student.student.name]['average_score'] = None

            # Check if QuarterlyGrades entry already exists for this student and quarter
            existing_entry = QuarterlyGrades.objects.filter(student=student.student, quarter=quarter).first()
            if not existing_entry:
                # Save grades to QuarterlyGrades model
                QuarterlyGrades.objects.create(
                    student=student.student,
                    quarter=quarter,
                    grades=subject_grades[student.student.name]
                )
            elif existing_entry.grades != subject_grades[student.student.name]:
                # Update the existing entry if the grades are different
                existing_entry.grades = subject_grades[student.student.name]
                existing_entry.save()

        subjects = list(students.first().grades_data.keys()) if students else []

        students = Student.objects.filter(grade=grade, section=section)
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section)

        final_grades = []
        for student in students:
            student_data = {
                'id': student.id,
                'name': student.name,
                'grade': grade,
                'section': section,
                'subjects': [],
                'student': student
            }

            for advisory_class in advisory_classes.filter(student=student):
                grades_data = advisory_class.grades_data
                for subject, subject_info in grades_data.items():
                    # Access grades data for each subject
                    subject_data = {
                        'subject': subject,
                        'quarter_grades': {
                            'first_quarter': subject_info.get('first_quarter', ''),
                            'second_quarter': subject_info.get('second_quarter', ''),
                            'third_quarter': subject_info.get('third_quarter', ''),
                            'fourth_quarter': subject_info.get('fourth_quarter', ''),
                            # Add more quarters if available
                        },
                        'final_grade': subject_info.get('final_grade', ''),
                        'teacher_name': subject_info.get('from_teacher_id', 'Unknown Teacher')
                    }
                    student_data['subjects'].append(subject_data)

            # Append student data to the final grades
            final_grades.append(student_data)
        
        # Compute the general average and save it for each student
        for student_data in final_grades:
            total_final_grade = 0
            num_subjects = len(student_data['subjects'])
            for subject_info in student_data['subjects']:
                final_grade = subject_info['final_grade']
                try:    
                    if final_grade is not None:
                        total_final_grade += float(final_grade) 
                except ValueError:
                    # Handle the case where final_grade is not a valid number
                    pass

            student_data['general_average'] = total_final_grade / num_subjects if num_subjects > 0 else 0
            save_general_average(student_data, grade, section)

        sorted_final_grades = sorted(final_grades, key=lambda x: x.get('general_average', 0), reverse=True)
        highest_per_quarter = {
            'first_quarter': [],
            'second_quarter': [],
            'third_quarter': [],
            'fourth_quarter': [],
        }

        # Populate data for each quarter
        for quarters in ['first_quarter', 'second_quarter', 'third_quarter', 'fourth_quarter']:
            sorted_students = []
            for student in final_grades:
                if 'subjects' in student and student['subjects']:  # Check if 'subjects' list exists and is not empty
                    quarter_grades = student['subjects'][0]['quarter_grades'].get(quarters)
                    if quarter_grades is not None and quarter_grades != '':
                        sorted_students.append(student)
            sorted_students = sorted(sorted_students, key=lambda x: float(x['subjects'][0]['quarter_grades'].get(quarters, '0') or '0'), reverse=True)
            highest_per_quarter[quarters] = sorted_students



            general_averages = GeneralAverage.objects.filter(grade=grade, section=section)

        # Sort GeneralAverage instances based on the general average from highest to lowest
            sorted_general_averages = general_averages.order_by('-general_average')

        user = request.user
        if hasattr(user, 'teacher'):
            # Retrieve the teacher associated with the user
            teacher = user.teacher

            # Filter class records based on the teacher
            class_records = AdvisoryClass.objects.filter(teacher=teacher)

            # Keep track of unique grade and section combinations
            unique_combinations = set()
            unique_class_records = []

  

            # Iterate through class records to filter out duplicates
            for record in class_records:
                combination = (record.grade, record.section)
                # Check if the combination is unique
                if combination not in unique_combinations:
                    unique_combinations.add(combination)
                    unique_class_records.append(record)

            
        

        context = {
            'grade': grade,
            'section': section,
            'unique_keys': unique_keys_context,
            'students': students_filtered,
            'class_type': class_type,
            'data': data,
            'quarter': quarter,
            'final_grades': sorted_final_grades,
            'highest_per_quarter': highest_per_quarter,
            'general_averages': sorted_general_averages,
            'class_records': unique_class_records,

         
        }

        return render(request, 'teacher_template/adviserTeacher/advisory_quarterly_grades.html', context)

def advisory_final_all_subject(request):
    # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher
        teacher_id = teacher.id 

        grade = request.GET.get('grade')    
        section = request.GET.get('section')
        class_type = request.GET.get('class_type')
        quarter = request.GET.get('quarter', '1st Quarter')  # Default to 1st Quarter if not provided

        # Fetch advisory classes based on teacher, grade, and section
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section, teacher=teacher)


        unique_keys = set()  # Initialize an empty set here

        if advisory_classes.exists():
            for advisory_class in advisory_classes:
                grades_data = advisory_class.grades_data
                if grades_data:
                    for key, value in grades_data.items():
                        unique_keys.add((key, value.get('from_teacher_id')))  # Add (key, from_teacher_id) tuple to the set

        else:
            print("No AdvisoryClass objects found for the specified criteria")

      
        # Filter students based on the class type
        students = Student.objects.filter(grade=grade, section=section)
        students_filtered = []
        for student in students:
            class_type_json = student.class_type
            if class_type_json and str(teacher_id) in class_type_json and class_type_json[str(teacher_id)] == class_type:
                students_filtered.append(student)

        # Unique keys context
        unique_keys_context = list(unique_keys)

        # Filter quarterly grades based on the selected quarter
        quarterly_grades = QuarterlyGrades.objects.filter(student__grade=grade, student__section=section, quarter=quarter)

        # Prepare data to pass to the template
        data = []
        for index, grades in enumerate(quarterly_grades, start=1):
            student_name = grades.student.name
            subjects_grades = grades.grades
            average_score = subjects_grades.pop('average_score', None)
            subjects_data = [{'subject': subject, 'score': score} for subject, score in subjects_grades.items()]

            data.append({
                'no': index,
                'student_name': student_name,
                'subjects_data': subjects_data,
                'average_score': average_score,
            })
        
        students = AdvisoryClass.objects.filter(grade=grade, section=section)
        # Dictionary to store subject-wise grades and average score for each student
        subject_grades = {}
        quarter_mapping = {
            '1st Quarter': 'first_quarter',
            '2nd Quarter': 'second_quarter',
            '3rd Quarter': 'third_quarter',
            '4th Quarter': 'fourth_quarter',
        }

        # Fetch subject-wise grades for each student
        for student in students:
            grades_data = student.grades_data
            subject_grades[student.student.name] = {}
            grades = []  # List to store grades for calculating mean

            for subject, grades_info in grades_data.items():
                # Skip subjects 'MUSIC', 'ARTS', 'PE', and 'HEALTH'
                if subject in ['MUSIC', 'ARTS', 'PE', 'HEALTH']:
                    continue
                
                if quarter_mapping[quarter] in grades_info:
                    subject_grade = grades_info[quarter_mapping[quarter]]
         
                    subject_grade_str = str(subject_grade) if subject_grade is not None else ""
                    if subject_grade_str.strip():  # Check if the string is not empty after stripping whitespace
                        subject_grades[student.student.name][subject] = subject_grade_str
                        if subject_grade is not None:
                            grades.append(float(subject_grade))

            # Calculate average score
            if grades:
                subject_grades[student.student.name]['average_score'] = round(mean(grades), 2)
            else:
                subject_grades[student.student.name]['average_score'] = ""

            # Check if QuarterlyGrades entry already exists for this student and quarter
            existing_entry = QuarterlyGrades.objects.filter(student=student.student, quarter=quarter).first()
            if not existing_entry:
                # Save grades to QuarterlyGrades model
                QuarterlyGrades.objects.create(
                    student=student.student,
                    quarter=quarter,
                    grades=subject_grades[student.student.name]
                )
            elif existing_entry.grades != subject_grades[student.student.name]:
                # Update the existing entry if the grades are different
                existing_entry.grades = subject_grades[student.student.name]
                existing_entry.save()

        subjects = list(students.first().grades_data.keys()) if students else []

        students = Student.objects.filter(grade=grade, section=section)
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section)

        final_grades = []
        for student in students:
            student_data = {
                'id': student.id,
                'name': student.name,
                'grade': grade,
                'section': section,
                'subjects': [],
                'student': student
            }

            for advisory_class in advisory_classes.filter(student=student):
                grades_data = advisory_class.grades_data
                for subject, subject_info in grades_data.items():
                    if subject.upper() in ['MUSIC', 'ARTS', 'PE', 'HEALTH']:
                        continue
                    # Access grades data for each subject
                    subject_data = {
                        'subject': subject,
                        'quarter_grades': {
                            'first_quarter': subject_info.get('first_quarter', ''),
                            'second_quarter': subject_info.get('second_quarter', ''),
                            'third_quarter': subject_info.get('third_quarter', ''),
                            'fourth_quarter': subject_info.get('fourth_quarter', ''),
                            # Add more quarters if available
                        },
                        'final_grade': subject_info.get('final_grade', ''),
                        'teacher_name': subject_info.get('from_teacher_id', 'Unknown Teacher')
                    }
                    student_data['subjects'].append(subject_data)

            # Append student data to the final grades
            final_grades.append(student_data)
        
        # Compute the general average and save it for each student
        for student_data in final_grades:
            total_final_grade = 0
            num_subjects = len(student_data['subjects'])
            for subject_info in student_data['subjects']:
                final_grade = subject_info['final_grade']
                try:    
                    if final_grade is not None:
                        total_final_grade += float(final_grade) 
                except ValueError:
                    # Handle the case where final_grade is not a valid number
                    pass

            student_data['general_average'] = total_final_grade / num_subjects if num_subjects > 0 else 0
            save_general_average(student_data, grade, section)

        sorted_final_grades = sorted(final_grades, key=lambda x: x.get('general_average', 0), reverse=True)

        highest_per_quarter = {
            'first_quarter': [],
            'second_quarter': [],
            'third_quarter': [],
            'fourth_quarter': [],
        }

        # Populate data for each quarter
        for quarters in ['first_quarter', 'second_quarter', 'third_quarter', 'fourth_quarter']:
            sorted_students = []
            for student in final_grades:
                if 'subjects' in student and student['subjects']:  # Check if 'subjects' list exists and is not empty
                    quarter_grades = student['subjects'][0]['quarter_grades'].get(quarters)
                    if quarter_grades is not None and quarter_grades != '':
                        sorted_students.append(student)
            sorted_students = sorted(sorted_students, key=lambda x: float(x['subjects'][0]['quarter_grades'].get(quarters, '0') or '0'), reverse=True)
            highest_per_quarter[quarters] = sorted_students



            general_averages = GeneralAverage.objects.filter(grade=grade, section=section)

        # Sort GeneralAverage instances based on the general average from highest to lowest
            sorted_general_averages = general_averages.order_by('-general_average')

        user = request.user
        if hasattr(user, 'teacher'):
            # Retrieve the teacher associated with the user
            teacher = user.teacher

            # Filter class records based on the teacher
            class_records = AdvisoryClass.objects.filter(teacher=teacher)

            # Keep track of unique grade and section combinations
            unique_combinations = set()
            unique_class_records = []


            # Iterate through class records to filter out duplicates
            for record in class_records:
                combination = (record.grade, record.section)
                # Check if the combination is unique
                if combination not in unique_combinations:
                    unique_combinations.add(combination)
                    unique_class_records.append(record)

            
        

        context = {
            'grade': grade,
            'section': section,
            'unique_keys': unique_keys_context,
            'students': students_filtered,
            'class_type': class_type,
            'data': data,
            'quarter': quarter,
            'final_grades': sorted_final_grades,
            'highest_per_quarter': highest_per_quarter,
            'general_averages': sorted_general_averages,
            'class_records': unique_class_records,

         
        }


        return render(request, 'teacher_template/adviserTeacher/advisory_final_all_subject.html', context)
    
def create_attendance_view(request):
    if request.method == 'GET':
        grade = request.GET.get('grade')
        section = request.GET.get('section')
        teacher = request.user.teacher
        teacher_id = teacher.id 
        class_type = request.GET.get('class_type')
        month = request.GET.get('month', '')  # Assuming month is passed via GET request



        # Check if the month already exists for the class
        students = Student.objects.filter(grade=grade, section=section)
        students_filtered = []
        for student in students:
            class_type_json = student.class_type
            if class_type_json and str(teacher_id) in class_type_json and class_type_json[str(teacher_id)] == class_type:
                students_filtered.append(student)
                

        context = {
            'students': students_filtered,
        }
        return render(request, 'teacher_template/adviserTeacher/create_attendance.html', context)
    
def save_attendance_record(request):
    if request.method == 'POST':
        month = request.POST.get('month')
        school_days = int(request.POST.get('school_days', 0))
        students = request.POST.getlist('student_id')
        response_data = {'message': 'Attendance records saved successfully'}
        error_response_data = {'error': f'Attendance records for {month} already exist'}

        if AttendanceRecord.objects.filter(attendance_record__has_key=month).exists():
            return JsonResponse(error_response_data, status=400)

        # Initialize totals
        total_school_days = 0
        total_days_present = 0
        total_days_absent = 0

        # Loop through the students and process their attendance data
        for student_id in students:
            # Retrieve days_present and days_absent data
            days_present_str = request.POST.get(f'days_present_{student_id}')
            days_present = int(days_present_str) if days_present_str else 0
            days_absent_str = request.POST.get(f'days_absent_{student_id}')
            days_absent = int(days_absent_str) if days_absent_str else 0

            total_school_days += school_days
            total_days_present += days_present
            total_days_absent += days_absent

            try:
                student = Student.objects.get(id=student_id)
            except Student.DoesNotExist:
                # Create a new student if not exists
                student = Student.objects.create(id=student_id)

            # Check if an attendance record exists for the student and month
            try:
                attendance_record = AttendanceRecord.objects.get(
                    student=student,
                    attendance_record__has_key=month  # Filter records based on the key
                )

                # Update the existing attendance record with new data
                existing_data = attendance_record.attendance_record.get(month, {})
                existing_data['No. of School Days'] = school_days
                existing_data['No. of Days Present'] = days_present
                existing_data['No. of Days Absent'] = days_absent
                attendance_record.attendance_record[month] = existing_data
                attendance_record.save()

            except AttendanceRecord.DoesNotExist:
                # If the AttendanceRecord does not exist for the student and month,
                # create a new record with the given data and month
                # Check if the student has any existing records, if not, create a new one
                if not AttendanceRecord.objects.filter(student=student).exists():
                    AttendanceRecord.objects.create(
                        student=student,
                        attendance_record={
                            month: {
                                'No. of School Days': school_days,
                                'No. of Days Present': days_present,
                                'No. of Days Absent': days_absent
                            }
                        }
                    )
                else:
                    # Retrieve all existing records for the student
                    existing_records = AttendanceRecord.objects.filter(student=student)
                    # Update all existing records by adding the new month and its data
                    for record in existing_records:
                        record.attendance_record[month] = {
                            'No. of School Days': school_days,
                            'No. of Days Present': days_present,
                            'No. of Days Absent': days_absent
                        }
                        record.save()

            total_attendance_record, created = AttendanceRecord.objects.get_or_create(
                student=student,  # For total attendance record
                defaults={
                    'attendance_record': {
                        'Total': {
                            'Total School Days': school_days,
                            'Total Days Present': days_present,
                            'Total Days Absent': days_absent
                        }
                    }
                }
            )

            if not created:
                # If the record already exists, update the existing data
                total_record_data = total_attendance_record.attendance_record.get('TOTAL', {})
                # Ensure that the 'Total' key exists and initialize it with default values if not
                if not total_record_data:
                    total_record_data = {
                        'Total School Days': 0,
                        'Total Days Present': 0,
                        'Total Days Absent': 0
                    }
                total_record_data['Total School Days'] += school_days
                total_record_data['Total Days Present'] += days_present
                total_record_data['Total Days Absent'] += days_absent
                total_attendance_record.attendance_record['TOTAL'] = total_record_data

                total_attendance_record.attendance_record = {
                    k: total_attendance_record.attendance_record[k] for k in sorted(total_attendance_record.attendance_record.keys()) if k != 'TOTAL'
                }
                total_attendance_record.attendance_record['TOTAL'] = total_record_data
                total_attendance_record.save()


        return JsonResponse(response_data, status=200)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)

@csrf_exempt
def delete_month(request):
    if request.method == 'POST':
        grade = request.POST.get('grade')
        section = request.POST.get('section')
        month = request.POST.get('month')

        try:
            # Get all students in the specified grade and section
            students = Student.objects.filter(grade=grade, section=section)
            
            for student in students:
                # Get the attendance record for the student
                try:
                    record = AttendanceRecord.objects.get(student=student)
                    if month in record.attendance_record:
                        # Extract the attendance data for the month being deleted
                        attendance_data = record.attendance_record[month]

                        # Update total school days, days present, and days absent if 'Total' exists
                        if 'TOTAL' in record.attendance_record:
                            total_school_days = attendance_data.get('No. of School Days', 0)
                            total_days_present = attendance_data.get('No. of Days Present', 0)
                            total_days_absent = attendance_data.get('No. of Days Absent', 0)

                            # Subtract the attendance data for the deleted month from the total records
                            record.attendance_record['TOTAL']['Total School Days'] -= total_school_days
                            record.attendance_record['TOTAL']['Total Days Present'] -= total_days_present
                            record.attendance_record['TOTAL']['Total Days Absent'] -= total_days_absent

                        # Delete the attendance record for the month
                        del record.attendance_record[month]
                        record.save()
                except AttendanceRecord.DoesNotExist:
                    pass

            return JsonResponse({'status': 'success'})
        
        except Student.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'No students found for the specified grade and section'})
    
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
def attendance_record_view(request, grade, section):
    # Filter students based on grade and section
    students = Student.objects.filter(grade=grade, section=section)
    # Initialize an empty list to store attendance records for all students
    attendance_records = []

    # Loop through each student and collect their attendance records
    for student in students:
        # Filter attendance records based on the current student
        records_for_student = AttendanceRecord.objects.filter(student=student)

        # Extend the attendance_records list with the records_for_student QuerySet
        attendance_records.append(records_for_student)

    months_set = set()
    for records in attendance_records:
        for record in records:
            if record.attendance_record:
                months_set.update(record.attendance_record.keys())

    # Get the list of month names in the correct order
    month_name = list(calendar.month_name)[1:]

    # Convert the set of months to a list and sort it based on the calendar order
    months_list = sorted(months_set, key=lambda m: month_name.index(m) if m in month_name else float('inf'))

    # Pass the attendance records and sorted list of months to the template for rendering
    context = {
        'grade': grade,
        'section': section,
        'attendance_records': attendance_records,
        'months': months_list
    }

    # Render the template with the attendance record data
    return render(request, 'teacher_template/adviserTeacher/attendance_record_view.html', context)

@csrf_exempt
def update_attendance_record(request):
    if request.method == 'POST':
        record_id = request.POST.get('student_id')
        month = request.POST.get('month')  # Month for which data is updated
        field_name = request.POST.get('key')  # Field to update: "No. of School Days", "No. of Days Present", "No. of Days Absent"
        new_value = request.POST.get('new_value')  # New value to update
        absent_days = request.POST.get('absent_days', '')
        total_present = request.POST.get('total_present')
        total_absent = request.POST.get('total_absent')  # New value for 'No. of Days Absent'



        try:
            # Fetch the AttendanceRecord object
            record = AttendanceRecord.objects.get(student=record_id)
            
            # Parse the JSON field
            attendance_record = record.attendance_record

            # Update the value based on the month and field name
            if month in attendance_record and field_name in attendance_record[month]:
                attendance_record[month][field_name] = new_value
                
                # Update the 'No. of Days Absent' if available
                if 'No. of Days Absent' in attendance_record[month]:
                    attendance_record[month]['No. of Days Absent'] = absent_days
                
                total_days_present = 0
         
                for month, data in attendance_record.items():
                    if 'No. of Days Present' in data:
                        total_days_present += int(data['No. of Days Present'])

                total_days_absent = 0
            
                for month, data in attendance_record.items():
                    if 'No. of Days Absent' in data:
                        total_days_absent += int(data['No. of Days Absent'])
                # Assign the calculated total days present to the record
                if 'Total Days Present' in attendance_record[month]:
                    attendance_record[month]['Total Days Present'] = total_days_present

                if 'Total Days Absent' in attendance_record[month]:
                    attendance_record[month]['Total Days Absent'] = total_days_absent
                
                # Save the updated JSON field back to the object
                record.attendance_record = attendance_record
                record.save()

                # Respond with a success message
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Month or field name not found'})
        
        except AttendanceRecord.DoesNotExist:
            # Respond with an error message if the record does not exist
            return JsonResponse({'status': 'error', 'message': 'Record not found'})
        
        except Exception as e:
            # Respond with an error message if any other exception occurs
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        # Respond with an error message for invalid request methods
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def display_advisory_data(request):
        # Assuming the user is logged in
    user = request.user

    # Check if the user is a teacher
    if user.is_authenticated and hasattr(user, 'teacher'):
        # Retrieve the teacher associated with the user
        teacher = user.teacher

        grade = request.GET.get('grade')    
        section = request.GET.get('section')
        key = request.GET.get('key')

        # Fetch students based on grade and section
        students = Student.objects.filter(grade=grade, section=section)
        # Fetch advisory classes based on teacher, grade, and section
        advisory_classes = AdvisoryClass.objects.filter(
            grade=grade, 
            section=section, 
            grades_data__has_key=key
        )
     

        for advisory_class in advisory_classes:

            grades_data = advisory_class.grades_data
            if grades_data:
                specific_key = key
            else:
                print("No grades data available")

       
        context = {
            'grade': grade,
            'subject': key,
            'section': section,
            'advisory_classes': advisory_classes,
            'students': students,
            # 'class_type': class_type,
           
        }
            
    return render(request, 'teacher_template/adviserTeacher/subject_quarter_advisory.html', context)


def update_final_grade(request):
    if request.method == 'POST' and request.headers.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        final_grade = request.POST.get('final_grade')
        subject = request.POST.get('subject')
        grade = request.POST.get('grade')
        section = request.POST.get('section')


        try:
            advisory_class = AdvisoryClass.objects.get(grade=grade, section=section)
        except AdvisoryClass.DoesNotExist:
            return JsonResponse({'error': 'Advisory class not found for grade and section'}, status=404)

        grades_data = advisory_class.grades_data or {}
        
        if subject in grades_data:
            grades_data[subject]= final_grade
        else:
            grades_data[subject] = {final_grade}
        
        advisory_class.grades_data = grades_data
        advisory_class.save()

        return JsonResponse({'message': 'Final grade updated successfully'}, status=200)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)

def display_student_transmuted_grades(request):
    grade = request.GET.get('grade')
    section = request.GET.get('section')
    subject = request.GET.get('subject')
    teacher_id = request.GET.get('teacher_id')

    advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section, subject=subject, from_teacher_id=teacher_id)

    context = {
        'advisory_classes': advisory_classes,
    }

    return render(request, 'teacher_template/adviserTeacher/advisory_final_grade_subject.html', context)


    
def edit_record(request, record_id):
    # Retrieve the specific record based on the record_id
    record = GradeScores.objects.get(pk=record_id)

    if request.method == 'POST':
        form = GradeScoresForm(request.POST, instance=record)
        if form.is_valid():
            # Save the updated form data to the database
            form.save()
            # Redirect to the same page or another page after successful update
            return redirect('edit_record', record_id=record_id)
    else:
        form = GradeScoresForm(instance=record)

    context = {
        'form': form,
        'record': record,
    }

    return render(request, 'teacher_template/adviserTeacher/edit_records.html', context)


def display_quarterly_summary(request, grade, section, subject, class_record_id):
    # Retrieve the specific class record based on the provided class_record_id
    class_record = get_object_or_404(ClassRecord, id=class_record_id, grade=grade, section=section, subject=subject)

    # Retrieve grade scores related to the class record
    grade_scores = GradeScores.objects.filter(class_record=class_record)

    # Handle None values for initial_grades and transmuted_grades
    for grade_score in grade_scores:
        if grade_score.initial_grades is None:
            grade_score.initial_grades = ""
        if grade_score.transmuted_grades is None:
            grade_score.transmuted_grades = ""

    context = {
        'class_record': class_record,
        'grade_scores': grade_scores,
    }

    return render(request, "teacher_template/adviserTeacher/summary_of_quarterly_grade.html", context)




def grade_summary(request, grade, section, quarter):
    students = AdvisoryClass.objects.filter(grade=grade, section=section)

    # Dictionary to store subject-wise grades and average score for each student
    subject_grades = {}
    quarter_mapping = {
        '1st Quarter': 'first_quarter',
        '2nd Quarter': 'second_quarter',
        '3rd Quarter': 'third_quarter',
        '4th Quarter': 'fourth_quarter',
    }

    # Fetch subject-wise grades for each student
    for student in students:
        grades_data = student.grades_data
        subject_grades[student.student.name] = {}
        grades = []  # List to store grades for calculating mean

        for subject, grades_info in grades_data.items():
            if quarter_mapping[quarter] in grades_info:
                subject_grade = grades_info[quarter_mapping[quarter]]
       
                subject_grades[student.student.name][subject] = subject_grade
                if subject_grade is not None:
                    grades.append(float(subject_grade))

        # Calculate average score
        if grades:
            subject_grades[student.student.name]['average_score'] = round(mean(grades), 2)
        else:
            subject_grades[student.student.name]['average_score'] = None

        # Check if QuarterlyGrades entry already exists for this student and quarter
        existing_entry = QuarterlyGrades.objects.filter(student=student.student, quarter=quarter).first()
        if not existing_entry:
            # Save grades to QuarterlyGrades model
            QuarterlyGrades.objects.create(
                student=student.student,
                quarter=quarter,
                grades=subject_grades[student.student.name]
            )
        elif existing_entry.grades != subject_grades[student.student.name]:
            # Update the existing entry if the grades are different
            existing_entry.grades = subject_grades[student.student.name]
            existing_entry.save()

    subjects = list(students.first().grades_data.keys()) if students else []

    context = {
        'students': students,
        'subject_grades': subject_grades,
        'subjects': subjects,
        'quarter': quarter,
    }

    return render(request, 'teacher_template/adviserTeacher/quarterly_summary.html', context)


def get_subject_score(student, subject, quarter):
    quarter_mapping = {
        '1st Quarter': 'first_quarter',
        '2nd Quarter': 'second_quarter',
        '3rd Quarter': 'third_quarter',
        '4th Quarter': 'fourth_quarter',
    }
    
    try:
        # Fetch the AdvisoryClass instance for the given student, subject, and quarter
        advisory_instance = AdvisoryClass.objects.get(
            student=student,
            grade=student.grade,
            section=student.section,
        )

        # Access the grades_data JSONField and retrieve the score for the given subject and quarter
        grades_data = advisory_instance.grades_data

        if subject in grades_data:
            quarter_grades = grades_data[subject]
            subject_score = quarter_grades.get(quarter_mapping[quarter])
            return subject_score  # Return the score if found

        # If subject or quarter not found, return None
        return None

    except AdvisoryClass.DoesNotExist:
        # Handle the case where the AdvisoryClass record does not exist
        return None

    except MultipleObjectsReturned:
        # Handle the case where multiple AdvisoryClass records are returned
        # For example, you can log a warning or return a default value
        return None

    except KeyError:
        # Handle the case where the subject or quarter key is not found in the grades_data
        return None
    
def get_subjects(student):
    # Fetch unique subjects for the given student from the GradeScores model
    subjects = GradeScores.objects.filter(
        student=student,  # Update this line to use the actual Student object
    ).values_list('class_record__subject', flat=True).distinct()

    return [subject for subject in subjects if subject]




def calculate_save_final_grades(grade, section, subject, students, subjects):
    quarters = ['1st Quarter', '2nd Quarter', '3rd Quarter', '4th Quarter']

    for student in students:
        student_final_grades = {}  # Store final grades for all subjects for this student
            
        # Check if a record already exists in the FinalGrade model
        existing_final_grade = FinalGrade.objects.filter(
            teacher__classrecord__grade=grade,
            teacher__classrecord__section=section,
            student=student
        ).first()

        if existing_final_grade:
            final_grade_info = json.loads(existing_final_grade.final_grade)  # Convert JSON string to dictionary
            student_final_grades = {item.get('subject'): item for item in final_grade_info}

        for subject_data in subjects:
            subject_name = subject_data.subject

            # Initialize subject data
            subject_info = {'subject': subject_name, 'quarter_grades': {}, 'final_grade': 0}

            # Retrieve initial grades per quarter
            for quarter in quarters:
                grade_score = GradeScores.objects.filter(
                    class_record__grade=grade,
                    class_record__section=section,
                    class_record__subject=subject_name,
                    class_record__quarters=quarter,
                    student=student
                ).first()

                subject_info['quarter_grades'][quarter] = grade_score.transmuted_grades if grade_score else 0

            # Get the teacher's name from the ClassRecord
            class_record = ClassRecord.objects.filter(
                grade=grade,
                section=section,
                subject=subject_name
            ).first()

            # Ensure that the class_record and its teacher exist
            if class_record and class_record.teacher:
                subject_info['teacher_name'] = f"{class_record.teacher.user.first_name} {class_record.teacher.user.last_name}"
            else:
                subject_info['teacher_name'] = "Unknown Teacher"

            # Calculate the final grade for the subject
            quarter_grades_values = subject_info['quarter_grades'].values()

            # Convert filtered values to a list before calculating the sum and length
            filtered_values = list(filter(lambda x: x is not None, quarter_grades_values))

            subject_info['final_grade'] = sum(filtered_values) / len(filtered_values) \
                if filtered_values else 0

            # Store subject info in the dictionary under the subject name
            student_final_grades[subject_name] = subject_info

        final_grade_data = {
            'teacher': class_record.teacher,
            'student': student,
            'grade': grade,
            'subject': subject,
            'section': section,
            'final_grade': json.dumps(list(student_final_grades.values()))  # Convert dictionary values to list and then to JSON string
        }

        if existing_final_grade:
            # Update existing record
            existing_final_grade.final_grade = json.dumps(list(student_final_grades.values()))
            existing_final_grade.save()
        else:
            # Insert new record
            final_grade = FinalGrade.objects.create(**final_grade_data)


def display_final_grades(request, grade, section, subject):
    teacher = request.user.teacher
    students = Student.objects.filter(grade=grade, section=section)
    
    advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section, teacher=teacher)

    final_grades = []
    for advisory_class in advisory_classes:
        for student in students.filter(id=advisory_class.student_id):
            student_data = {'id': student.id, 'name': student.name, 'grade': grade, 'section': section, 'subjects': []}

            grades_data = advisory_class.grades_data
            if grades_data and subject in grades_data:
                subject_info = grades_data[subject]
                subject_data = {
                    'subject': subject,
                    'quarter_grades': {
                        'first_quarter': subject_info.get('first_quarter', ''),
                        'second_quarter': subject_info.get('second_quarter', ''),
                        'third_quarter': subject_info.get('third_quarter', ''),
                        'fourth_quarter': subject_info.get('fourth_quarter', ''),
                        # Add more quarters if available
                    },
                    'final_grade': subject_info.get('final_grade', ''),
                    'teacher_name': subject_info.get('from_teacher_id', 'Unknown Teacher')
                }
                student_data['subjects'].append(subject_data)

            final_grades.append(student_data)


    context = {
        'grade': grade,
        'section': section,
        'final_grades': final_grades,
        'subject': subject,  # Pass the subject to the template
    }

    return render(request, "teacher_template/adviserTeacher/new_final_grades.html", context)


def determine_remarks(general_average):
    if general_average is None:
        return 'No Grade'
     
    if general_average >= 98:
        return 'WITH HIGHEST HONOR'
    elif general_average >= 95:
        return 'WITH HIGH HONOR'
    elif general_average >= 90:
        return 'WITH HONOR'
    elif general_average >= 75:
        return 'PASSED'
    else:
        return 'FAILED'
    
def determine_status(general_average):
    if general_average is None:
        return 'No Grade'
     
    if general_average >= 98:
        return 'PASSED'
    elif general_average >= 95:
        return 'PASSED'
    elif general_average >= 90:
        return 'PASSED'
    elif general_average >= 75:
        return 'PASSED'
    else:
        return 'FAILED'


def save_general_average(student_data, grade, section):
    # Check if 'general_average' key is present in student_data and is not None
    if 'general_average' in student_data and student_data['general_average'] is not None:
        # Extract relevant information from student_data
        student = student_data['student']
        general_average = round(student_data['general_average'], 2)
        remarks = determine_remarks(general_average)
        status = determine_status(general_average)

        # Filter GeneralAverage objects based on student, grade, and section
        general_average_records = GeneralAverage.objects.filter(
            student=student,
            grade=grade,
            section=section,
        )

        # Check if any records exist
        if general_average_records.exists():
            # Update the first record
            general_average_record = general_average_records.first()
            general_average_record.general_average = general_average
            general_average_record.remarks = remarks
            general_average_record.status = status
            general_average_record.save()
        else:
            # Create a new record if none exist
            GeneralAverage.objects.create(
                student=student,
                grade=grade,
                section=section,
                general_average=general_average,
                remarks=remarks,
                status=status
            )

def display_all_final_grades(request, grade, section):
    students = Student.objects.filter(grade=grade, section=section)
    advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section)

    final_grades = []
    for student in students:
        student_data = {
            'id': student.id,
            'name': student.name,
            'grade': grade,
            'section': section,
            'subjects': [],
            'student': student
        }

        for advisory_class in advisory_classes.filter(student=student):
            grades_data = advisory_class.grades_data
            for subject, subject_info in grades_data.items():
                # Access grades data for each subject
                subject_data = {
                    'subject': subject,
                    'quarter_grades': {
                        'first_quarter': subject_info.get('first_quarter', ''),
                        'second_quarter': subject_info.get('second_quarter', ''),
                        'third_quarter': subject_info.get('third_quarter', ''),
                        'fourth_quarter': subject_info.get('fourth_quarter', ''),
                        # Add more quarters if available
                    },
                    'final_grade': subject_info.get('final_grade', ''),
                    'teacher_name': subject_info.get('from_teacher_id', 'Unknown Teacher')
                }
                student_data['subjects'].append(subject_data)

        # Append student data to the final grades
        final_grades.append(student_data)
    
    # Compute the general average and save it for each student
    for student_data in final_grades:
        total_final_grade = 0
        num_subjects = len(student_data['subjects'])
        for subject_info in student_data['subjects']:
            final_grade = subject_info['final_grade']
            try:
                total_final_grade += float(final_grade)
            except ValueError:
                # Handle the case where final_grade is not a valid number
                pass

        student_data['general_average'] = total_final_grade / num_subjects if num_subjects > 0 else 0
        save_general_average(student_data, grade, section)

        sorted_final_grades = sorted(final_grades, key=lambda x: x.get('general_average', 0), reverse=True)
        highest_per_quarter = {
            'first_quarter': [],
            'second_quarter': [],
            'third_quarter': [],
            'fourth_quarter': [],
        }

        # Populate data for each quarter
        for quarter in ['first_quarter', 'second_quarter', 'third_quarter', 'fourth_quarter']:
            sorted_students = sorted(final_grades, key=lambda x: float(x['subjects'][0]['quarter_grades'].get(quarter, '0') or '0'), reverse=True)
            highest_per_quarter[quarter] = sorted_students  



        general_averages = GeneralAverage.objects.filter(grade=grade, section=section)

    # Sort GeneralAverage instances based on the general average from highest to lowest
        sorted_general_averages = general_averages.order_by('-general_average')


        # Now you have a dictionary 'highest_per_quarter' where each quarter maps to a list of students
        # The students are sorted in descending order of their grades for each quarter

        
        context = {
            'grade': grade,
            'section': section,
            'final_grades': sorted_final_grades,
            'highest_per_quarter': highest_per_quarter,
             'general_averages': sorted_general_averages
        }
        
    return render(request, "teacher_template/adviserTeacher/new_all_final_grades.html", context)


    
@login_required
@transaction.atomic
def update_score(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        student_name = request.POST.get('student_name')
        new_score = request.POST.get('new_score')
        class_record_id = request.POST.get('class_record_id')
        score_data = request.POST.getlist('scoreData[]')  # Changed to match AJAX data key
        assessment_type = request.POST.get('assessment_type')

         # Assuming you want to filter students by both name and teacher
        students = Student.objects.filter(name=student_name)
        scores_per_assessment = "scores_per_assessment"
        # Ensure there is exactly one matching student
        if students.count() == 1:
            student = students.first()
        else:
            return JsonResponse({'success': False, 'error': 'Error in finding the student'})

        try:
            # Retrieve the GradeScores object based on student name and class record id
            grade_score = GradeScores.objects.get(
                student=student,
                class_record__id=class_record_id
            )

        except GradeScores.DoesNotExist:
            print(f"Not Found: /update_score/")
            return HttpResponse("GradeScores not found for the given student and class record ID.", status=404)
        
        scores_hps_data = grade_score.grade_scores['scores_hps'][assessment_type]['SCORES']

        # Check if any score exceeds the corresponding scores_hps value

        previous_scores = grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'][:]

        for idx, score in enumerate(score_data):
            current_score = grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'][idx]
            if not (current_score.isdigit() or current_score == ''):  # Check if score is not a numerical value at the current index
                grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'][idx] = ''
                grade_score.save()
                return JsonResponse({'success': False, 'error': 'Invalid score: Please enter a numerical value'})
            
            if score.isdigit() and scores_hps_data[idx].isdigit():  # Only check if both are numeric
                if int(score) > int(scores_hps_data[idx]):
                    grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'][idx] = ''
                    grade_score.save()
                    return JsonResponse({'success': False, 'error': 'Invalid score: Exceeds highest possible score'})
                
            elif scores_hps_data[idx] == '' and score_data[idx] != '':   # Check if the corresponding scores_hps_data entry is an empty string
                grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'][idx] = ''
                grade_score.save()
                return JsonResponse({'success': False, 'error': 'Invalid score: HPS score is empty'})
  
        grade_score.grade_scores[scores_per_assessment][assessment_type]['scores'] = score_data
        total_score_data = sum(int(score) for score in score_data if score.isdigit())
        total_hps_data = grade_score.grade_scores['scores_hps'][assessment_type]['TOTAL_HPS']
        weight_input = float(grade_score.grade_scores['scores_hps'][assessment_type]['WEIGHT'])

        total_hps_data = float(total_hps_data)
        # Calculate percentage score
        if total_hps_data != 0:
            percentage_score_data = round((total_score_data / total_hps_data) * 100, 2)
        else:
            percentage_score_data = None


        if percentage_score_data is not None:
            total_weighted_score = round((percentage_score_data / 100) * weight_input, 2)
        else:
            total_weighted_score = None

        grade_score.grade_scores[scores_per_assessment][assessment_type]['total_score'] = total_score_data
        grade_score.grade_scores[scores_per_assessment][assessment_type]['percentage_score'] = percentage_score_data
        grade_score.grade_scores[scores_per_assessment][assessment_type]['total_weighted_score'] = total_weighted_score

        assessments = grade_score.grade_scores[scores_per_assessment].keys()

        # Compute initial grades by summing total_weighted_score of each assessment
        initial_grades = sum(
            float(grade_score.grade_scores[scores_per_assessment][assmt_type]['total_weighted_score'] or 0)
            for assmt_type in assessments
        )

        transmuted_grades = transmuted_grade(initial_grades)

        grade_score.initial_grades = initial_grades
        grade_score.transmuted_grades = transmuted_grades

        try:
            grade_score.save()
            # print("Grade score saved successfully!")
        except Exception as e:
            print("Error saving grade score:", e)

        response_data = {'success': True}

        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Invalid request'})


def update_highest_possible_scores(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        class_record_id = request.POST.get('class_record_id')
        new_hps_data = request.POST.getlist('new_hps_data[]')
        assessment_type = request.POST.get('assessment_type')


        try:
            # Retrieve GradeScores objects based on the given class record id
            grade_scores = GradeScores.objects.filter(class_record_id=class_record_id)
        except GradeScores.DoesNotExist:
            return JsonResponse({'error': 'GradeScores not found for the given class record ID.'}, status=404)

        # Update the highest possible scores data for each GradeScores object
        for grade_score in grade_scores:
            # Update scores_hps field with new_hps_data

            if not all(value.isdigit() or value == '' for value in new_hps_data):
              
                return JsonResponse({'success': False, 'error': 'Invalid highest possible scores data. Please provide valid numeric values.'})

            grade_score.grade_scores['scores_hps'][assessment_type]['SCORES'] = new_hps_data
            total_hps = sum(int(value) for value in new_hps_data if value.isdigit())


            grade_score.grade_scores['scores_hps'][assessment_type]['TOTAL_HPS'] = total_hps
            grade_score.save()

            response_data = {'success': True}

        return JsonResponse(response_data)

    return JsonResponse({'error': 'Invalid request'}, status=400)

def update_total_max_quarterly(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        class_record_id = request.POST.get('class_record_id')
        new_score = request.POST.get('new_score')



        try:
            # Retrieve GradeScores objects based on the given class record id
            grade_scores = GradeScores.objects.filter(class_record_id=class_record_id)
        except GradeScores.DoesNotExist:
            return JsonResponse({'error': 'GradeScores not found for the given class record ID.'}, status=404)

        # Determine the field to update based on the section_id
        section_id = 'quarterly_assessment'
        hps_field = 'scores_hps_quarterly'
        total_hps_field = 'total_qa_hps'

        # Update the total_qa_hps for each GradeScores object
        for grade_score in grade_scores:
            try:
                # Convert new_score to an integer (handle empty strings as well)
                total_qa_hps_value = int(float(new_score)) if new_score.strip() != '' else 0

                # Update the total_qa_hps field
                setattr(grade_score, total_hps_field, total_qa_hps_value)
                grade_score.save()

            except Exception as e:
                # Print or log the exception for debugging purposes
                print(f"Error updating grade_score {grade_score.id}: {e}")

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request'})



def validate_score(request):
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        student_name = request.POST.get('student_name')
        new_score = request.POST.get('new_score')
        column_index = int(request.POST.get('column_index'))
        section_id = request.POST.get('section_id')
        class_record_id = request.POST.get('class_record_id')

        try:
            # Assuming you want to filter GradeScores by student name and class record ID
            grade_score = GradeScores.objects.get(
                student__name=student_name,
                class_record__id=class_record_id
            )
        except GradeScores.DoesNotExist:
            return JsonResponse({'error': 'GradeScores not found for the given student and class record ID.'}, status=404)

        if section_id == 'written_works':
            scores_field = 'written_works_scores'
            hps_field = 'scores_hps_written'
        elif section_id == 'performance_task':
            scores_field = 'performance_task_scores'
            hps_field = 'scores_hps_performance'
        elif section_id == 'quarterly_assessment':
            scores_field = 'quarterly_assessment_scores'
            hps_field = 'scores_hps_quarterly'
        else:
            return JsonResponse({'error': 'Invalid section_id'})

        # Ensure scores_list and hps_list have enough elements, initialize with zeros if necessary
        scores_list = getattr(grade_score, scores_field, [0] * (column_index + 1))
        hps_list = getattr(grade_score, hps_field, [0] * (column_index + 1))

        # Validate the new score against its corresponding HPS
        highest_possible_score_str = hps_list[column_index]
        if highest_possible_score_str != '':
            highest_possible_score = int(highest_possible_score_str)  # Convert HPS to integer
            if new_score != '' and int(new_score) > highest_possible_score:  # Convert score to integer
                return JsonResponse({'error': f'Score cannot be greater than the highest possible score for index {column_index}.'})
        else:
            return JsonResponse({'error': f'Highest possible score is empty for index {column_index}.'})

        # Validate HPS against written scores
        for i, score in enumerate(scores_list):
            if hps_list[i] != '' and int(score) > int(hps_list[i]):
                return JsonResponse({'error': f'Written score cannot be greater than the highest possible score for index {i}.'})

        return JsonResponse({'success': 'Validation passed.'})

    return JsonResponse({'error': 'Invalid request'})



@require_POST
def delete_classrecord(request, class_record_id):
    class_record = get_object_or_404(ClassRecord, id=class_record_id)
    user = request.user
    class_record_name = class_record.name

    # Assuming you have some permission checks here before deleting

    # Delete associated GradeScores records based on class record id
    GradeScores.objects.filter(class_record_id=class_record_id).delete()

    action = f'{user} deleted a Classrecord named "{class_record_name}"'
    details = f'{user} deleted a Classrecord named "{class_record_name}" in the system along with its GradeScores.'
    log_activity(user, action, details)

    # Now, delete the ClassRecord
    class_record.delete()

    return JsonResponse({'message': 'Record deleted successfully'})

# Your other views remain the same
def class_records_list(request):
    class_records = ClassRecord.objects.all()
    return render(request, 'teacher_template/adviserTeacher/view_classrecord.html', {'class_records': class_records})

def tempo_newupload(request):
    return render(request, 'teacher_template/adviserTeacher/tempo_newupload.html')

def read_excel_values(excel_file, sheet_name):
    # Loading the workbook
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[sheet_name]

    data = []
    start_reading = False

    for row in sheet.iter_rows():
        row_values = []
        for cell in row[1:]:
            cell_value = cell.value
            row_values.append(cell_value)
        
        if any(cell_value is not None and cell_value != 0 for cell_value in row_values):
            if any(isinstance(cell_value, str) and ("MALE" in cell_value or "FEMALE" in cell_value) for cell_value in row_values):
                start_reading = True
            
            if start_reading:
                data.append(row_values)

    return data


def process_row(row):
    # Find the index of the first non-None element after index 0
    index = 1
    while index < len(row) and row[index] is None:
        index += 1

    # Remove the three None values next to index 0
    row = [row[0]] + row[index:]

    # Find the index of the last non-None element
    last_index = len(row) - 1
    while last_index >= 0 and row[last_index] is None:
        last_index -= 1

    # Remove all None values from the end of the list until reaching a non-None value
    row = row[:last_index + 1]

    return row

def divide_scores(score_list):
    # Convert all values to string
    score_list = [str(value) if value is not None else '' for value in score_list]

    written_works_scores = score_list[:10]
    
    # Check if there are enough elements in the score_list for written works
    if len(score_list) >= 13:
        total_scores_written = score_list[10]
        percentage_score_written = score_list[11]
        weighted_score_written = score_list[12]
    else:
        # Set default values if there are not enough elements for written works
        total_scores_written = ''
        percentage_score_written = ''
        weighted_score_written = ''
    
    written_works_dict = {
        "written_works_scores": written_works_scores,
        "total_scores_written": total_scores_written,
        "percentage_score_written": percentage_score_written,
        "weighted_score_written": weighted_score_written
    }

    
    # Extract Performance Task scores and related values
    performance_task_scores = score_list[13:23]
    
    # Check if there are enough elements in the score_list for performance task
    if len(score_list) >= 26:
        total_score_performance = score_list[23]
        percentage_score_performance = score_list[24]
        weighted_score_performance = score_list[25]
    else:
        # Set default values if there are not enough elements for performance task
        total_score_performance = ''
        percentage_score_performance = ''
        weighted_score_performance = ''
    
    performance_task_dict = {
        "performance_task_scores": performance_task_scores,
        "total_score_performance": total_score_performance,
        "percentage_score_performance": percentage_score_performance,
        "weighted_score_performance": weighted_score_performance
    }

    if len(score_list) >= 27:
        quarterly_assessment_scores = [score_list[26]] + [''] * 9
    else:
        quarterly_assessment_scores = [''] * 10

    # Check if there are enough elements in the score_list for Quarterly Assessment
    if len(score_list) >= 29:
        quarterly_assessment = {
            "quarterly_assessment_scores": quarterly_assessment_scores,
            "total_score_quarterly": score_list[26],
            "percentage_score_quarterly": score_list[27],
            "weighted_score_quarterly": score_list[28]
        }
    else:
        quarterly_assessment = None


    # Check if there are enough elements in the score_list for Initial Grade
    if len(score_list) >= 30:
        initial_grade = {"initial_grades": score_list[29]}
    else:
        initial_grade = None
    
    # Check if there are enough elements in the score_list for Quarterly Grade
    if len(score_list) >= 31:
        quarterly_grade = {"transmuted_grades": score_list[30]}
    else:
        quarterly_grade = None
    
    return {
        "WRITTEN WORKS": written_works_dict,
        "PERFORMANCE TASK": performance_task_dict,
        "QUARTERLY ASSESSMENT": quarterly_assessment,
        "INITIAL GRADE": initial_grade,
        "QUARTERLY GRADE": quarterly_grade
    }



def class_record_details(excel_file, sheet_name):
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[sheet_name]

    teacher_row = None
    school_row = None

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row if cell.value is not None]  # Filter out None values

        
        if "TEACHER:" in row_values:
            grade_section = row_values[2].split('-')
            if len(grade_section) > 1:
                quarter = map_quarter(row_values[0])  # Map the quarter using the map_quarter function
                grade = grade_section[0].strip()
                section = grade_section[1].strip()
                # Map the grade using the map_grade function
                grade = map_grade(grade)
                teacher_row = {
                    "quarters": quarter,
                    "grade": grade,
                    "section": section,
                    "subject": row_values[6]
                }
            else:
                print("Grade section not formatted correctly:", row_values[2])
        
        if "SCHOOL NAME" in row_values:
            school_row = {
                "school_name": row_values[1],
                "school_id": row_values[3],
                "school_year": row_values[5]
            }

        if "REGION" in row_values:
            region_row = {
                "region": row_values[1],
                "division": row_values[3],
                "district": row_values[5]
            }

        if teacher_row and school_row and region_row:
            break


    return teacher_row, school_row, region_row


def divide_hps(row):
    # Divide into sections
 if len(row) >= 27:
        hps_written_works = row[:13]
        hps_performance = row[13:26]
        hps_quarterly = row[26:]

        hps_written_works = [str(value) if value is not None else '' for value in hps_written_works]
        hps_performance = [str(value) if value is not None else '' for value in hps_performance]
        hps_quarterly = [str(value) if value is not None else '' for value in hps_quarterly]

        # Extract values for hps_written_works
        scores_hps_written = hps_written_works[:10]
        total_ww_hps = hps_written_works[10] if len(hps_written_works) > 10 else ''
        percentage_hps_written = hps_written_works[11] if len(hps_written_works) > 11 else ''
        weight_input_written = hps_written_works[12] if len(hps_written_works) > 12 else ''

        # Extract values for hps_performance
        scores_hps_performance = hps_performance[:10]
        total_pt_hps = hps_performance[10] if len(hps_performance) > 10 else ''
        percentage_hps_performance = hps_performance[11] if len(hps_performance) > 11 else ''
        weight_input_performance = hps_performance[12] if len(hps_performance) > 12 else ''

        # Extract values for hps_quarterly
        scores_hps_quarterly = [hps_quarterly[0]] + [''] * 9
        total_qa_hps = hps_quarterly[0] if len(hps_quarterly) > 0 else ''
        percentage_hps_quarterly = hps_quarterly[1] if len(hps_quarterly) > 1 else ''
        weight_input_quarterly = hps_quarterly[2] if len(hps_quarterly) > 2 else ''


        return {
            "written-works": {
                "SCORES": scores_hps_written,
                "TOTAL_HPS": total_ww_hps,
                "percentage_hps_written": percentage_hps_written,
                "WEIGHT": int(float(weight_input_written) * 100)
            },
            "performance-task": {
               "SCORES": scores_hps_performance,
                "TOTAL_HPS": total_pt_hps,
                # "percentage_hps_performance": percentage_hps_performance,
                "WEIGHT": int(float(weight_input_performance) * 100)
            },
            "quarterly-assessment": {
                "SCORES": scores_hps_quarterly,
                 "TOTAL_HPS": total_qa_hps,
                # "percentage_hps_quarterly": percentage_hps_quarterly,
                "WEIGHT": (int(float(weight_input_quarterly) * 100))
            }
        }


def find_highest_possible_scores(excel_file, sheet_name):
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[sheet_name]

    hps_row = None

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        
        # Check if the row contains the header "HIGHEST POSSIBLE SCORE" or "HIGHEST POSSIBLE SCORES"
        if "HIGHEST POSSIBLE SCORE" in row_values:
            # Exclude the first element (index 0) from the row_values list
            row_values = row_values[1:]
            # Process the row to remove leading and trailing None values
            hps_row = process_row(row_values)
            break

    return [hps_row] if hps_row is not None else []

def map_grade(grade):
    grade_map = {
        '1': 'Grade 1',
        'one': 'Grade 1',
        'grade 1': 'Grade 1',
        'GRADE ONE': 'Grade 1',
        '2': 'Grade 2',
        'two': 'Grade 2',
        'grade 2': 'Grade 2',
        'GRADE TWO': 'Grade 2',
        '3': 'Grade 3',
        'three': 'Grade 3',
        'grade 3': 'Grade 3',
        'GRADE THREE': 'Grade 3',
        '4': 'Grade 4',
        'four': 'Grade 4',
        'grade 4': 'Grade 4',
        'GRADE FOUR': 'Grade 4',
        '5': 'Grade 5',
        'five': 'Grade 5',
        'grade 5': 'Grade 5',
        'GRADE FIVE': 'Grade 5',
        '6': 'Grade 6',
        'six': 'Grade 6',
        'grade 6': 'Grade 6',
        'I': 'Grade 1',
        'II': 'Grade 2',
        'III': 'Grade 3',
        'IV': 'Grade 4',
        'V': 'Grade 5',
        'VI': 'Grade 6',

    }
    return grade_map.get(grade.lower(), 'Unknown Grade')

def map_quarter(quarter):
    quarter_map = {
        '1': '1st Quarter',
        'first quarter': '1st Quarter',
        'first': '1st Quarter',
        '1st': '1st Quarter',
        '2': '2nd Quarter',
        'second quarter': '2nd Quarter',
        'second': '2nd Quarter',
        '2nd': '2nd Quarter',
        '3': '3rd Quarter',
        'third quarter': '3rd Quarter',
        'third': '3rd Quarter',
        '3rd': '3rd Quarter',
        '4': '4th Quarter',
        'fourth quarter': '4th Quarter',
        'fourth': '4th Quarter',
        '4th': '4th Quarter',
        '1ST QUARTER': '1st Quarter',
        'FIRST QUARTER': '1st Quarter',
        '2ND QUARTER': '2nd Quarter',
        'SECOND QUARTER': '2nd Quarter',
        '3RD QUARTER': '3rd Quarter',
        'THIRD QUARTER': '3rd Quarter',
        '4TH QUARTER': '4th Quarter',
        'FOURTH QUARTER': '4th Quarter',
    }
    return quarter_map.get(quarter.lower(), 'Unknown Quarter')

def map_data_to_model(json_data, teacher_id, request):
    # Initialize messages list
    messages_list = []
    success = True

    # Extract details from JSON
    teacher_info = json_data['details']['teacher_info']
    students_scores = json_data['students_scores']
    hps_class_record = json_data['hps_class_record']['HIGHEST POSSIBLE SCORE']

    # Create ClassRecord instance
    class_record_instance = ClassRecord.objects.create(
        name=f"{teacher_info['grade']} - {teacher_info['section']} - {teacher_info['subject']} - {teacher_info['quarters']}",
        grade=teacher_info['grade'],
        section=teacher_info['section'],
        subject=teacher_info['subject'],
        quarters=teacher_info['quarters'],
        teacher_id=teacher_id
    )

    # Iterate over student names in JSON data
    for student_name, student_data in students_scores.items():
        # Find the student with the same name and teacher ID
        try:
            student = Student.objects.get(name=student_name)
        except Student.DoesNotExist:
            messages_list.append(('error', f"Student '{student_name}' does not exist in the database for the provided criteria."))
            continue  # Skip this student if not found
        
        total_weighted_score_written = student_data.get('WRITTEN WORKS', {}).get('weighted_score_written', "")
        if total_weighted_score_written and isinstance(total_weighted_score_written, (int, float)):
            total_weighted_score_written *= 100
        # Prepare grade_scores data
        grade_scores_data = {
            "scores_hps": hps_class_record,

            "scores_per_assessment": {  
                "written-works": {
                    "scores": student_data.get('WRITTEN WORKS', {}).get('written_works_scores', []),
                    "total_score": student_data.get('WRITTEN WORKS', {}).get('total_scores_written', ""),
                    "total_percentage_score": student_data.get('WRITTEN WORKS', {}).get('percentage_score_written', "None"),
                    "total_weighted_score": student_data.get('WRITTEN WORKS', {}).get('weighted_score_written', "")
                },
                "performance-task": {
                     "scores": student_data.get('PERFORMANCE TASK', {}).get('performance_task_scores', []),
                    "total_score": student_data.get('PERFORMANCE TASK', {}).get('total_score_performance', "None"),
                    "total_percentage_score": student_data.get('PERFORMANCE TASK', {}).get('percentage_score_performance', ""),
                    "total_weighted_score": student_data.get('PERFORMANCE TASK', {}).get('weighted_score_performance', "")

                },
                "quarterly-assessment": {
                     "scores": student_data.get('QUARTERLY ASSESSMENT', {}).get('quarterly_assessment_scores', []),
                    "total_score": student_data.get('QUARTERLY ASSESSMENT', {}).get('total_score_quarterly', ""),
                     "total_percentage_score": student_data.get('QUARTERLY ASSESSMENT', {}).get('percentage_score_quarterly', ""),
                    "total_weighted_score": student_data.get('QUARTERLY ASSESSMENT', {}).get('weighted_score_quarterly', "")
                }
            },
        }



        # Convert the dictionary to JSON and save it
        GradeScores.objects.create(
            student=student,
            class_record=class_record_instance,
            initial_grades=student_data.get('INITIAL GRADE', {}).get('initial_grades', None),
            transmuted_grades=student_data.get('QUARTERLY GRADE', {}).get('transmuted_grades', None),
            grade_scores=grade_scores_data
        )

    messages_list.append(('success', 'Class record uploaded successfully.'))

    for message_type, message_text in messages_list:
        if message_type == 'success':
            messages.success(request, message_text)
        elif message_type == 'error':
            messages.error(request, message_text)

    return success

@login_required
@require_POST
def class_record_upload(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        sheet_name = request.POST.get('sheet_name')

        user = request.user
        action = f'{user} upload Classrecord "{excel_file}" - "{sheet_name}"'
        details = f'{user} upload Classrecord "{excel_file}" - "{sheet_name}"'
        log_activity(user, action, details)
        # try:
        class_record_data = read_excel_values(excel_file, sheet_name)

            # Process each row in class_record_data
        class_record_data_scores = [process_row(row) for row in class_record_data]


            # Create a dictionary where student's name is the key and the rest of the row values are stored as a list
        class_record_data_scores_with_names = {}
        for row in class_record_data_scores:
                # Check if the student name contains 'FEMALE' or 'MALE'
                if isinstance(row[0], str) and "FEMALE" not in row[0]:
                    student_name = row[0]
                    student_info = {'student_name': student_name}
                    
                    # Find the index of the first non-zero and non-None value
                    index = 1
                    while index < len(row) and (row[index] == 0 or row[index] is None):
                        index += 1

                    # Remove leading occurrences of 0 or None
                    cleaned_row = row[index:] if index < len(row) else []

                    # Update the dictionary with the cleaned row values
                    student_info.update(divide_scores(cleaned_row))
                    class_record_data_scores_with_names[student_name] = student_info

            # Remove the first student from the dictionary
        if class_record_data_scores_with_names:
                del class_record_data_scores_with_names[next(iter(class_record_data_scores_with_names))]

            # Store the result as JSON in the variable students_scores
        students_scores = class_record_data_scores_with_names

        teacher_info, school_info, region_info = class_record_details(excel_file, sheet_name)

    
            
        highest_possible_scores = find_highest_possible_scores(excel_file, sheet_name)
        highest_possible_scores_with_label = {}

        for row in highest_possible_scores:
                hps_label = row.pop(0)
                highest_possible_scores_with_label[hps_label] = divide_hps(row)

        hps_class_record = highest_possible_scores_with_label

            # Store both results in a single dictionary
        extracted_class_record = {
                "details": {
                    "teacher_info": teacher_info,
                    "school_info": school_info,
                    "region_info": region_info
                },
                "students_scores": students_scores,
                "hps_class_record": hps_class_record
        }

            
            # Save extracted_class_record as a JSON file
        json_filename = 'result.json'
        json_path = os.path.join(settings.MEDIA_ROOT, json_filename)

            # Ensure the directory exists, if not create it
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

        with open(json_path, 'w') as json_file:
                json.dump(extracted_class_record, json_file, indent=4)  
            
            # Map the loaded JSON data to the model
        success = map_data_to_model(extracted_class_record, request.user.teacher.id, request)  # Passing request object

        if success:
                return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html')
        else:
                # If there's an error, return the same template with an error flag
                return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html', {'error': True})
        # except Exception as e:
        #     # Add error message
        #     messages.error(request, f'Error processing file: {e}')

        return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html')
    else:
        # Return a response in case of non-POST requests
        return HttpResponse("Only POST requests are allowed.")



def teacher_upload_documents_ocr(request):
    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            
            uploaded_file = request.FILES['document']
            name = uploaded_file.name
            teacher = request.user.teacher
            
        
            # Sanitize the filename by replacing spaces and special characters with underscores
            filename = 'processed_documents/' + name.replace(' ', '_').replace(',', '').replace('(', '').replace(')', '')
            file_extension = os.path.splitext(filename)[-1].lower()
            print(filename)
            
            
            if ProcessedDocument.objects.filter(document=filename).exists():
                messages.error(request, 'Document with the same name already exists.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            # Replace 'YOUR_PROJECT_ID' with your Google Cloud project ID.

            user = request.user
            action = f'{user} upload SF10 "{name}"'
            details = f'{user} upload SF10 "{name}" in the system.'
            log_activity(user, action, details)

            logs = user, action, details    
            print(logs)

            project_id = '1083879771832'


            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = r"ces-ocr-5a2441a9fd54.json"

            client = documentai.DocumentProcessorServiceClient()

            # Define the processor resource name.
            processor_name = f"projects/{project_id}/locations/us/processors/84dec1544028cc60"

            

            # Read the document content from the uploaded file.
            content = uploaded_file.read()

            # Determine the MIME type based on the file extension.
            file_extension = os.path.splitext(uploaded_file.name)[-1].lower()
            if file_extension in ['.pdf']:
                mime_type = "application/pdf"
            elif file_extension in ['.jpg', '.jpeg']:
                mime_type = "image/jpeg"
            else:
                # Handle unsupported file types or provide an error message.
                return render(request, 'unsupported_file_type.html')

            # Configure the processing request.
            processing_request = {
                "name": processor_name,
                "document": {"content": content, "mime_type": mime_type},
            }

            # Process the document.
            response = client.process_document(request=processing_request)

            # Access the extracted text from the document content.
            document = response.document
            text = document.text

            data_by_type = {
                'Type': [],
                'Raw Value': [],
                'Normalized Value': [],
                'Confidence': [],
            }

            # Iterate through your data extraction process and populate the dictionary
            for entity in document.entities:
                data_by_type['Type'].append(entity.type_)
                data_by_type['Raw Value'].append(entity.mention_text)
                data_by_type['Normalized Value'].append(entity.normalized_value.text)
                data_by_type['Confidence'].append(f"{entity.confidence:.0%}")

                # Get Properties (Sub-Entities) with confidence scores
                for prop in entity.properties:
                    data_by_type['Type'].append(prop.type_)
                    data_by_type['Raw Value'].append(prop.mention_text)
                    data_by_type['Normalized Value'].append(prop.normalized_value.text)
                    data_by_type['Confidence'].append(f"{prop.confidence:.0%}")

            print(data_by_type)

            # Create a ProcessedDocument instance and save it
            processed_document = ProcessedDocument(document=uploaded_file, upload_date=timezone.now(), teacher=teacher)
            processed_document.save()

            my_data = ExtractedData(processed_document=processed_document)

            # Define a mapping of keys from data_by_type to ExtractedData fields
            key_mapping = {
                'Last_Name': 'last_name',
                'First_Name': 'first_name',
                'Middle_Name': 'middle_name',
                'SEX': 'sex',
                'Classified_as_Grade': 'classified_as_grade',
                'LRN': 'lrn',
                'Name_of_School': 'name_of_school',
                'School_Year': 'school_year',
                'General_Average': 'general_average',
                'Birthdate': 'birthdate',
            }

            last_values = {}

            for i in range(len(data_by_type['Type'])):
                data_type = data_by_type['Type'][i]
                raw_value = data_by_type['Raw Value'][i]

                # Update the last value for the type
                last_values[data_type] = {'value': raw_value}

            # Set the last values to the corresponding fields in my_data
            for key, field_name in key_mapping.items():
                if key in last_values:
                    setattr(my_data, field_name, last_values[key]['value'])

            # # Handle birthdate separately
            #     if 'Birthdate' in key_mapping:
            #         birthdate_index = data_by_type['Type'].index('Birthdate') if 'Birthdate' in data_by_type['Type'] else None
            #         if birthdate_index is not None:
            #             birthdate_str = data_by_type['Raw Value'][birthdate_index]
            #             try:
            #                 # Provide a specific format string based on the expected format
            #                 my_data.birthdate = parser.parse(birthdate_str).date()
            #             except ValueError as e:
            #                 print(f"Error parsing birthdate: {e}")
            if 'Birthdate' in key_mapping:
                birthdate_index = data_by_type['Type'].index('Birthdate') if 'Birthdate' in data_by_type['Type'] else None
                if birthdate_index is not None:
                    birthdate_str = data_by_type['Raw Value'][birthdate_index]
                    try:
                        # Provide a specific format string based on the expected format
                        my_data.birthdate = parser.parse(birthdate_str).date()
                    except ValueError as e:
                        print(f"Error parsing birthdate: {e}")

            my_data.save()

            # response = FileResponse(open(processed_document.document.path, 'rb'), content_type='application/pdf')
            # response['Content-Disposition'] = f'inline; filename="{uploaded_file.name}"'
            # return response

            pdf_content_base64 = base64.b64encode(content).decode('utf-8')

        return redirect('teacher_sf10_views')
        # return render(request, 'admin_template/edit_extracted_data.html', {
        #         # 'extracted_data': extracted_data_for_review,
        #         'document_text': text,
        #         'uploaded_document_url': processed_document.document.url,
        #         # 'all_extracted_data': all_extracted_data,
        #         'processed_document': processed_document,
        #         'download_link': processed_document.document.url,
        #         'data_by_type': data_by_type,
        #         # 'extracted_text': extracted_text 
        #         'extracted_data': my_data,
        #         'pdf_content_base64': pdf_content_base64, 
        #     })
    else: 
        form = DocumentUploadForm()

    return render(request, 'teacher_template/adviserTeacher/teacher_upload_documents.html', {'form': form})


def teacher_sf10_views(request):
    # Retrieve the search query from the request's GET parameters
    search_query = request.GET.get('search', '')
    teacher = request.user.teacher
    print(teacher)
    # If a search query is present, filter the ExtractedData model
    if search_query:
        # You can customize the fields you want to search on
        search_fields = ['last_name', 'first_name', 'middle_name', 'lrn', 'name_of_school', 'sex', 'birthdate', 'school_year', 'classified_as_grade', 'general_average','processed_document__teacher__user__first_name', 'processed_document__teacher__user__last_name']
        
        # Use Q objects to create a complex OR query
        query = Q()
        for field in search_fields:
            query |= Q(**{f'{field}__icontains': search_query})

        # Filter the ExtractedData model based on the search query
        all_extracted_data = ExtractedData.objects.filter(query)
    else:
        # If no search query, retrieve all records
        all_extracted_data = ExtractedData.objects.filter(processed_document__teacher=teacher)

    # Pass the filtered data and search query to the template context
    context = {
        'all_extracted_data': all_extracted_data,
        'search_query': search_query,
        }

        # Render the sf10.html template with the context data
    return render(request, 'teacher_template/adviserTeacher/teacher_sf10.html', context)


def teacher_batch_process_documents(request):

    if request.method == 'POST':
        form = DocumentBatchUploadForm(request.POST, request.FILES)

    
        if form.is_valid():
            uploaded_files = request.FILES.getlist('documents')
            teacher = request.user.teacher
            for uploaded_file in uploaded_files:
                
                
                name = uploaded_file.name
                filename = 'processed_documents/' + name.replace(' ', '_').replace(',', '').replace('(', '').replace(')', '')

                if ProcessedDocument.objects.filter(document=filename).exists():
                    messages.error(request, 'Document with the same name already exists.')
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


                user = request.user
                action = f'{user} upload SF10 "{name}"'
                details = f'{user} upload SF10 "{name}" in the system.'
                log_activity(user, action, details)

                logs = user, action, details    
                print(logs)

                project_id = '1083879771832'
                os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = r"ces-ocr-5a2441a9fd54.json"
                client = documentai.DocumentProcessorServiceClient()
                processor_name = f"projects/{project_id}/locations/us/processors/84dec1544028cc60"

                content = uploaded_file.read()
                file_extension = os.path.splitext(uploaded_file.name)[-1].lower()
                if file_extension in ['.pdf']:
                    mime_type = "application/pdf"
                elif file_extension in ['.jpg', '.jpeg']:
                    mime_type = "image/jpeg"
                else:
                    messages.error(request, f'Unsupported file type for document "{name}".')
                    continue

                processing_request = {
                    "name": processor_name,
                    "document": {"content": content, "mime_type": mime_type},
                }

                try:
                    response = client.process_document(request=processing_request)
                    document = response.document
                    text = document.text
                    data_by_type = {
                        'Type': [],
                        'Raw Value': [],
                        'Normalized Value': [],
                        'Confidence': [],
                    }

                    # Iterate through your data extraction process and populate the dictionary
                    for entity in document.entities:
                        data_by_type['Type'].append(entity.type_)
                        data_by_type['Raw Value'].append(entity.mention_text)
                        data_by_type['Normalized Value'].append(entity.normalized_value.text)
                        data_by_type['Confidence'].append(f"{entity.confidence:.0%}")

                        # Get Properties (Sub-Entities) with confidence scores
                        for prop in entity.properties:
                            data_by_type['Type'].append(prop.type_)
                            data_by_type['Raw Value'].append(prop.mention_text)
                            data_by_type['Normalized Value'].append(prop.normalized_value.text)
                            data_by_type['Confidence'].append(f"{prop.confidence:.0%}")

                    print(data_by_type)

                    # Create a ProcessedDocument instance and save it
                    processed_document = ProcessedDocument(document=uploaded_file, upload_date=timezone.now(), teacher=teacher)
                    processed_document.save()

                    my_data = ExtractedData(processed_document=processed_document)

                    # Define a mapping of keys from data_by_type to ExtractedData fields
                    key_mapping = {
                        'Last_Name': 'last_name',
                        'First_Name': 'first_name',
                        'Middle_Name': 'middle_name',
                        'SEX': 'sex',
                        'Classified_as_Grade': 'classified_as_grade',
                        'LRN': 'lrn',
                        'Name_of_School': 'name_of_school',
                        'School_Year': 'school_year',
                        'General_Average': 'general_average',
                        'Birthdate': 'birthdate',
                    }

                    last_values = {}

                    for i in range(len(data_by_type['Type'])):
                        data_type = data_by_type['Type'][i]
                        raw_value = data_by_type['Raw Value'][i]

                        # Update the last value for the type
                        last_values[data_type] = {'value': raw_value}

                    # Set the last values to the corresponding fields in my_data
                    for key, field_name in key_mapping.items():
                        if key in last_values:
                            setattr(my_data, field_name, last_values[key]['value'])

                    # # Handle birthdate separately
                    #     if 'Birthdate' in key_mapping:
                    #         birthdate_index = data_by_type['Type'].index('Birthdate') if 'Birthdate' in data_by_type['Type'] else None
                    #         if birthdate_index is not None:
                    #             birthdate_str = data_by_type['Raw Value'][birthdate_index]
                    #             try:
                    #                 # Provide a specific format string based on the expected format
                    #                 my_data.birthdate = parser.parse(birthdate_str).date()
                    #             except ValueError as e:
                    #                 print(f"Error parsing birthdate: {e}")
                    if 'Birthdate' in key_mapping:
                        birthdate_index = data_by_type['Type'].index('Birthdate') if 'Birthdate' in data_by_type['Type'] else None
                        if birthdate_index is not None:
                            birthdate_str = data_by_type['Raw Value'][birthdate_index]
                            try:
                                # Provide a specific format string based on the expected format
                                my_data.birthdate = parser.parse(birthdate_str).date()
                            except ValueError as e:
                                print(f"Error parsing birthdate: {e}")

                    my_data.save()


                except Exception as e:
                    messages.error(request, f'Error processing document "{name}": {str(e)}')

            messages.success(request, 'Documents processed successfully.')
            return redirect('teacher_sf10_views')
            # return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    else:
        form = DocumentBatchUploadForm()

    return render(request, 'teacher_template/adviserTeacher/teacher_batch_process_documents.html', {'form': form})

def teacher_sf10_edit_view(request, id):
    extracted_data = get_object_or_404(ExtractedData, id=id)

    # Assuming you have 'processed_document' field in your ExtractedData model
    processed_document = extracted_data.processed_document

    # Access the PDF content from the 'document' field of the 'ProcessedDocument' object
    pdf_content = processed_document.document.read()

    # Convert the content to base64 encoding
    pdf_content_base64 = base64.b64encode(pdf_content).decode('utf-8')

    return render(request, 'teacher_template/adviserTeacher/teacher_edit_sf10.html', {'extracted_data': extracted_data, 'pdf_content_base64': pdf_content_base64})

def teacher_sf10_edit(request, id):

    extracted_data = get_object_or_404(ExtractedData, id=id)

    if request.method == 'POST':
        # Assuming form data is sent via POST request
        # Retrieve and process the form data for editing
        extracted_data.last_name = request.POST.get('Last_Name', '')
        extracted_data.first_name = request.POST.get('First_Name', '')
        extracted_data.middle_name = request.POST.get('Middle_Name', '')
        extracted_data.sex = request.POST.get('SEX', '')
        extracted_data.classified_as_grade = request.POST.get('Classified_as_Grade', '')
        extracted_data.lrn = request.POST.get('LRN', '')
        extracted_data.name_of_school = request.POST.get('Name_of_School', '')
        extracted_data.school_year = request.POST.get('School_Year', '')
        extracted_data.general_average = request.POST.get('General_Average', '')

        sf10_name = f"{extracted_data.first_name} {extracted_data.last_name}"
        user = request.user
        action = f'{user} updates information of the SF10 of "{sf10_name}"'
        details = f'{user} updates information of the SF10 of "{sf10_name} in the system.'
        log_activity(user, action, details)

        logs = user, action, details    
        print(logs)
  
        # Handle birthdate format conversion
        birthdate_str = request.POST.get('Birthdate', '')
        # Attempt to create the announcement
        try:
            birthdate_obj = datetime.strptime(birthdate_str, "%b. %d, %Y")
            extracted_data.birthdate = birthdate_obj.strftime("%Y-%m-%d")
        except ValueError:
            # Handle invalid birthdate format
            pass  # You may want to add proper error handling here

        # Save the changes to the ExtractedData instance
        extracted_data.save()


        # Redirect to a success page or any other appropriate URL
        return HttpResponseRedirect(reverse('teacher_sf10_views') + '?success=true')

    # Render the edit_sf10.html template with the ExtractedData instance
    return render(request, 'teacher_template/adviserTeacher/teacher_edit_sf10.html', {'extracted_data': extracted_data})


def create_core_values(request):
    if request.method == 'POST':
        form = CoreValuesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create_core_values')
    else:
        form = CoreValuesForm()
    return render(request, 'teacher_template/adviserTeacher/create_core_values.html', {'form': form})

def create_behavior_statements(request):
    if request.method == 'POST':
        form = BehaviorStatementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create_behavior_statements')
    else:
        form = BehaviorStatementForm()
    return render(request, 'teacher_template/adviserTeacher/create_behavior_statements.html', {'form': form})


# def create_learners_observation(request, grade, section):
#     behavior_statements = BehaviorStatement.objects.all()
#     students = Student.objects.filter(grade=grade, section=section)
#     core_values = CoreValues.objects.all()
#     quarters = ['1st Quarter', '2nd Quarter', '3rd Quarter', '4th Quarter'] 

#     if request.method == 'POST':
#         for student in students:
#             for quarter in range(1, 5):
#                 for behavior_statement in behavior_statements:
#                     field_name = f'student_{student.id}_quarter_{quarter}_behavior_{behavior_statement.id}'
#                     marking = request.POST.get(field_name)
#                     # Process the marking data and save to the database as needed

#     # Create a dictionary to hold marking data for each student for each quarter
#     student_markings = {}
#     for student in students:
#         student_markings[student] = {}
#         for quarter in quarters:
#             student_markings[student][quarter] = {}
#             for behavior_statement in behavior_statements:
#                 # Initialize marking to None
#                 student_markings[student][quarter][behavior_statement] = None

#     print(behavior_statements)
#     # print(student_markings)
#     context = {
#         'students': students,
#         'behavior_statements': behavior_statements,
#         'quarters': quarters,
#         'student_markings': student_markings,
#         'core_values': core_values  # Pass the student markings to the template
#     }
#     # Render the form template with the behavior statements and students
#     return render(request, 'teacher_template/adviserTeacher/create_learners_observation.html', context)


def students_behavior_view(request, grade, section):
    students = Student.objects.filter(grade=grade, section=section)
    core_values = CoreValues.objects.all()
    behavior_statements = BehaviorStatement.objects.all()
    quarters = ['Quarter 1', 'Quarter 2', 'Quarter 3', 'Quarter 4'] 

    core_values_length = len(core_values)
    behavior_statements_length = len(behavior_statements)
    rowspan = core_values_length * behavior_statements_length

    context = {
        'grade': grade,
        'section': section,
        'students': students, 
        'behavior_statements': behavior_statements,
        'core_values' : core_values,
        'quarters': quarters,
        'rowspan': rowspan,
    }
    return render(request, 'teacher_template/adviserTeacher/students_behavior.html', context)

def save_observations(request):
    if request.method == 'POST':
        try:
            observations = json.loads(request.POST.get('observations'))

            # Collect existing observations for each quarter
            existing_quarters = {}

            for observation in observations:
                student_id = observation.get('student_id')
                quarter_data = observation.get('quarter')

                quarter_field_map = {
                    '1st Quarter': 'quarter_1',
                    '2nd Quarter': 'quarter_2',
                    '3rd Quarter': 'quarter_3',
                    '4th Quarter': 'quarter_4'
                }

                quarter_field = quarter_field_map.get(quarter_data)

                if not quarter_field:
                    raise ValueError('Invalid quarter data')

                learner_observation, _ = LearnersObservation.objects.get_or_create(student_id=student_id)
                current_observations = getattr(learner_observation, quarter_field)

                if current_observations:
                    existing_quarters[(student_id, quarter_field)] = quarter_data

            for observation in observations:
                student_id = observation.get('student_id')
                quarter_data = observation.get('quarter')
                core_value = observation.get('core_value')
                behavior_statement = observation.get('behavior_statement')
                marking = observation.get('marking')

                quarter_field_map = {
                    '1st Quarter': 'quarter_1',
                    '2nd Quarter': 'quarter_2',
                    '3rd Quarter': 'quarter_3',
                    '4th Quarter': 'quarter_4'
                }

                quarter_field = quarter_field_map.get(quarter_data)

                if not quarter_field:
                    raise ValueError('Invalid quarter data')

                # Check if the quarter field already has data
                if (student_id, quarter_field) in existing_quarters:
                    raise ValueError(f'Observations for {quarter_data} already exist')

                learner_observation, _ = LearnersObservation.objects.get_or_create(student_id=student_id)

                current_observations = getattr(learner_observation, quarter_field)
                if current_observations is None:
                    current_observations = {}

                if core_value not in current_observations:
                    current_observations[core_value] = []

                current_observations[core_value].append({
                    'behavior_statement': behavior_statement,
                    'marking': marking
                })

                setattr(learner_observation, quarter_field, current_observations)
                learner_observation.save()

            return JsonResponse({'message': 'Observations saved successfully.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=405)


    
def display_learners_observation(request, grade, section):
    learners_observation = LearnersObservation.objects.filter(student__grade=grade, student__section=section)
    quarters = ['quarter_1', 'quarter_2', 'quarter_3', 'quarter_4']
    context = {
        'learners_observation': learners_observation,
        'quarters': quarters,
         'grade': grade,
        'section': section,
        
    }
    return render(request, 'teacher_template/adviserTeacher/display_learners_observation.html', context)


def update_markings(request):
    if request.method == 'POST':
        observation_id = request.POST.get('observation_id')
        quarter = request.POST.get('quarter')
        core_value = request.POST.get('core_value')
        behavior = request.POST.get('behavior')
        marking = request.POST.get('marking')

        # Update marking for the observation
        # Example:
        observation = LearnersObservation.objects.get(pk=observation_id)
        quarter_data = getattr(observation, quarter)
        for core_val, behaviors in quarter_data.items():
            if core_val == core_value:
                for item in behaviors:
                    if item['behavior_statement'] == behavior:
                        item['marking'] = marking
                        break

        # Save the updated data
        setattr(observation, quarter, quarter_data)
        observation.save()

        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error'}, status=400)


@login_required
@require_POST
def sf2_upload(request):
    if request.method == 'POST' and 'pdf_file' in request.FILES:
        pdf_file = request.FILES['pdf_file']
        content = pdf_file.read()



        # Set up Google Cloud Document AI client
        project_id = "1083879771832"
        location = "us"  # Format is "us" or "eu"
        processor_id = "827ebb48ef18ecd"  # Create processor before running sample
        processor_version = "pretrained-form-parser-v2.0-2022-11-10"  # Refer to https://cloud.google.com/document-ai/docs/manage-processor-versions for more information


        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = r"ces-ocr-5a2441a9fd54.json"
            # Process the document and extract tables using Document AI
        document = process_document_form_sample(project_id, location, processor_id, processor_version, content, "application/pdf")


            # Extract table data
        table_data = []


        for page in document.pages:
                for table in page.tables:
                    # Extract text content of header and body rows
                    for row in table.header_rows:
                        row_content = [layout_to_text(cell.layout, document.text) for cell in row.cells]
                        row_hps = "HIGHEST POSSIBLE SCORE"
                        table_data.append(row_content)


                    for row in table.body_rows:
                        row_content = [layout_to_text(cell.layout, document.text) for cell in row.cells]
                        table_data.append(row_content)
        # # print(table_data)
        # sf2_data_scores = [sf2_process_row(row) for row in table_data]
        # # print(f"sf2_Data_scores {sf2_data_scores}")


        # sf2_names = {}


        # for row in sf2_data_scores:
        #     if len(row) >= 2 and isinstance(row[1], str):  # Check if the row has at least two elements and the element at index 1 is a string
        #         student_name = row[1]
        #         student_info = {'student_name': student_name}
               
        #         index = 1
        #         while index < len(row) and (row[index] == 0 or row[index] is None):
        #             index += 1
        #         # You can add additional processing of the row here if needed
        #         cleaned_row = row[index:] if index < len(row) else []
        #         sf2_names[student_name] = student_info
        #         student_info.update(sf2_scores_pdf(cleaned_row))


        # school_info = {}
        # grade_info = None


        # for row in table_data:
        #     # print("Row values:", row)


        #     for element in row:
        #         # Check if the element starts with the desired string
        #         if element.startswith("Report for the Month of"):
        #             month_row = element.strip().split("\n")[-1]
        #             month = month_row.split()[-1]
        #             school_info['month'] = month
        #             # print("Month:", month)


        #             break  # Exit t


        #         # if element.startswith("Month"):
        #         #     month_part = element.split(':')[1].strip()
        #         #     month = month_part.split()[0]  # Extract the first word, which should be the month name
        #         #     school_info['month'] = month
        #         #     # print("Month:", month)  # Print "Month:" along with the month value


        #             break


        #     if "Name of School" in row:
        #         grade_info = {
        #             "school_name": row[1],
        #             "grade": row[3],
        #             "section": row[5]
        #         }
        #     if school_info and grade_info:
        #         break


        # # print(school_info)


        # # print(table_data)
        # # print(f"sf2_name {sf2_names}")


        # if sf2_names:
        #     keys_to_remove = list(sf2_names.keys())[1:3]  # Get keys for the next two items
        #     for key in keys_to_remove:
        #         del sf2_names[key]


        # students_scores = sf2_names

        # # print(f"students_score: {students_scores}")


        # extracted_sf2 = {
        #         "details": {
        #             "school_info": school_info,
        #         },
        #         "students_scores": students_scores,
        # }


        # json_filename = 'result_sf3.json'
        # json_path = os.path.join(settings.MEDIA_ROOT, json_filename)


        #     # Ensure the directory exists, if not create it
        # os.makedirs(settings.MEDIA_ROOT, exist_ok=True)


        # with open(json_path, 'w') as json_file:
        #         json.dump(extracted_sf2, json_file, indent=4)  
           
        #     # Map the loaded JSON data to the model
        # success = sf2_map_data_to_model_pdf(extracted_sf2, request.user.teacher.id, request)  # Passing request object
        # print(success)
        # form_fields_data = []


        # for page in document.pages:
        #         for field in page.form_fields:
        #             name = layout_to_text(field.field_name, document.text).strip()
        #             value = layout_to_text(field.field_value, document.text).strip()
        #             form_field_data = {'name': name, 'value': value}
        #             form_fields_data.append(form_field_data)


        # json_data = {'table_data': table_data, 'form_fields_data': form_fields_data}
        json_data = {'table_data': table_data}
        json_file_path = 'document_data.json'
        with open(json_file_path, 'w') as json_file:
                json.dump(json_data, json_file)


        print(f'Table data saved to {json_file_path}')


        return render(request, 'admin_template/table_data.html', {'table_data': table_data})
# return render(request, 'admin_template/table_data.html', {'table_data': table_data, 'form_fields_data': form_fields_data})

    elif request.method == 'POST':
        excel_file = request.FILES['excel_file']
        sheet_name = request.POST.get('sheet_name')


        sf2_data = sf2_read_excel_values(excel_file, sheet_name)

        sf2_data_scores = [sf2_process_row(row) for row in sf2_data]

        sf2_names = {}

        # print(sf2_data_scores)

        for row in sf2_data_scores:
            if isinstance(row[1], str):  # Check if the element at index 1 is a string
                student_name = row[1]
                student_info = {'student_name': student_name}
                
                index = 1
                while index < len(row) and (row[index] == 0 or row[index] is None):
                        index += 1
                # You can add additional processing of the row here if needed
                cleaned_row = row[index:] if index < len(row) else []
                sf2_names[student_name] = student_info
                student_info.update(sf2_scores(cleaned_row))

        if sf2_names:
            del sf2_names[next(iter(sf2_names))]

        students_scores = sf2_names

        school_info, grade_info = sf2_class_record_details(excel_file, sheet_name)
        
        # print(f"sf2_Data: {sf2_data}")
        # print(f"sf2_data_scores {sf2_data_scores}")
        # print(sf2_names)
        
        # print(school_info, grade_info)

        extracted_sf2 = {
                "details": {
                    "grade_info": grade_info,
                    "school_info": school_info,
                },
                "students_scores": students_scores,
        }

        print(extracted_sf2)
        json_filename = 'result_sf2.json'
        json_path = os.path.join(settings.MEDIA_ROOT, json_filename)

            # Ensure the directory exists, if not create it
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

        with open(json_path, 'w') as json_file:
                json.dump(extracted_sf2, json_file, indent=4)  
            
            # Map the loaded JSON data to the model
        success = sf2_map_data_to_model(extracted_sf2, request.user.teacher.id, request)  # Passing request object

        if success:
                return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html')
        else:
                # If there's an error, return the same template with an error flag
                return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html', {'error': True})
        # except Exception as e:
        #     # Add error message
        #     messages.error(request, f'Error processing file: {e}')

        return render(request, 'teacher_template/adviserTeacher/class_record_from_excel.html')
    else:
        # Return a response in case of non-POST requests
        return HttpResponse("Only POST requests are allowed.")

def process_document_form_sample(
    project_id: str,
    location: str,
    processor_id: str,
    processor_version: str,
    content: bytes,
    mime_type: str,
) -> documentai.Document:
    # Set up Google Cloud Document AI client
    client_options = ClientOptions(api_endpoint=f"{location}-documentai.googleapis.com")
    client = documentai.DocumentProcessorServiceClient(client_options=client_options)

    # The full resource name of the processor version
    name = client.processor_version_path(project_id, location, processor_id, processor_version)

    # Configure the process request
    request = documentai.ProcessRequest(
        name=name,
        raw_document=documentai.RawDocument(content=content, mime_type=mime_type),
    )

    # Process the document and extract tables using Document AI
    result = client.process_document(request=request)

    # Return the processed document
    return result.document

def layout_to_text(layout, document_text):
    """
    Extracts text content from the layout of a Document AI element.
    """
    text_content = ""
    for text_segment in layout.text_anchor.text_segments:
        start_index = text_segment.start_index
        end_index = text_segment.end_index
        text_content += document_text[start_index:end_index]
    return text_content

def sf2_read_excel_values(excel_file, sheet_name):
    # Loading the workbook
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[sheet_name]

    data = []
    start_reading = False

    for row in sheet.iter_rows():
        row_values = []
        for cell in row[1:]:
            cell_value = cell.value
            row_values.append(cell_value)
        
        if any(cell_value is not None and cell_value != 0 for cell_value in row_values):
            start_reading = True
            
            if start_reading:
                data.append(row_values)

    return data

def sf2_process_row(row):
    # Find the index of the first non-None element after index 0
    index = 1
    while index < len(row) and row[index] is None:
        index += 1

    # # Remove the three None values next to index 0
    # row = [row[0]] + row[index:]

    # Find the index of the last non-None element
    last_index = len(row) - 1
    while last_index >= 0 and row[last_index] is None:
        last_index -= 1

    # Remove all None values from the end of the list until reaching a non-None value
    row = row[:last_index + 1]

    return row

def sf2_scores(score_list):
    # Check if there are enough elements in the score_list for Absent
    if len(score_list) >= 37:
        absent = {"absent": score_list[36]}
    else:
        absent = None
    
    # Check if there are enough elements in the score_list for Present
    if len(score_list) >= 39:
        present = {"present": score_list[38]}
    else:
        present = None
    
    return {
        "ABSENT": absent,
        "PRESENT": present
    }

def sf2_scores_pdf(score_list):
    student_absent = None
    student_present = None


        # Check if there are enough elements in the student_info list for Absent
    if len(score_list) >= 3:  # Adjust the index as per your data structure
        if score_list[-1] == '':
            student_absent = {"absent": score_list[-3]}
        else:
            student_absent = {"absent": score_list[-2]} # Adjust the index as per your data structure


        # Check if there are enough elements in the student_info list for Present
    if len(score_list) >= 2:
            if score_list[-1] == "":
                student_present = {"present": score_list[-2]}
            else:  # Adjust the index as per your data structure
                student_present = {"present": score_list[-1]}  # Adjust the index as per your data structure


    return {
            # Assuming student name is at index 1
            "ABSENT": student_absent,
            "PRESENT": student_present
        }


def sf2_class_record_details(excel_file, sheet_name):
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[sheet_name]

    school_info = None
    grade_info = None

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row if cell.value is not None]  # Filter out None values
        print("Row values:", row_values)  # Debug print statement

        if "Report for the Month of" in row_values:
            school_info = {
                    "school_id": row_values[1],
                    "school_year": row_values[3],
                    "month": row_values[5],
                }
        
        if "Name of School" in row_values:
            grade_info = {
                "school_name": row_values[1],
                "grade": row_values[3],
                "section": row_values[5]
            }

        if school_info and grade_info:
            break

    return school_info, grade_info



@transaction.atomic
def sf2_map_data_to_model(json_data, teacher_id, request):
    # Initialize messages list
    messages_list = []
    success = True

    # Extract details from JSON
    details = json_data['details']
    students_scores = json_data['students_scores']


        # Extract grade and school info
    grade_info = details['grade_info']
    school_info = details['school_info']

        # Extract month from school info
    month = school_info.get('month', 'UNKNOWN')
    response_data = {'message': 'Attendance records saved successfully'}
    error_response_data = {'error': f'Attendance records for {month} already exist'}
    total_school_days = 0
    total_days_present = 0
    total_days_absent = 0

   

    if AttendanceRecord.objects.filter(attendance_record__has_key=month).exists():
            return JsonResponse(error_response_data, status=400)
    
        # Iterate over student names in JSON data
    for student_name, student_data in students_scores.items():
            school_days = student_data.get('ABSENT', {}).get('absent', 0) + student_data.get('PRESENT', {}).get('present', 0)
            days_present = student_data.get('PRESENT', {}).get('present', 0)
            days_absent = student_data.get('ABSENT', {}).get('absent', 0)

            total_school_days += school_days
            total_days_present += days_present
            total_days_absent += days_absent
            # Find the student with the same name
            try:
                student = Student.objects.get(name=student_name)
            except Student.DoesNotExist:
                # messages_list.append(('error', f"Student '{student_name}' does not exist in the database."))
                continue  # Skip this student if not found

            # Prepare attendance record data
            # attendance_record = {
            #     month: {
            #         "No. of School Days": student_data.get('ABSENT', {}).get('absent', 0) + student_data.get('PRESENT', {}).get('present', 0),
            #         "No. of Days Present": student_data.get('PRESENT', {}).get('present', 0),
            #         "No. of Days Absent": student_data.get('ABSENT', {}).get('absent', 0)
            #     }
            # }

            # # Update or create AttendanceRecord for the student
            # record, created = AttendanceRecord.objects.update_or_create(
            #     student=student,
            #     defaults={'attendance_record': attendance_record}
            # )

            # Check if an attendance record exists for the student and month
            try:
                attendance_record = AttendanceRecord.objects.get(
                    student=student,
                    attendance_record__has_key=month  # Filter records based on the key
                )

                # Update the existing attendance record with new data
                existing_data = attendance_record.attendance_record.get(month, {})
                existing_data['No. of School Days'] = school_days
                existing_data['No. of Days Present'] = days_present
                existing_data['No. of Days Absent'] = days_absent
                attendance_record.attendance_record[month] = existing_data
                attendance_record.save()

            except AttendanceRecord.DoesNotExist:
                # If the AttendanceRecord does not exist for the student and month,
                # create a new record with the given data and month
                # Check if the student has any existing records, if not, create a new one
                if not AttendanceRecord.objects.filter(student=student).exists():
                    AttendanceRecord.objects.create(
                        student=student,
                        attendance_record={
                            month: {
                                'No. of School Days': school_days,
                                'No. of Days Present': days_present,
                                'No. of Days Absent': days_absent
                            }
                        }
                    )
                else:
                    # Retrieve all existing records for the student
                    existing_records = AttendanceRecord.objects.filter(student=student)
                    # Update all existing records by adding the new month and its data
                    for record in existing_records:
                        record.attendance_record[month] = {
                            'No. of School Days': school_days,
                            'No. of Days Present': days_present,
                            'No. of Days Absent': days_absent
                        }
                        record.save()

            total_attendance_record, created = AttendanceRecord.objects.get_or_create(
                student=student,  # For total attendance record
                defaults={
                    'attendance_record': {
                        'Total': {
                            'Total School Days': school_days,
                            'Total Days Present': days_present,
                            'Total Days Absent': days_absent
                        }
                    }
                }
            )

            if not created:
                # If the record already exists, update the existing data
                total_record_data = total_attendance_record.attendance_record.get('TOTAL', {})
                # Ensure that the 'Total' key exists and initialize it with default values if not
                if not total_record_data:
                    total_record_data = {
                        'Total School Days': 0,
                        'Total Days Present': 0,
                        'Total Days Absent': 0
                    }
                total_record_data['Total School Days'] += school_days
                total_record_data['Total Days Present'] += days_present
                total_record_data['Total Days Absent'] += days_absent
                total_attendance_record.attendance_record['TOTAL'] = total_record_data

                total_attendance_record.attendance_record = {
                    k: total_attendance_record.attendance_record[k] for k in sorted(total_attendance_record.attendance_record.keys()) if k != 'TOTAL'
                }
                total_attendance_record.attendance_record['TOTAL'] = total_record_data
                total_attendance_record.save()

            # Check if the record was created or updated

    if created:
            messages_list.append(('success', f"Total attendance record created successfully."))
    else:
            messages_list.append(('success', f"Total attendance record updated successfully."))

    for message_type, message_text in messages_list:
        if message_type == 'success':
            messages.success(request, message_text)
        elif message_type == 'error':
            messages.error(request, message_text)

    return success

def delete_grade_data_subject(request, grade, section, subject):
    try:
        advisory_classes = AdvisoryClass.objects.filter(grade=grade, section=section)
    except AdvisoryClass.DoesNotExist:
        return JsonResponse({'error': 'AdvisoryClass not found'}, status=404)

    deleted = False  # Flag to check if any deletion occurred

    # Check if the provided subject is present in grades_data for each advisory class
    for advisory_class in advisory_classes:
        if subject in advisory_class.grades_data:
            # Delete the key associated with the subject from grades_data
            del advisory_class.grades_data[subject]
            # Save the changes
            advisory_class.save()
            deleted = True

    if deleted:
        return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        return JsonResponse({'error': f'{subject} not found in grades_data for any matching AdvisoryClass objects'}, status=400)
